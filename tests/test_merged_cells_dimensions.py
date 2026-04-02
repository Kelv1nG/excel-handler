"""Tests for merged cells of various dimensions during table writes."""
import pytest
import polars as pl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

from excel.template_writer import ExcelTemplateWriter
from excel.protocols import TypedValue


@pytest.fixture
def template_various_merges(tmp_path):
    """Create template with various merge sizes."""
    wb = Workbook()
    ws = wb.active
    
    # 1x3 horizontal (above insertion zone)
    ws["A1"] = "Title"
    ws.merge_cells("A1:C1")
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # Headers
    ws["A3"] = "Index"
    ws["B3"] = "Col1"
    ws["C3"] = "Col2"
    
    # Template rows
    ws["A4"] = "a"
    ws["B4"] = "{{data|table(join=outer,on=Index)}}"
    ws["A5"] = "b"
    
    # 3x1 vertical (below insertion zone)
    ws["A7"] = "Status"
    ws.merge_cells("A7:A9")
    ws["A7"].font = Font(italic=True)
    ws["A7"].alignment = Alignment(vertical="center")
    
    # 2x2 rectangular (below insertion zone)
    ws["D7"] = "Summary"
    ws.merge_cells("D7:E8")
    ws["D7"].font = Font(bold=True)
    ws["D7"].alignment = Alignment(horizontal="center", vertical="center")
    
    path = tmp_path / "template_various_merges.xlsx"
    wb.save(path)
    return str(path)


def _run_outer_join(template_path, tmp_path):
    """Fill template with outer join DF that adds one row."""
    df = pl.DataFrame({
        "Index": ["a", "b", "c"],
        "Col1": [10, 20, 30],
        "Col2": [100, 200, 300],
    })
    output = str(tmp_path / "output.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(df, "table")}, output)
    return load_workbook(output), output


class TestHorizontalMerge1x3:
    """A1:C1 merge (1 row × 3 cols) — above insertion zone."""
    
    def test_merge_preserved(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        merges = {str(m) for m in ws.merged_cells.ranges}
        assert "A1:C1" in merges, f"1x3 merge destroyed. Got: {merges}"
        wb.close()
    
    def test_value_preserved(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        assert ws["A1"].value == "Title"
        wb.close()
    
    def test_font_preserved(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        assert ws["A1"].font.bold is True
        wb.close()


class TestVerticalMerge3x1:
    """A7:A9 merge (3 rows × 1 col) — below insertion zone, shifts to A8:A10."""
    
    def test_merge_shifts(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        merges = {str(m) for m in ws.merged_cells.ranges}
        assert "A8:A10" in merges, f"3x1 merge did not shift to A8:A10. Got: {merges}"
        wb.close()
    
    def test_value_preserved(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        assert ws["A8"].value == "Status", f"3x1 value lost, got: {ws['A8'].value}"
        wb.close()
    
    def test_font_preserved(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        assert ws["A8"].font.italic is True
        wb.close()


class TestRectangularMerge2x2:
    """D7:E8 merge (2 rows × 2 cols) — below insertion zone, shifts to D8:E9."""
    
    def test_merge_shifts(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        merges = {str(m) for m in ws.merged_cells.ranges}
        assert "D8:E9" in merges, f"2x2 merge did not shift to D8:E9. Got: {merges}"
        wb.close()
    
    def test_value_preserved(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        assert ws["D8"].value == "Summary", f"2x2 value lost, got: {ws['D8'].value}"
        wb.close()
    
    def test_font_and_alignment_preserved(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        assert ws["D8"].font.bold is True
        assert ws["D8"].alignment.horizontal == "center"
        assert ws["D8"].alignment.vertical == "center"
        wb.close()


class TestDataFillCorrectness:
    """All template rows filled correctly despite complex merges."""
    
    def test_all_rows_filled(self, template_various_merges, tmp_path):
        wb, _ = _run_outer_join(template_various_merges, tmp_path)
        ws = wb.active
        
        assert ws["A4"].value == "a" and ws["B4"].value == 10 and ws["C4"].value == 100
        assert ws["A5"].value == "b" and ws["B5"].value == 20 and ws["C5"].value == 200
        assert ws["A6"].value == "c" and ws["B6"].value == 30 and ws["C6"].value == 300
        wb.close()
