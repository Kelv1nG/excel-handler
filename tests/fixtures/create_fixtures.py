"""
Run this script to (re)generate all Excel test fixtures.

    uv run python tests/fixtures/create_fixtures.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color

FIXTURES_DIR = Path(__file__).parent


def _header_style(ws, row: int, cols: range):
    """Bold + light-blue fill for header cells."""
    fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    for col in cols:
        cell = ws.cell(row, col)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# simple_table.xlsx
# Single clean table at A1 — Name / Amount / Category
# Tests: extract_table_by_column_names, partial columns, unordered columns,
#        extract_table_by_range, extract_table_from_cell, extract_table_near_cell
# ---------------------------------------------------------------------------
def create_simple_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Amount", "Category"])
    ws.append(["Alice", 100, "Food"])
    ws.append(["Bob", 200, "Travel"])
    ws.append(["Carol", 300, "Food"])
    _header_style(ws, 1, range(1, 4))
    wb.save(FIXTURES_DIR / "simple_table.xlsx")
    wb.close()
    print("  simple_table.xlsx")


# ---------------------------------------------------------------------------
# no_headers.xlsx
# Pure data rows, no header row
# Tests: extract_table_by_range(has_headers=False, column_names=[...])
#        extract_table_by_range(has_headers=False)  → auto col_0, col_1, ...
# ---------------------------------------------------------------------------
def create_no_headers():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append([1, "Alice", 100])
    ws.append([2, "Bob", 200])
    ws.append([3, "Carol", 300])
    wb.save(FIXTURES_DIR / "no_headers.xlsx")
    wb.close()
    print("  no_headers.xlsx")


# ---------------------------------------------------------------------------
# merged_cells.xlsx
# Region column is merged across two rows per region
# Tests: unmerge_cells=True + fill_forward=True removes nulls
#        unmerge_cells=False + fill_forward=False leaves nulls
# ---------------------------------------------------------------------------
def create_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Region"
    ws["B1"] = "Country"
    ws["C1"] = "Sales"
    ws["A2"] = "Europe"
    ws["B2"] = "France"
    ws["C2"] = 100
    ws["B3"] = "Germany"
    ws["C3"] = 150
    ws["A4"] = "Asia"
    ws["B4"] = "Japan"
    ws["C4"] = 200
    ws["B5"] = "China"
    ws["C5"] = 250
    ws.merge_cells("A2:A3")  # "Europe" spans rows 2-3
    ws.merge_cells("A4:A5")  # "Asia"   spans rows 4-5
    _header_style(ws, 1, range(1, 4))
    wb.save(FIXTURES_DIR / "merged_cells.xlsx")
    wb.close()
    print("  merged_cells.xlsx")


# ---------------------------------------------------------------------------
# multiple_sheets.xlsx
# Same columns (Name, Amount) exist in Sheet1 AND Sheet2
# Tests: extract_table_by_column_names → returns first sheet match (Sheet1)
#        extract_table_by_column_names_from_sheet → targets Sheet2 specifically
# ---------------------------------------------------------------------------
def create_multiple_sheets():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Name", "Amount"])
    ws1.append(["Alice", 100])
    _header_style(ws1, 1, range(1, 3))

    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["Name", "Amount"])
    ws2.append(["Bob", 200])
    _header_style(ws2, 1, range(1, 3))

    wb.save(FIXTURES_DIR / "multiple_sheets.xlsx")
    wb.close()
    print("  multiple_sheets.xlsx")


# ---------------------------------------------------------------------------
# multiple_tables.xlsx
# Two tables with identical headers on the same sheet (rows 1 and 3)
# Tests: extract_table_by_column_names → raises MultipleTablesFoundError
# ---------------------------------------------------------------------------
def create_multiple_tables():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Amount"])  # row 1 — table 1 header
    ws.append(["Alice", 100])      # row 2 — table 1 data
    ws.append(["Name", "Amount"])  # row 3 — table 2 header
    ws.append(["Bob", 200])        # row 4 — table 2 data
    _header_style(ws, 1, range(1, 3))
    _header_style(ws, 3, range(1, 3))
    wb.save(FIXTURES_DIR / "multiple_tables.xlsx")
    wb.close()
    print("  multiple_tables.xlsx")


# ---------------------------------------------------------------------------
# empty_table.xlsx
# Headers present (Name / Amount) but no data rows below
# Tests: extract_table_by_column_names → returns empty DataFrame with schema
#        extract_table_near_cell from below headers → raises TableNotFoundError
# ---------------------------------------------------------------------------
def create_empty_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Amount"])
    _header_style(ws, 1, range(1, 3))
    wb.save(FIXTURES_DIR / "empty_table.xlsx")
    wb.close()
    print("  empty_table.xlsx")


# ---------------------------------------------------------------------------
# offset_table.xlsx
# Table starting at C3 — rows 1-2 and columns A-B are empty
# Tests: extract_table_from_cell("C3") → finds table directly
#        extract_table_near_cell("A1") → scans right+down to C3
# ---------------------------------------------------------------------------
def create_offset_table():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["C3"] = "Name"
    ws["D3"] = "Amount"
    ws["E3"] = "Category"
    ws["C4"] = "Alice"
    ws["D4"] = 100
    ws["E4"] = "Food"
    ws["C5"] = "Bob"
    ws["D5"] = 200
    ws["E5"] = "Travel"
    _header_style(ws, 3, range(3, 6))
    wb.save(FIXTURES_DIR / "offset_table.xlsx")
    wb.close()
    print("  offset_table.xlsx")


# ---------------------------------------------------------------------------
# template.xlsx
# Template with {{ variable }} tags spread across two sheets.
# A third sheet (EmptySheet) has no tags — it must be excluded from results.
#
# Sheet1 layout:
#   B2  → {{ revenue }}
#   C4  → {{ title | orientation=horizontal }}
#   D6  → {{ count | skip=2, flag=True }}
#   A8  → plain text (no tag) — must not appear in results
#   B8  → 42  (numeric) — must not appear in results
#
# Sheet2 layout:
#   A1  → {{ summary }}
#
# Tests: ExcelTemplateReader.read() happy path + error cases,
#        MarkedCell.parse_metadata() coercion
# ---------------------------------------------------------------------------
def create_template():
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["B2"] = "{{ revenue }}"
    ws1["C4"] = "{{ title | orientation=horizontal }}"
    ws1["D6"] = "{{ count | skip=2, flag=True }}"
    ws1["A8"] = "plain text — no tag here"
    ws1["B8"] = 42

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "{{ summary }}"

    wb.create_sheet("EmptySheet")  # no tags — must be excluded from results

    wb.save(FIXTURES_DIR / "template.xlsx")
    wb.close()
    print("  template.xlsx")




# ---------------------------------------------------------------------------
# template_merge_preservation.xlsx
# Template with:
#   - A1:C1 merged title cell (bold, centered, wrap_text) ABOVE the table
#   - A table at rows 3-4 with an outer-join tag
#   - A6:C6 merged footer cell (italic) BELOW the table
# Used to verify that merged cells outside the insertion zone survive intact
# after ExcelTemplateWriter inserts extra outer-join rows.
# ---------------------------------------------------------------------------
def create_template_merge_preservation():
    from openpyxl.styles import Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Title: A1:C1 merged — bold, centered, wrap_text, light-blue fill
    ws["A1"] = "Report Title"
    title_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    thin = Side(style="thin")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells("A1:C1")

    # Headers row 2
    ws["A2"] = "Index"
    ws["B2"] = "col1"
    ws["C2"] = "col2"
    _header_style(ws, 2, range(1, 4))

    # Data rows 3-4 with outer-join tag in B3
    ws["A3"] = "a"
    ws["B3"] = "{{ data | table(join=outer, on=Index) }}"
    ws["A4"] = "b"

    # Empty row 5 (separator)

    # Footer: A6:C6 merged — italic
    ws["A6"] = "Footer Note"
    ws["A6"].font = Font(italic=True)
    ws["A6"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A6:C6")

    wb.save(FIXTURES_DIR / "template_merge_preservation.xlsx")
    wb.close()
    print("  template_merge_preservation.xlsx")

# ---------------------------------------------------------------------------
# template_complex_merges.xlsx
# Template with:
#   - A1:C1 merged title ABOVE table (1x3)
#   - A table at rows 3-4 with outer-join tag
#   - A6:B7 merged 2x2 "BigNote" BELOW table
#   - A9:C9 merged 1x3 "Footer2" BELOW 2x2 merge
# Tests: multiple merges of different shapes all shift correctly on multi-insert
# ---------------------------------------------------------------------------
def create_template_complex_merges():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Complex Report"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells("A1:C1")

    ws["A2"] = "Index"
    ws["B2"] = "col1"
    ws["C2"] = "col2"
    _header_style(ws, 2, range(1, 4))

    ws["A3"] = "a"
    ws["B3"] = "{{ data | table(join=outer, on=Index) }}"
    ws["A4"] = "b"

    # 2x2 merge at A6:B7
    ws["A6"] = "BigNote"
    ws["A6"].font = Font(bold=True, italic=True)
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.merge_cells("A6:B7")

    # Footer at A9:C9
    ws["A9"] = "Footer2"
    ws["A9"].font = Font(italic=True, color="FF0000")
    ws["A9"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A9:C9")

    wb.save(FIXTURES_DIR / "template_complex_merges.xlsx")
    wb.close()
    print("  template_complex_merges.xlsx")


# ---------------------------------------------------------------------------
# template_tight_footer.xlsx
# Footer merge immediately below last data row (no blank separator).
# ---------------------------------------------------------------------------
def create_template_tight_footer():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "col1"
    ws["C1"] = "col2"
    _header_style(ws, 1, range(1, 4))

    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer, on=Index) }}"
    ws["A3"] = "b"

    # Footer immediately at row 4 (no separator)
    ws["A4"] = "Tight Footer"
    ws["A4"].font = Font(italic=True)
    ws["A4"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A4:C4")

    wb.save(FIXTURES_DIR / "template_tight_footer.xlsx")
    wb.close()
    print("  template_tight_footer.xlsx")


# ---------------------------------------------------------------------------
# template_left_join.xlsx
# Left join — no row insertion happens. Merges must be completely untouched.
# ---------------------------------------------------------------------------
def create_template_left_join():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Left Join Report"
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells("A1:C1")

    ws["A2"] = "Index"
    ws["B2"] = "col1"
    ws["C2"] = "col2"
    _header_style(ws, 2, range(1, 4))

    ws["A3"] = "a"
    ws["B3"] = "{{ data | table(join=left, on=Index) }}"
    ws["A4"] = "b"

    ws["A6"] = "Left Join Footer"
    ws["A6"].font = Font(italic=True, bold=True)
    ws["A6"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A6:C6")

    wb.save(FIXTURES_DIR / "template_left_join.xlsx")
    wb.close()
    print("  template_left_join.xlsx")


# ---------------------------------------------------------------------------
# template_vertical_merge.xlsx
# Table with a 3-row vertical merge (A5:A7) immediately below the last data
# row (A4 is blank separator, A5:A7 = "Status" italic).
# Used to verify that the vertical merge is correctly shifted down (not split)
# when outer join inserts extra rows.
# Variant 2: merge IMMEDIATELY adjacent (no blank separator) at A4:A6.
# ---------------------------------------------------------------------------
def create_template_vertical_merge():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws["A1"] = "Index"
    ws["B1"] = "col1"
    ws["C1"] = "col2"
    _header_style(ws, 1, range(1, 4))

    # Data rows
    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer, on=Index) }}"
    ws["A3"] = "b"

    # Row 4: blank separator

    # Vertical 3-row merge below the table
    ws["A5"] = "Status"
    ws["A5"].font = Font(italic=True)
    ws["A5"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells("A5:A7")

    wb.save(FIXTURES_DIR / "template_vertical_merge.xlsx")
    wb.close()
    print("  template_vertical_merge.xlsx")


def create_template_vertical_merge_adjacent():
    """Vertical merge immediately below the last data row (no blank separator)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "col1"
    ws["C1"] = "col2"
    _header_style(ws, 1, range(1, 4))

    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer, on=Index) }}"
    ws["A3"] = "b"

    # Vertical 3-row merge IMMEDIATELY after last data row
    ws["A4"] = "Status"
    ws["A4"].font = Font(italic=True)
    ws["A4"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells("A4:A6")

    wb.save(FIXTURES_DIR / "template_vertical_merge_adjacent.xlsx")
    wb.close()
    print("  template_vertical_merge_adjacent.xlsx")


def create_template_data_col_vertical_merge():
    """Vertical 3-row merge in a DATA column (not join col) at the boundary row.

    Template:
      Row 1: headers
      Row 2: a  {{tag|outer}}
      Row 3: b
      Row 4: Status (A4, not in DF)  +  B4:B6 = "Section Header" italic merge

    _find_last_data_row must stop at row 3 (before the merge at row 4) so
    that _copy_row_styles uses source_row=3, making B4:B6 a "below" merge
    that shifts down intact (to B5:B7 after 1 insertion, or B6:B8 for 2, etc.)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "col1"
    ws["C1"] = "col2"
    _header_style(ws, 1, range(1, 4))

    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer, on=Index) }}"
    ws["A3"] = "b"

    # Row 4: join col has a value NOT in DF, data col has 3-row merge
    ws["A4"] = "Status"
    ws["B4"] = "Section Header"
    ws["B4"].font = Font(italic=True)
    ws["B4"].alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells("B4:B6")

    wb.save(FIXTURES_DIR / "template_data_col_vertical_merge.xlsx")
    wb.close()
    print("  template_data_col_vertical_merge.xlsx")


# ---------------------------------------------------------------------------
# template_record.xlsx
# Template for testing record (dot-notation) variable access.
#
# Sheet1:
#   Row 1: label/value header
#   Row 2: "Company"  | {{ result.Company }}
#   Row 3: "Revenue"  | {{ result.Revenue }}
#   Row 4: "Quarter"  | {{ other.Quarter }}
#   Row 5: "Title"    | {{ title }}          <- plain scalar for mixed test
# ---------------------------------------------------------------------------
def create_template_record():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Field"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "Company"
    ws["B2"] = "{{ result.Company }}"

    ws["A3"] = "Revenue"
    ws["B3"] = "{{ result.Revenue }}"

    ws["A4"] = "Quarter"
    ws["B4"] = "{{ other.Quarter }}"

    ws["A5"] = "Title"
    ws["B5"] = "{{ title }}"

    wb.save(FIXTURES_DIR / "template_record.xlsx")
    wb.close()
    print("  template_record.xlsx")


# ---------------------------------------------------------------------------
# template_sorted_outer_asc.xlsx
# Sorted outer join — ascending, no fixed zone, Option A end_table.
# Row 1: Index, Value  (headers)
# Row 2: c,  {{ data | table(join=outer, on=Index, order_by=asc) }}
# Row 3: a,  (empty)
# Row 4: (empty), {{ end_table }}   ← Option A: own row, no join value
# ---------------------------------------------------------------------------
def create_template_sorted_outer_asc():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "c"
    ws["B2"] = "{{ data | table(join=outer, on=Index, order_by=asc) }}"
    ws["A3"] = "a"

    ws["B4"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_sorted_outer_asc.xlsx")
    wb.close()
    print("  template_sorted_outer_asc.xlsx")


# ---------------------------------------------------------------------------
# template_sorted_outer_desc.xlsx
# Same layout as asc but tag uses order_by=desc.
# ---------------------------------------------------------------------------
def create_template_sorted_outer_desc():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "c"
    ws["B2"] = "{{ data | table(join=outer, on=Index, order_by=desc) }}"
    ws["A3"] = "a"

    ws["B4"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_sorted_outer_desc.xlsx")
    wb.close()
    print("  template_sorted_outer_desc.xlsx")


# ---------------------------------------------------------------------------
# template_sorted_outer_fixed.xlsx
# Sorted outer join with a fixed lower zone via {{ insert_data }}.
# Row 1: Index, Value  (headers)
# Row 2: c,  {{ data | table(join=outer, on=Index, order_by=asc) }}
# Row 3: a,  (empty)
# Row 4: (empty), {{ insert_data }}   ← insertion point marker
# Row 5: total, {{ end_table }}       ← Option B: on data row
# ---------------------------------------------------------------------------
def create_template_sorted_outer_fixed():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "c"
    ws["B2"] = "{{ data | table(join=outer, on=Index, order_by=asc) }}"
    ws["A3"] = "a"

    ws["B4"] = "{{ insert_data }}"

    ws["A5"] = "total"
    ws["B5"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_sorted_outer_fixed.xlsx")
    wb.close()
    print("  template_sorted_outer_fixed.xlsx")


# ---------------------------------------------------------------------------
# template_sorted_outer_by_col.xlsx
# Sorted outer join — sort by a non-join column (Value:desc).
# Row 1: Index, Value  (headers)
# Row 2: c,  {{ data | table(join=outer, on=Index, order_by=Value:desc) }}
# Row 3: a,  (empty)
# Row 4: (empty), {{ end_table }}
# ---------------------------------------------------------------------------
def create_template_sorted_outer_by_col():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "c"
    ws["B2"] = "{{ data | table(join=outer, on=Index, order_by=Value:desc) }}"
    ws["A3"] = "a"

    ws["B4"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_sorted_outer_by_col.xlsx")
    wb.close()
    print("  template_sorted_outer_by_col.xlsx")


# ---------------------------------------------------------------------------
# template_sorted_outer_shorter.xlsx
# Sorted outer join where df has fewer rows than upper zone template slots.
# Row 1: Index, Value  (headers)
# Row 2: c,  {{ data | table(join=outer, on=Index, order_by=asc) }}
# Row 3: a,  (empty)
# Row 4: b,  (empty)
# Row 5: (empty), {{ end_table }}   ← 3 upper slots, but df only provides 2 rows
# ---------------------------------------------------------------------------
def create_template_sorted_outer_shorter():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "c"
    ws["B2"] = "{{ data | table(join=outer, on=Index, order_by=asc) }}"
    ws["A3"] = "a"
    ws["A4"] = "b"

    ws["B5"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_sorted_outer_shorter.xlsx")
    wb.close()
    print("  template_sorted_outer_shorter.xlsx")


# ---------------------------------------------------------------------------
# template_sorted_outer_tmpl_rows.xlsx
# Sorted outer join where the upper zone has template-only rows (foo, bar)
# that are NOT present in the DataFrame — they must survive the sort.
# Row 1: Index, Value  (headers)
# Row 2: foo, {{ data | table(join=outer, on=Index, order_by=asc) }}
# Row 3: bar
# Row 4: (B4={{ insert_data }})         ← insertion point marker
# Row 5: No Sector                       ← lower zone (fixed)
# Row 6: Total, {{ end_table }}          ← lower zone fixed row + Option B end_table
# ---------------------------------------------------------------------------
def create_template_sorted_outer_tmpl_rows():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "foo"
    ws["B2"] = "{{ data | table(join=outer, on=Index, order_by=asc) }}"
    ws["A3"] = "bar"

    ws["B4"] = "{{ insert_data }}"

    ws["A5"] = "No Sector"
    ws["A6"] = "Total"
    ws["B6"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_sorted_outer_tmpl_rows.xlsx")
    wb.close()
    print("  template_sorted_outer_tmpl_rows.xlsx")


# ---------------------------------------------------------------------------
# template_fill_global.xlsx
# Outer join with fill=0 — every null data column in the output gets 0.
# Row 1: Index, col1, col2  (headers)
# Row 2: a,  {{ data | table(join=outer, on=Index, fill=0) }}
# Row 3: x  (template-only, not in df)
# Row 4: (empty), {{ end_table }}   ← Option A
# ---------------------------------------------------------------------------
def create_template_fill_global():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "col1"
    ws["C1"] = "col2"
    _header_style(ws, 1, range(1, 4))

    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer, on=Index, fill=0) }}"
    ws["A3"] = "x"

    ws["B4"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_fill_global.xlsx")
    wb.close()
    print("  template_fill_global.xlsx")


# ---------------------------------------------------------------------------
# template_fill_per_col.xlsx
# Outer join with fill=col1:0;col2:N/A — per-column fill values.
# Same layout as template_fill_global.xlsx.
# ---------------------------------------------------------------------------
def create_template_fill_per_col():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "col1"
    ws["C1"] = "col2"
    _header_style(ws, 1, range(1, 4))

    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer, on=Index, fill=col1:0;col2:N/A) }}"
    ws["A3"] = "x"

    ws["B4"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_fill_per_col.xlsx")
    wb.close()
    print("  template_fill_per_col.xlsx")


# ---------------------------------------------------------------------------
# template_fill_lower_zone.xlsx
# Outer join with fill=0, insert_data marker, and an unmatched lower zone row.
# The lower zone row has no df match — fill=0 must still apply to its nulls.
# Row 1: Index, col1  (headers)
# Row 2: a,  {{ data | table(join=outer, on=Index, fill=0) }}
# Row 3:     {{ insert_data }}    ← insertion point
# Row 4: Total                   ← lower zone (NOT in df; col1 → 0 via fill)
# Row 5:     {{ end_table }}     ← Option A
# ---------------------------------------------------------------------------
def create_template_fill_lower_zone():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "col1"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer, on=Index, fill=0) }}"

    ws["B3"] = "{{ insert_data }}"

    ws["A4"] = "Total"

    ws["B5"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_fill_lower_zone.xlsx")
    wb.close()
    print("  template_fill_lower_zone.xlsx")


# ---------------------------------------------------------------------------
# template_fill_sorted_outer_lower_zone.xlsx
# Sorted outer join with fill=0, insert_data, and an unmatched lower zone row.
# The unmatched lower zone row must get fill=0, not stay blank.
# Row 1: Sector, ColValue  (headers)
# Row 2: Foo,  {{ data | table(join=outer, on=Sector, order_by=asc, fill=0) }}
# Row 3: Bar
# Row 4:       {{ insert_data }}    ← insertion point
# Row 5: No Sector                  ← lower zone: NOT in df → ColValue must be 0
# Row 6: Total  {{ end_table }}     ← lower zone: matched in df
# ---------------------------------------------------------------------------
def create_template_fill_sorted_outer_lower_zone():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Sector"
    ws["B1"] = "ColValue"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "Foo"
    ws["B2"] = "{{ data | table(join=outer, on=Sector, order_by=asc, fill=0) }}"
    ws["A3"] = "Bar"

    ws["B4"] = "{{ insert_data }}"

    ws["A5"] = "No Sector"
    ws["A6"] = "Total"
    ws["B6"] = "{{ end_table }}"

    wb.save(FIXTURES_DIR / "template_fill_sorted_outer_lower_zone.xlsx")
    wb.close()
    print("  template_fill_sorted_outer_lower_zone.xlsx")


# ---------------------------------------------------------------------------
# anchored_cells.xlsx
# Fixture for ExcelCellReader.get_relative / get_many_relative tests.
#
# Sheet1 layout:
#   A1 = "Revenue Label"   B1 = 5000   C1 = "USD"
#   A2 = "Tax Label"       B2 = 250
#   A3 = "Note"
#
# Sheet2 layout:
#   A1 = "Revenue Label"   B1 = 9999   ← duplicate keyword for error tests
#
# Tests:
#   - get_relative(cell_ref="Sheet1!A1", right=1)   → 5000
#   - get_relative(cell_ref="Sheet1!A1", right=2)   → "USD"
#   - get_relative(cell_ref="Sheet1!B1", left=1)    → "Revenue Label"
#   - get_relative(cell_ref="Sheet1!A2", up=1)      → "Revenue Label"
#   - get_relative(keyword="Revenue Label", sheet="Sheet1", right=1)           → 5000
#   - get_relative(keyword="Tax Label",    sheet="Sheet1", right=1)            → 250
#   - get_relative(keyword="Revenue Label", sheet="Sheet1", down=1)            → "Tax Label"
#   - get_relative(keyword="Revenue Label")  → raises MultipleKeywordsFoundError
#   - get_relative(keyword="No Such Label") → raises KeywordNotFoundError
#   - get_many_relative(cell_ref="Sheet1!A1", offsets={...})  → dict
#   - get_many_relative(keyword="Tax Label", sheet="Sheet1", offsets={...}) → dict
# ---------------------------------------------------------------------------
def create_anchored_cells():
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = "Revenue Label"
    ws1["B1"] = 5000
    ws1["C1"] = "USD"
    ws1["A2"] = "Tax Label"
    ws1["B2"] = 250
    ws1["A3"] = "Note"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Revenue Label"
    ws2["B1"] = 9999

    wb.save(FIXTURES_DIR / "anchored_cells.xlsx")
    wb.close()
    print("  anchored_cells.xlsx")


# ---------------------------------------------------------------------------
# bug-on-insert.xlsx
# Template for testing scalar cells below an expanding outer-join table.
# Includes border styling on rows 17-18 and 21-27 for comprehensive merge/
# border preservation tests.
# 
# Row 3-10: First test case - simple outer join with scalar below
# Row 12-14: Second test case - placeholder + style=first
# Row 17-18: table border styles (thick/thin mix, bold text with theme colors)
# Row 21-27: complex border structure + fill + scalar placeholder
# ---------------------------------------------------------------------------
def create_bug_on_insert():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # === FIRST TEST CASE: rows 3-10 ===
    ws["B3"] = "index"
    ws["C3"] = "colA"
    ws["D3"] = "colB"
    
    ws["B4"] = "a"
    ws["C4"] = "{{ my_table | table(join=outer) }}"
    
    ws["B5"] = "b"
    ws["B6"] = "c"
    
    ws["C7"] = "{{ insert_data }}"
    
    ws["B8"] = "d"
    ws["D8"] = None
    ws["B8"].font = Font(bold=True, color=Color(theme=1))
    ws["D8"].font = Font(bold=True, color=Color(theme=1))
    
    ws["B9"] = "Total"
    ws["C9"] = "{{ end_table }}"
    ws["B9"].font = Font(bold=True, color=Color(theme=1))
    ws["C9"].font = Font(bold=True, color=Color(theme=1))
    
    ws["B10"] = None
    ws["C10"] = None
    ws["B10"].font = Font(bold=True, color=Color(theme=1))
    ws["C10"].font = Font(bold=True, color=Color(theme=1))
    
    # === SECOND TEST CASE: rows 12-14 (placeholder=True, style=first) ===
    ws["B12"] = "index"
    ws["C12"] = "colA"
    ws["D12"] = "colB"
    
    ws["C13"] = "{{ my_table | table(join=outer, placeholder=True, style=first) }}"
    
    ws["B14"] = "Total"
    ws["C14"] = "{{ end_table | insert=above }}"
    ws["D14"] = None
    ws["B14"].font = Font(bold=True, color=Color(theme=1))
    ws["C14"].font = Font(bold=True, color=Color(theme=1))
    ws["D14"].font = Font(bold=True, color=Color(theme=1))
    
    # === STYLED SECTION 1: rows 17-18 (borders) ===
    ws["B17"] = "headerColumn1"
    ws["B17"].font = Font(bold=True, color=Color(theme=1))
    ws["B17"].border = Border(
        left=Side(style="thick", color=Color(theme=1)),
        right=Side(style="thin", color=Color(theme=1)),
        top=Side(style="thick", color=Color(theme=1)),
        bottom=Side(style="thin", color=Color(theme=1))
    )
    ws["B17"].alignment = Alignment(horizontal="center")
    
    ws["C17"] = None
    ws["C17"].font = Font(color=Color(theme=1))
    ws["C17"].border = Border(
        right=Side(style="thin", color=Color(theme=1)),
        top=Side(style="thick", color=Color(theme=1)),
        bottom=Side(style="thin", color=Color(theme=1))
    )
    
    ws["D17"] = "MergedCol1"
    ws["D17"].font = Font(bold=True, color=Color(theme=1))
    ws["D17"].border = Border(
        left=Side(style="thin", color=Color(theme=1)),
        right=Side(style="thin", color=Color(theme=1)),
        top=Side(style="thick", color=Color(theme=1)),
        bottom=Side(style="thin", color=Color(theme=1))
    )
    ws["D17"].alignment = Alignment(horizontal="center")
    
    ws["B18"] = "subheaderColumn1"
    ws["B18"].font = Font(bold=True, color=Color(theme=1))
    ws["B18"].border = Border(
        left=Side(style="thick", color=Color(theme=1)),
        right=Side(style="thin", color=Color(theme=1)),
        top=Side(style="thin", color=Color(theme=1)),
        bottom=Side(style="thick", color=Color(theme=1))
    )
    
    ws["C18"] = "subheaderColumn2"
    ws["C18"].font = Font(bold=True, color=Color(theme=1))
    ws["C18"].border = Border(
        left=Side(style="thin", color=Color(theme=1)),
        right=Side(style="thin", color=Color(theme=1)),
        top=Side(style="thin", color=Color(theme=1)),
        bottom=Side(style="thick", color=Color(theme=1))
    )
    
    ws["D18"] = None
    ws["D18"].font = Font(color=Color(theme=1))
    ws["D18"].border = Border(
        left=Side(style="thin", color=Color(theme=1)),
        right=Side(style="thin", color=Color(theme=1)),
        bottom=Side(style="thick", color=Color(theme=1))
    )
    
    # === STYLED SECTION 2: rows 21-27 (borders + fill + scalar) ===
    ws["B21"] = "some text here"
    ws["B21"].font = Font(color=Color(theme=0))
    ws["B21"].fill = PatternFill(
        start_color=Color(theme=4, tint=-0.249977111117893),
        end_color=Color(theme=4, tint=-0.249977111117893),
        fill_type="solid"
    )
    ws["B21"].border = Border(
        left=Side(style="thick", color=Color(theme=1)),
        right=Side(style="thick", color=Color(theme=1)),
        top=Side(style="thick", color=Color(theme=1))
    )
    ws["B21"].alignment = Alignment(horizontal="center")
    
    ws["C21"] = None
    ws["C21"].font = Font(color=Color(theme=1))
    ws["C21"].border = Border(top=Side(style="thick", color=Color(theme=1)))
    
    ws["D21"] = None
    ws["D21"].font = Font(color=Color(theme=1))
    ws["D21"].border = Border(top=Side(style="thick", color=Color(theme=1)))
    
    ws["B22"] = "{{ some_value }}"
    ws["B22"].font = Font(color=Color(theme=1))
    ws["B22"].border = Border(
        left=Side(style="thin", color=Color(theme=1)),
        right=Side(style="thin", color=Color(theme=1)),
        top=Side(style="thin", color=Color(theme=1)),
        bottom=Side(style="thin", color=Color(theme=1))
    )
    ws["B22"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws["C22"] = None
    ws["C22"].font = Font(color=Color(theme=1))
    ws["C22"].border = Border(top=Side(style="thin", color=Color(theme=1)))
    
    ws["D22"] = None
    ws["D22"].font = Font(color=Color(theme=1))
    ws["D22"].border = Border(top=Side(style="thin", color=Color(theme=1)))
    
    # Left border column B23-B27
    for row in range(23, 27):
        ws[f"B{row}"] = None
        ws[f"B{row}"].font = Font(color=Color(theme=1))
        ws[f"B{row}"].border = Border(left=Side(style="thin", color=Color(theme=1)))
    
    # Bottom border row 27
    ws["B27"] = None
    ws["B27"].font = Font(color=Color(theme=1))
    ws["B27"].border = Border(
        left=Side(style="thin", color=Color(theme=1)),
        bottom=Side(style="thin", color=Color(theme=1))
    )
    
    ws["C27"] = None
    ws["C27"].font = Font(color=Color(theme=1))
    ws["C27"].border = Border(bottom=Side(style="thin", color=Color(theme=1)))
    
    ws["D27"] = None
    ws["D27"].font = Font(color=Color(theme=1))
    ws["D27"].border = Border(bottom=Side(style="thin", color=Color(theme=1)))
    
    wb.save(FIXTURES_DIR / "bug-on-insert.xlsx")
    wb.close()
    print("  bug-on-insert.xlsx")


# ---------------------------------------------------------------------------
# template_placeholder_outer.xlsx
# Template for placeholder=true + end_table|insert=above tests.
# Row 1: headers (Index, Value)
# Row 2: blank Index, {{ data | table(join=outer, placeholder=true) }} — plain
# Row 3: Total, {{ end_table | insert=above }} — bold + yellow fill
# Tests: placeholder row deleted when unmatched; Total pinned via Option C
# ---------------------------------------------------------------------------
def create_template_placeholder_outer():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    # Row 2: blank join col, tag cell (plain style — no bold, no fill)
    ws["B2"] = "{{ data | table(join=outer, placeholder=true) }}"

    # Row 3: Total row — bold + yellow fill
    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"] = "Total"
    ws["B3"] = "{{ end_table | insert=above }}"
    for col in range(1, 3):
        ws.cell(3, col).font = Font(bold=True)
        ws.cell(3, col).fill = total_fill

    wb.save(FIXTURES_DIR / "template_placeholder_outer.xlsx")
    wb.close()
    print("  template_placeholder_outer.xlsx")


# ---------------------------------------------------------------------------
# template_style_src_last.xlsx — style=last (default)
# template_style_src_first.xlsx — style=first
# Both: Row 1 headers, Row 2 plain data row (tag), Row 3 Total (bold + yellow)
# Tests: style=last → inserted rows inherit bold/yellow;
#        style=first → inserted rows inherit plain style
# ---------------------------------------------------------------------------
def create_template_style_src_last():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer) }}"

    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"] = "Total"
    ws["B3"] = 100
    for col in range(1, 3):
        ws.cell(3, col).font = Font(bold=True)
        ws.cell(3, col).fill = total_fill

    wb.save(FIXTURES_DIR / "template_style_src_last.xlsx")
    wb.close()
    print("  template_style_src_last.xlsx")


def create_template_style_src_first():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["A2"] = "a"
    ws["B2"] = "{{ data | table(join=outer, style=first) }}"

    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"] = "Total"
    ws["B3"] = 100
    for col in range(1, 3):
        ws.cell(3, col).font = Font(bold=True)
        ws.cell(3, col).fill = total_fill

    wb.save(FIXTURES_DIR / "template_style_src_first.xlsx")
    wb.close()
    print("  template_style_src_first.xlsx")


# ---------------------------------------------------------------------------
# template_empty_outer_style_first.xlsx — empty table with outer+style=first
# Minimal template: just headers + tag row, no pre-existing data rows,
# with end_table|insert=above marker.
# Tests: that style=first works correctly when table is empty (no data rows
# between tag and end_table).  When data is added, inserted rows should copy
# style from the tag row (plain), not from end_table row (if styled).
# ---------------------------------------------------------------------------
def create_template_empty_outer_style_first():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Index"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    # Plain tag row with outer join + style=first
    ws["A2"] = None  # Join column placeholder
    ws["B2"] = "{{ data | table(join=outer, style=first) }}"

    # End marker row with bold + yellow fill (styled)
    end_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"] = "{{ end_table | insert=above }}"
    for col in range(1, 3):
        ws.cell(3, col).font = Font(bold=True)
        ws.cell(3, col).fill = end_fill

    wb.save(FIXTURES_DIR / "template_empty_outer_style_first.xlsx")
    wb.close()
    print("  template_empty_outer_style_first.xlsx")


# ---------------------------------------------------------------------------
# template_combo_outer_merges_below.xlsx
# Outer join (placeholder=True) with two adjacent same-span merges directly
# below the table.  Primary regression test for the merge-shift bug.
#
# Row 1: headers Key / Value
# Row 2: blank Key, {{ data | table(join=outer, placeholder=True) }}
# Row 3: "Total", {{ end_data | insert=above }} (bold, yellow)
# Row 5: A5:B5 merged "Note First" (bold, blue fill)
# Row 6: A6:B6 merged "Note Second" (italic, orange fill)
#
# DF: 3 real rows (a, b, c) + Total → placeholder deleted, a/b/c inserted
# before Total → net +2 shift → merges land at A7:B7 and A8:B8.
# Tests: both adjacent merges survive, correct values + styles, no stale
# ---------------------------------------------------------------------------
def create_template_combo_outer_merges_below():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Key"
    ws["B1"] = "Value"
    _header_style(ws, 1, range(1, 3))

    ws["B2"] = "{{ data | table(join=outer, placeholder=True) }}"

    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"] = "Total"
    ws["B3"] = "{{ end_table | insert=above }}"
    for col in range(1, 3):
        ws.cell(3, col).font = Font(bold=True)
        ws.cell(3, col).fill = total_fill

    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A5"] = "Note First"
    ws.merge_cells("A5:B5")
    ws["A5"].font = Font(bold=True, color="FFFFFF")
    ws["A5"].fill = blue_fill

    orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    ws["A6"] = "Note Second"
    ws.merge_cells("A6:B6")
    ws["A6"].font = Font(italic=True)
    ws["A6"].fill = orange_fill

    wb.save(FIXTURES_DIR / "template_combo_outer_merges_below.xlsx")
    wb.close()
    print("  template_combo_outer_merges_below.xlsx")


# ---------------------------------------------------------------------------
# template_combo_left_with_merges.xlsx
# Left join (no row insertion) with two adjacent merges below — merges must
# remain completely untouched.
#
# Row 1: headers K / V1 / V2
# Row 2: "x", {{ data | table(join=left) }}
# Row 3: "y"
# Row 4: "z"
# Row 6: A6:C6 merged "Bottom Header" (bold, green fill)
# Row 7: A7:C7 merged "Bottom Note" (italic, gray fill)
#
# DF: exactly x/y/z → no insertion → merges at A6:C6 and A7:C7 unchanged.
# ---------------------------------------------------------------------------
def create_template_combo_left_with_merges():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "K"
    ws["B1"] = "V1"
    ws["C1"] = "V2"
    _header_style(ws, 1, range(1, 4))

    ws["A2"] = "x"
    ws["B2"] = "{{ data | table(join=left) }}"
    ws["A3"] = "y"
    ws["A4"] = "z"

    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws["A6"] = "Bottom Header"
    ws.merge_cells("A6:C6")
    ws["A6"].font = Font(bold=True, color="FFFFFF")
    ws["A6"].fill = green_fill

    gray_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
    ws["A7"] = "Bottom Note"
    ws.merge_cells("A7:C7")
    ws["A7"].font = Font(italic=True)
    ws["A7"].fill = gray_fill

    wb.save(FIXTURES_DIR / "template_combo_left_with_merges.xlsx")
    wb.close()
    print("  template_combo_left_with_merges.xlsx")


# ---------------------------------------------------------------------------
# template_combo_scalar_with_outer.xlsx
# Outer join (placeholder=True) with scalar cells below — scalars must shift
# to the correct row after table expansion.
#
# Row 1: headers Key / Data
# Row 2: blank Key, {{ tbl | table(join=outer, placeholder=True) }}
# Row 3: "Total", {{ end_table | insert=above }} (bold, yellow)
# Row 5: A5={{ title }} (bold, red fill), B5={{ summary }} (italic, blue fill)
#
# DF: a/b/c + Total → net +2 shift → scalars land at row 7.
# Tests: scalar values at shifted rows, raw tags gone, styles preserved
# ---------------------------------------------------------------------------
def create_template_combo_scalar_with_outer():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Key"
    ws["B1"] = "Data"
    _header_style(ws, 1, range(1, 3))

    ws["B2"] = "{{ tbl | table(join=outer, placeholder=True) }}"

    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"] = "Total"
    ws["B3"] = "{{ end_table | insert=above }}"
    for col in range(1, 3):
        ws.cell(3, col).font = Font(bold=True)
        ws.cell(3, col).fill = total_fill

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws["A5"] = "{{ title }}"
    ws["A5"].font = Font(bold=True, size=12)
    ws["A5"].fill = red_fill

    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["B5"] = "{{ summary }}"
    ws["B5"].font = Font(italic=True)
    ws["B5"].fill = blue_fill

    wb.save(FIXTURES_DIR / "template_combo_scalar_with_outer.xlsx")
    wb.close()
    print("  template_combo_scalar_with_outer.xlsx")


# ---------------------------------------------------------------------------
# template_combo_triple_adjacent_merges.xlsx
# PRIMARY regression test for the stale-merge-registry bug.
# Three adjacent same-span (A:D) merges below an outer-join table — ALL three
# must survive after the row-shift.  The bug caused the first one to be
# silently dropped when a stale large merge was in the registry.
#
# Row 1: headers Sec / A / B / C
# Row 2: blank, {{ tbl | table(join=outer, placeholder=True) }}
# Row 3: "Total", {{ end_table | insert=above }} (bold, yellow)
# Row 5: A5:D5 "Section One"   (bold, blue, medium border)
# Row 6: A6:D6 "Section Two"   (italic, orange, medium border)
# Row 7: A7:D7 "Section Three" (bold, green, medium border)
# Row 9: B9:C9 "Narrow Note"   (italic, red) — different column span
#
# DF: p/q/r + Total → net +2 shift → merges land at rows 7/8/9, narrow at 11.
# ---------------------------------------------------------------------------
def create_template_combo_triple_adjacent_merges():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Sec"
    ws["B1"] = "A"
    ws["C1"] = "B"
    ws["D1"] = "C"
    _header_style(ws, 1, range(1, 5))

    ws["B2"] = "{{ tbl | table(join=outer, placeholder=True) }}"

    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws["A3"] = "Total"
    ws["B3"] = "{{ end_table | insert=above }}"
    for col in range(1, 5):
        ws.cell(3, col).font = Font(bold=True)
        ws.cell(3, col).fill = total_fill

    med = Side(style="medium")
    border = Border(left=med, right=med, top=med, bottom=med)

    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A5"] = "Section One"
    ws.merge_cells("A5:D5")
    ws["A5"].font = Font(bold=True, color="FFFFFF")
    ws["A5"].fill = blue_fill
    ws["A5"].border = border

    orange_fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    ws["A6"] = "Section Two"
    ws.merge_cells("A6:D6")
    ws["A6"].font = Font(italic=True)
    ws["A6"].fill = orange_fill
    ws["A6"].border = border

    green_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws["A7"] = "Section Three"
    ws.merge_cells("A7:D7")
    ws["A7"].font = Font(bold=True, color="FFFFFF")
    ws["A7"].fill = green_fill
    ws["A7"].border = border

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    ws["B9"] = "Narrow Note"
    ws.merge_cells("B9:C9")
    ws["B9"].font = Font(italic=True, color="FFFFFF")
    ws["B9"].fill = red_fill

    wb.save(FIXTURES_DIR / "template_combo_triple_adjacent_merges.xlsx")
    wb.close()
    print("  template_combo_triple_adjacent_merges.xlsx")


# ---------------------------------------------------------------------------
# template_combo_two_outer_tables.xlsx
# Two stacked outer-join tables (each inserting 1 extra row), with an adjacent
# pair of merges below both.  Merges must shift by the cumulative +2.
#
# Row 1:  K1/V1 headers
# Row 2:  "a", {{ tbl1 | table(join=outer) }}
# Row 3:  "b"
# Row 4:  "Total1", {{ end_tbl1 | insert=above }} (bold, yellow)
# Row 6:  K2/V2 headers
# Row 7:  "x", {{ tbl2 | table(join=outer) }}
# Row 8:  "y"
# Row 9:  "Total2", {{ end_tbl2 | insert=above }} (bold, yellow)
# Row 11: A11:B11 "Grand Footer" (bold, blue, medium border)
# Row 12: A12:B12 "Sub-Footer"   (italic, gray)
#
# DF1: a/b/c + Total1 → +1 shift.  DF2: x/y/z + Total2 → +1 shift.
# Net +2 → Grand Footer at A13:B13, Sub-Footer at A14:B14.
# ---------------------------------------------------------------------------
def create_template_combo_two_outer_tables():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    total_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    med = Side(style="medium")
    border = Border(left=med, right=med, top=med, bottom=med)

    # First outer-join table (rows 1-4)
    ws["A1"] = "Key1"
    ws["B1"] = "V1"
    _header_style(ws, 1, range(1, 3))
    ws["A2"] = "a"
    ws["B2"] = "{{ tbl1 | table(join=outer) }}"
    ws["A3"] = "b"
    ws["A4"] = "Total1"
    ws["B4"] = "{{ end_table | insert=above }}"
    for col in range(1, 3):
        ws.cell(4, col).font = Font(bold=True)
        ws.cell(4, col).fill = total_fill

    # Second outer-join table (rows 6-9)
    ws["A6"] = "Key2"
    ws["B6"] = "V2"
    _header_style(ws, 6, range(1, 3))
    ws["A7"] = "x"
    ws["B7"] = "{{ tbl2 | table(join=outer) }}"
    ws["A8"] = "y"
    ws["A9"] = "Total2"
    ws["B9"] = "{{ end_table | insert=above }}"
    for col in range(1, 3):
        ws.cell(9, col).font = Font(bold=True)
        ws.cell(9, col).fill = total_fill

    # Adjacent merges below both tables (rows 11-12)
    blue_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws["A11"] = "Grand Footer"
    ws.merge_cells("A11:B11")
    ws["A11"].font = Font(bold=True, color="FFFFFF")
    ws["A11"].fill = blue_fill
    ws["A11"].border = border

    gray_fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
    ws["A12"] = "Sub-Footer"
    ws.merge_cells("A12:B12")
    ws["A12"].font = Font(italic=True)
    ws["A12"].fill = gray_fill

    wb.save(FIXTURES_DIR / "template_combo_two_outer_tables.xlsx")
    wb.close()
    print("  template_combo_two_outer_tables.xlsx")


# ---------------------------------------------------------------------------
# template_master.xlsx
# Master template for comprehensive border-preservation regression testing.
#
# The layout captures every scenario affected by the non-top-left MergedCell
# border-data loss bugs fixed in _copy_row_styles and _sync_merges_after_delete.
#
# Layout (Sheet1):
#   Rows  3–10  — Table 1: outer join with {{ insert_data }} + {{ end_table }}
#   Rows 12–14  — Table 2: outer join, placeholder=True, end_table|insert=above
#   Rows 17–18  — Styled Section 1: three overlapping merges with thick/thin borders
#                   B17:C17  horizontal merge  (C17 is non-TL MC)
#                   D17:D18  vertical merge    (D18 is non-TL MC)
#                   E17:E18  vertical merge    (E18 is non-TL MC)
#                   B18, C18 are separate real cells (bottom=thick)
#   Rows 21–27  — Styled Section 2: wide fill box + multi-row merge with scalar
#                   B21:H21  single-row wide merge  (C21–H21 are non-TL MCs)
#                   B22:H27  6-row × 7-col merge   (all non-TL positions are MCs)
#
# Reference input for tests:
#   my_table: DataFrame with index=[1,2,3,'Total',4], colA=[…], colB=[…]
#   some_value: 'X'
#
# Net row shift after both tables expand: +6
#   Section 1 lands at output rows 23–24
#   Section 2 lands at output rows 27–33
# ---------------------------------------------------------------------------
def create_template_master():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    thin = Side(style="thin")
    thick = Side(style="thick")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # --- Table 1 (rows 3–10): outer join with insert_data + end_table ---
    ws["B3"] = "index"
    ws["C3"] = "colA"
    ws["D3"] = "colB"

    ws["B4"] = "a"
    ws["C4"] = "{{ my_table | table(join=outer) }}"
    ws["B5"] = "b"
    ws["B6"] = "c"
    ws["C7"] = "{{ insert_data }}"
    ws["B8"] = "d"
    ws["B8"].font = Font(bold=True)
    ws["B9"] = "Total"
    ws["C9"] = "{{ end_table }}"
    ws["B9"].font = Font(bold=True)

    # --- Table 2 (rows 12–14): outer join / placeholder=True / end_table|insert=above ---
    ws["B12"] = "index"
    ws["C12"] = "colA"
    ws["D12"] = "colB"

    ws["C13"] = "{{ my_table | table(join=outer, placeholder=True, style=first) }}"
    ws["B14"] = "Total"
    ws["C14"] = "{{ end_table | insert=above }}"
    ws["B14"].font = Font(bold=True)
    for c in range(2, 5):
        ws.cell(14, c).fill = yellow

    # --- Styled Section 1 (rows 17–18): bordered header band ---
    # Top-left cells of each merge — borders must survive row-shift
    ws["B17"] = "headerColumn1"
    ws["B17"].font = Font(bold=True)
    ws["B17"].border = Border(left=thick, right=thin, top=thick, bottom=thin)
    ws["B17"].alignment = Alignment(horizontal="center")

    ws["D17"] = "MergedCol1"
    ws["D17"].font = Font(bold=True)
    ws["D17"].border = Border(left=thin, right=thin, top=thick, bottom=thin)
    ws["D17"].alignment = Alignment(horizontal="center")

    ws["E17"] = "MergedCol2"
    ws["E17"].font = Font(bold=True)
    ws["E17"].border = Border(left=thin, right=thick, top=thick, bottom=thin)
    ws["E17"].alignment = Alignment(horizontal="center")

    # Separate real cells in row 18 (NOT merged — their borders are independent)
    ws["B18"] = "subheaderColumn1"
    ws["B18"].font = Font(bold=True)
    ws["B18"].border = Border(left=thick, right=thin, top=thin, bottom=thick)

    ws["C18"] = "subheaderColumn2"
    ws["C18"].font = Font(bold=True)
    ws["C18"].border = Border(left=thin, right=thin, top=thin, bottom=thick)

    # Create section 1 merges — non-TL borders must be set AFTER merge_cells
    ws.merge_cells("B17:C17")  # C17 → MergedCell
    ws.merge_cells("D17:D18")  # D18 → MergedCell
    ws.merge_cells("E17:E18")  # E18 → MergedCell

    # C17: non-TL of B17:C17 — top=thick, bottom=thin, right=thin (no left)
    ws._cells.get((17, 3)).border = Border(top=thick, bottom=thin, right=thin)
    # D18: non-TL of D17:D18 — bottom=thin, left=thin, right=thin (no top)
    ws._cells.get((18, 4)).border = Border(bottom=thin, left=thin, right=thin)
    # E18: non-TL of E17:E18 — bottom=thin, left=thin, right=thick (no top)
    ws._cells.get((18, 5)).border = Border(bottom=thin, left=thin, right=thick)

    # --- Styled Section 2 (rows 21–27): wide fill box + multi-row merge ---
    blue_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    # B21: top-left of B21:H21 — thick-framed header with fill
    ws["B21"] = "some text here"
    ws["B21"].font = Font(bold=True, color="FFFFFF")
    ws["B21"].fill = blue_fill
    ws["B21"].border = Border(left=thick, right=thick, top=thick)
    ws["B21"].alignment = Alignment(horizontal="center")

    # B22: top-left of B22:H27 — contains scalar placeholder, thin border
    ws["B22"] = "{{ some_value }}"
    ws["B22"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws["B22"].alignment = Alignment(horizontal="center", vertical="center")

    # Create section 2 merges — non-TL borders must be set AFTER merge_cells
    ws.merge_cells("B21:H21")  # C21–H21 → MergedCells
    ws.merge_cells("B22:H27")  # all non-(22,2) positions → MergedCells

    # B21:H21 non-TL borders
    # C21–G21: top=thick only
    for col in range(3, 8):
        ws._cells.get((21, col)).border = Border(top=thick)
    # H21: top=thick, right=thick
    ws._cells.get((21, 8)).border = Border(top=thick, right=thick)

    # B22:H27 non-TL borders
    # C22–G22: top=thin only
    for col in range(3, 8):
        ws._cells.get((22, col)).border = Border(top=thin)
    # H22: top=thin, right=thin
    ws._cells.get((22, 8)).border = Border(top=thin, right=thin)
    # B23–B26: left=thin only  |  H23–H26: right=thin only
    for row in range(23, 27):
        ws._cells.get((row, 2)).border = Border(left=thin)
        ws._cells.get((row, 8)).border = Border(right=thin)
    # B27: bottom=thin, left=thin  |  C27–G27: bottom=thin  |  H27: bottom=thin, right=thin
    ws._cells.get((27, 2)).border = Border(bottom=thin, left=thin)
    for col in range(3, 8):
        ws._cells.get((27, col)).border = Border(bottom=thin)
    ws._cells.get((27, 8)).border = Border(bottom=thin, right=thin)

    wb.save(FIXTURES_DIR / "template_master.xlsx")
    wb.close()
    print("  template_master.xlsx")


if __name__ == "__main__":
    print("Creating fixtures in", FIXTURES_DIR)
    create_simple_table()
    create_no_headers()
    create_merged_cells()
    create_multiple_sheets()
    create_multiple_tables()
    create_empty_table()
    create_offset_table()
    create_template()
    create_template_merge_preservation()
    create_template_complex_merges()
    create_template_tight_footer()
    create_template_left_join()
    create_template_vertical_merge()
    create_template_vertical_merge_adjacent()
    create_template_data_col_vertical_merge()
    create_template_record()
    create_template_sorted_outer_asc()
    create_template_sorted_outer_desc()
    create_template_sorted_outer_fixed()
    create_template_sorted_outer_by_col()
    create_template_sorted_outer_shorter()
    create_template_sorted_outer_tmpl_rows()
    create_template_fill_global()
    create_template_fill_per_col()
    create_template_fill_lower_zone()
    create_template_fill_sorted_outer_lower_zone()
    create_anchored_cells()
    create_template_placeholder_outer()
    create_template_style_src_last()
    create_template_style_src_first()
    create_template_empty_outer_style_first()
    create_template_combo_outer_merges_below()
    create_template_combo_left_with_merges()
    create_template_combo_scalar_with_outer()
    create_template_combo_triple_adjacent_merges()
    create_template_combo_two_outer_tables()
    create_bug_on_insert()
    create_template_master()
    print("Done.")
