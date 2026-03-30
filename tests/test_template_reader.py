import pytest
from openpyxl import Workbook

from excel.template_reader import ExcelTemplateReader, MarkedCell
from excel.exceptions import TemplateReadError


# ==================== ExcelTemplateReader.read ====================


class TestRead:
    def test_tagged_sheets_present_in_result(self, template_path):
        reader = ExcelTemplateReader()
        result = reader.read(template_path)
        assert "Sheet1" in result
        assert "Sheet2" in result

    def test_untagged_sheet_excluded(self, template_path):
        reader = ExcelTemplateReader()
        result = reader.read(template_path)
        assert "EmptySheet" not in result

    def test_sheet1_tag_count(self, template_path):
        """Sheet1 has exactly three tagged cells."""
        reader = ExcelTemplateReader()
        result = reader.read(template_path)
        assert len(result["Sheet1"]) == 3

    def test_sheet2_tag_count(self, template_path):
        """Sheet2 has exactly one tagged cell."""
        reader = ExcelTemplateReader()
        result = reader.read(template_path)
        assert len(result["Sheet2"]) == 1

    def test_non_string_cells_ignored(self, template_path):
        """Numeric cells (e.g. 42) must not appear as tagged cells."""
        reader = ExcelTemplateReader()
        result = reader.read(template_path)
        names = [c.name for c in result["Sheet1"]]
        assert "42" not in names

    def test_plain_text_cells_ignored(self, template_path):
        """Cells without {{ }} must not appear in the result."""
        reader = ExcelTemplateReader()
        result = reader.read(template_path)
        names = [c.name for c in result["Sheet1"]]
        assert all("plain" not in n for n in names)

    def test_file_not_found_raises(self, tmp_path):
        reader = ExcelTemplateReader()
        with pytest.raises(TemplateReadError):
            reader.read(str(tmp_path / "missing.xlsx"))

    def test_corrupted_file_raises(self, tmp_path):
        bad = tmp_path / "bad.xlsx"
        bad.write_text("not an excel file")
        reader = ExcelTemplateReader()
        with pytest.raises(TemplateReadError):
            reader.read(str(bad))

    def test_empty_tag_name_raises(self, tmp_path):
        """{{ }} with no variable name must raise TemplateReadError."""
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "{{ }}"
        path = tmp_path / "bad_tag.xlsx"
        wb.save(str(path))
        reader = ExcelTemplateReader()
        with pytest.raises(TemplateReadError):
            reader.read(str(path))

    def test_workbook_with_no_tags_returns_empty_dict(self, tmp_path):
        wb = Workbook()
        wb.active["A1"] = "no tags here"
        path = tmp_path / "no_tags.xlsx"
        wb.save(str(path))
        result = ExcelTemplateReader().read(str(path))
        assert result == {}


# ==================== MarkedCell attributes ====================


class TestMarkedCellAttributes:
    def test_revenue_name(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "revenue")
        assert cell.name == "revenue"

    def test_revenue_cell_addr(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "revenue")
        assert cell.cell_addr == "B2"

    def test_revenue_sheet(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "revenue")
        assert cell.sheet == "Sheet1"

    def test_revenue_raw(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "revenue")
        assert cell.raw == "{{ revenue }}"

    def test_revenue_metadata_empty_string(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "revenue")
        assert cell.metadata == ""

    def test_title_metadata_raw_string(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "title")
        assert cell.metadata == "orientation=horizontal"

    def test_summary_on_sheet2(self, template_path):
        reader = ExcelTemplateReader()
        cell = reader.read(template_path)["Sheet2"][0]
        assert cell.name == "summary"
        assert cell.cell_addr == "A1"


# ==================== MarkedCell.parse_metadata ====================


class TestParseMetadata:
    def test_no_metadata_returns_empty_dict(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "revenue")
        assert cell.parse_metadata() == {}

    def test_string_value(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "title")
        meta = cell.parse_metadata()
        assert meta == {"orientation": "horizontal"}
        assert isinstance(meta["orientation"], str)

    def test_int_coercion(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "count")
        meta = cell.parse_metadata()
        assert meta["skip"] == 2
        assert isinstance(meta["skip"], int)

    def test_bool_true_coercion(self, template_path):
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "count")
        meta = cell.parse_metadata()
        assert meta["flag"] is True

    def test_bool_false_coercion(self):
        mc = MarkedCell(name="x", sheet="S", cell_addr="A1", raw="{{ x | active=False }}", metadata="active=False")
        assert mc.parse_metadata()["active"] is False

    def test_float_coercion(self):
        mc = MarkedCell(name="x", sheet="S", cell_addr="A1", raw="{{ x | ratio=1.5 }}", metadata="ratio=1.5")
        result = mc.parse_metadata()
        assert result["ratio"] == 1.5
        assert isinstance(result["ratio"], float)

    def test_multiple_key_value_pairs(self):
        mc = MarkedCell(
            name="x", sheet="S", cell_addr="A1",
            raw="{{ x | a=1, b=hello, c=True }}",
            metadata="a=1, b=hello, c=True",
        )
        assert mc.parse_metadata() == {"a": 1, "b": "hello", "c": True}

    def test_invalid_fragment_raises(self):
        mc = MarkedCell(name="x", sheet="S", cell_addr="A1", raw="{{ x | badkey }}", metadata="badkey")
        with pytest.raises(TemplateReadError):
            mc.parse_metadata()

    def test_parse_metadata_is_idempotent(self, template_path):
        """Calling parse_metadata() twice returns the same result."""
        reader = ExcelTemplateReader()
        cell = next(c for c in reader.read(template_path)["Sheet1"] if c.name == "count")
        assert cell.parse_metadata() == cell.parse_metadata()
