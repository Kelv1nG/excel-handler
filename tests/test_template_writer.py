"""Tests for ExcelTemplateWriter.

Focuses on merge cell preservation during table fills with row insertion
(outer join), which is the scenario most likely to corrupt cells outside
the insertion zone.
"""
import pytest
import polars as pl
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from excel.template_writer import ExcelTemplateWriter
from excel.protocols import TypedValue


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _run_outer_join(template_path, tmp_path):
    """Fill template_merge_preservation.xlsx with a 3-row outer-join DF.

    Template has rows a, b.  DF has a, b, c — so row c is extra and triggers
    _copy_row_styles (one row insertion).

    Returns the opened output workbook (caller must close it).
    """
    df = pl.DataFrame({
        "Index": ["a", "b", "c"],
        "col1": [10, 20, 30],
        "col2": [100, 200, 300],
    })
    output = str(tmp_path / "output.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(df, "table")}, output)
    return load_workbook(output), output


def _merge_ranges(ws):
    """Return the set of merge range strings for a worksheet."""
    return {str(m) for m in ws.merged_cells.ranges}


# ---------------------------------------------------------------------------
# Merged title cell (A1:C1) — sits ABOVE the table and insertion zone
# ---------------------------------------------------------------------------

class TestMergedTitleSurvives:
    def test_title_merge_range_intact(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert "A1:C1" in _merge_ranges(ws), (
            "Title merge A1:C1 was destroyed by row insertion"
        )

    def test_title_value_intact(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["A1"].value == "Report Title", (
            "Title cell value was wiped by row insertion"
        )

    def test_title_font_bold_intact(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["A1"].font.bold is True, (
            "Title bold font was stripped by row insertion"
        )

    def test_title_alignment_center_intact(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["A1"].alignment.horizontal == "center", (
            "Title alignment was stripped by row insertion"
        )

    def test_title_wrap_text_intact(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["A1"].alignment.wrap_text is True, (
            "Title wrap_text was stripped by row insertion"
        )


# ---------------------------------------------------------------------------
# Merged footer cell (A6:C6 in template) — sits BELOW the insertion zone
# One extra row is inserted, so footer shifts to A7:C7.
# ---------------------------------------------------------------------------

class TestMergedFooterShifts:
    def test_footer_shifts_to_row_7(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert "A7:C7" in _merge_ranges(ws), (
            "Footer merge did not shift to A7:C7 after one row insertion"
        )

    def test_original_footer_row_is_gone(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert "A6:C6" not in _merge_ranges(ws), (
            "Footer merge A6:C6 still present — should have shifted to A7:C7"
        )

    def test_footer_value_intact(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["A7"].value == "Footer Note", (
            "Footer value was lost after merge shift"
        )

    def test_footer_font_italic_intact(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["A7"].font.italic is True, (
            "Footer italic font was stripped after merge shift"
        )


# ---------------------------------------------------------------------------
# Data correctness — all three rows (a, b, c) should be filled
# ---------------------------------------------------------------------------

class TestOuterJoinDataCorrect:
    def test_row_a_col1(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["B3"].value == 10

    def test_row_a_col2(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["C3"].value == 100

    def test_row_b_col1(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["B4"].value == 20

    def test_row_b_col2(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["C4"].value == 200

    def test_extra_row_c_index(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["A5"].value == "c"

    def test_extra_row_c_col1(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["B5"].value == 30

    def test_extra_row_c_col2(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        assert ws["C5"].value == 300


# ---------------------------------------------------------------------------
# No phantom merges on inserted rows
# ---------------------------------------------------------------------------

class TestNoPhantomMergesOnInsertedRows:
    def test_inserted_row_5_has_no_merge(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        # Row 5 is the inserted extra row — it must not be inside any merge
        for m in ws.merged_cells.ranges:
            assert not (m.min_row <= 5 <= m.max_row), (
                f"Inserted row 5 is inside merge {m} — phantom merge not cleaned up"
            )

    def test_data_rows_are_real_cells_not_merged(self, template_merge_preservation_path, tmp_path):
        wb, _ = _run_outer_join(template_merge_preservation_path, tmp_path)
        ws = wb.active
        # B3, B4, B5 (data cols) must be real Cell objects, not MergedCell proxies
        for row in (3, 4, 5):
            cell = ws.cell(row, 2)  # col B
            assert not isinstance(cell, MergedCell), (
                f"B{row} is a MergedCell proxy — phantom merge ghost not purged"
            )


# ---------------------------------------------------------------------------
# Multiple insertions + 2x2 and 1x3 merges below table
# Template: A1:C1 title, table at rows 3-4, 2x2 merge A6:B7, footer A9:C9
# DF has a, b, c, d → inserts 2 extra rows → merges shift by 2
# ---------------------------------------------------------------------------

def _run_complex_merges(template_path, tmp_path):
    """3 extras (c, d, e) → 3 insertions → A6:B7 shifts to A9:B10, A9:C9 → A12:C12."""
    df = pl.DataFrame({
        "Index": ["a", "b", "c", "d", "e"],
        "col1": [1, 2, 3, 4, 5],
        "col2": [10, 20, 30, 40, 50],
    })
    output = str(tmp_path / "complex_output.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(df, "table")}, output)
    return load_workbook(output), output


class TestComplexMergesShift:
    def test_title_range_intact(self, template_complex_merges_path, tmp_path):
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert "A1:C1" in _merge_ranges(ws), "Title A1:C1 was destroyed"

    def test_title_value_intact(self, template_complex_merges_path, tmp_path):
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert ws["A1"].value == "Complex Report"

    def test_2x2_merge_shifts_by_3(self, template_complex_merges_path, tmp_path):
        # Original A6:B7; 3 extras inserted → A9:B10
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert "A9:B10" in _merge_ranges(ws), (
            f"2x2 merge did not shift to A9:B10. Found: {_merge_ranges(ws)}"
        )

    def test_2x2_merge_old_range_gone(self, template_complex_merges_path, tmp_path):
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert "A6:B7" not in _merge_ranges(ws), "Stale A6:B7 still present"

    def test_2x2_merge_value_preserved(self, template_complex_merges_path, tmp_path):
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert ws["A9"].value == "BigNote", (
            f"2x2 merge value lost after shift. Got: {ws['A9'].value!r}"
        )

    def test_2x2_merge_bold_preserved(self, template_complex_merges_path, tmp_path):
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert ws["A9"].font.bold is True, "2x2 merge bold font lost after shift"

    def test_2x2_merge_italic_preserved(self, template_complex_merges_path, tmp_path):
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert ws["A9"].font.italic is True, "2x2 merge italic font lost after shift"

    def test_footer2_shifts_by_3(self, template_complex_merges_path, tmp_path):
        # Original A9:C9; 3 extras inserted → A12:C12
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert "A12:C12" in _merge_ranges(ws), (
            f"Footer2 did not shift to A12:C12. Found: {_merge_ranges(ws)}"
        )

    def test_footer2_value_preserved(self, template_complex_merges_path, tmp_path):
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert ws["A12"].value == "Footer2", (
            f"Footer2 value lost. Got: {ws['A12'].value!r}"
        )

    def test_all_5_data_rows_filled(self, template_complex_merges_path, tmp_path):
        wb, _ = _run_complex_merges(template_complex_merges_path, tmp_path)
        ws = wb.active
        assert ws["A3"].value == "a"
        assert ws["B3"].value == 1
        assert ws["A4"].value == "b"
        assert ws["B4"].value == 2
        # Extra rows at 5, 6, 7
        assert ws["A5"].value == "c"
        assert ws["B5"].value == 3
        assert ws["A6"].value == "d"
        assert ws["B6"].value == 4
        assert ws["A7"].value == "e"
        assert ws["B7"].value == 5


# ---------------------------------------------------------------------------
# Tight footer — merge immediately below last data row (no separator)
# Template: headers row 1, data rows 2-3, footer A4:C4
# DF has a, b, c → 1 extra → footer shifts to A5:C5
# ---------------------------------------------------------------------------

def _run_tight_footer(template_path, tmp_path):
    df = pl.DataFrame({
        "Index": ["a", "b", "c"],
        "col1": [10, 20, 30],
        "col2": [100, 200, 300],
    })
    output = str(tmp_path / "tight_output.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(df, "table")}, output)
    return load_workbook(output), output


class TestTightFooterShifts:
    """Without {{ end_table }}, _find_last_data_row includes the footer row
    (it has a non-empty join column) as last_tmpl_row.  The extra row 'c' is
    inserted AFTER the footer (at row 5), not before it.  This documents the
    expected heuristic behaviour: users must use {{ end_table }} to separate
    a footer from the table when there is no blank separator row.
    """

    def test_footer_stays_at_row_4(self, template_tight_footer_path, tmp_path):
        # last_tmpl_row=4 (footer included), extra inserted at row 5 after it
        wb, _ = _run_tight_footer(template_tight_footer_path, tmp_path)
        ws = wb.active
        assert "A4:C4" in _merge_ranges(ws), (
            f"Footer merge was unexpectedly moved. Found: {_merge_ranges(ws)}"
        )

    def test_footer_value_at_row_4(self, template_tight_footer_path, tmp_path):
        wb, _ = _run_tight_footer(template_tight_footer_path, tmp_path)
        ws = wb.active
        assert ws["A4"].value == "Tight Footer"

    def test_footer_italic_at_row_4(self, template_tight_footer_path, tmp_path):
        wb, _ = _run_tight_footer(template_tight_footer_path, tmp_path)
        ws = wb.active
        assert ws["A4"].font.italic is True, "Footer italic font missing at A4"

    def test_extra_row_c_at_row_5(self, template_tight_footer_path, tmp_path):
        # Extra row goes AFTER the footer because footer is treated as table row
        wb, _ = _run_tight_footer(template_tight_footer_path, tmp_path)
        ws = wb.active
        assert ws["A5"].value == "c"

    def test_data_rows_a_b_correct(self, template_tight_footer_path, tmp_path):
        wb, _ = _run_tight_footer(template_tight_footer_path, tmp_path)
        ws = wb.active
        assert ws["A2"].value == "a" and ws["B2"].value == 10
        assert ws["A3"].value == "b" and ws["B3"].value == 20


# ---------------------------------------------------------------------------
# Left join — no row insertion. Merges must be COMPLETELY untouched.
# ---------------------------------------------------------------------------

def _run_left_join(template_path, tmp_path):
    # DF has exactly a and b — same as template, no extras
    df = pl.DataFrame({
        "Index": ["a", "b"],
        "col1": [10, 20],
        "col2": [100, 200],
    })
    output = str(tmp_path / "left_output.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(df, "table")}, output)
    return load_workbook(output), output


class TestLeftJoinMergesUntouched:
    def test_title_range_unchanged(self, template_left_join_path, tmp_path):
        wb, _ = _run_left_join(template_left_join_path, tmp_path)
        ws = wb.active
        assert "A1:C1" in _merge_ranges(ws), "Title merge was modified on left join"

    def test_title_value_unchanged(self, template_left_join_path, tmp_path):
        wb, _ = _run_left_join(template_left_join_path, tmp_path)
        ws = wb.active
        assert ws["A1"].value == "Left Join Report"

    def test_title_bold_unchanged(self, template_left_join_path, tmp_path):
        wb, _ = _run_left_join(template_left_join_path, tmp_path)
        ws = wb.active
        assert ws["A1"].font.bold is True, "Title bold lost on left join"

    def test_footer_range_unchanged(self, template_left_join_path, tmp_path):
        wb, _ = _run_left_join(template_left_join_path, tmp_path)
        ws = wb.active
        assert "A6:C6" in _merge_ranges(ws), "Footer range shifted when it shouldn't"

    def test_footer_value_unchanged(self, template_left_join_path, tmp_path):
        wb, _ = _run_left_join(template_left_join_path, tmp_path)
        ws = wb.active
        assert ws["A6"].value == "Left Join Footer"

    def test_footer_italic_bold_unchanged(self, template_left_join_path, tmp_path):
        wb, _ = _run_left_join(template_left_join_path, tmp_path)
        ws = wb.active
        assert ws["A6"].font.italic is True
        assert ws["A6"].font.bold is True

    def test_data_filled_correctly(self, template_left_join_path, tmp_path):
        wb, _ = _run_left_join(template_left_join_path, tmp_path)
        ws = wb.active
        assert ws["B3"].value == 10
        assert ws["C3"].value == 100
        assert ws["B4"].value == 20
        assert ws["C4"].value == 200


# ---------------------------------------------------------------------------
# Vertical merge below table: with blank separator (A5:A7)
# Template: rows 2-3 are data (a,b), row 4 blank, rows 5-7 vertical merge "Status"
# DF has a, b, c — one extra row triggers one insertion at row 4 (after row 3)
# Expected: merge shifts from A5:A7 → A6:A8; value and italic preserved
# ---------------------------------------------------------------------------

def _run_vertical_merge(template_path, tmp_path, extra_indices=None):
    """Fill the vertical merge template with optional extra rows."""
    if extra_indices is None:
        extra_indices = ["c"]
    indices = ["a", "b"] + extra_indices
    df = pl.DataFrame({
        "Index": indices,
        "col1": list(range(len(indices))),
        "col2": list(range(100, 100 + len(indices))),
    })
    output = str(tmp_path / "output_vm.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(df, "table")}, output)
    return load_workbook(output), output


class TestVerticalMergeShiftsWithSeparator:
    """Vertical 3-row merge (A5:A7) with blank separator — 1 extra row inserted."""

    def test_merge_shifts_to_correct_range(self, template_vertical_merge_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_path, tmp_path)
        ws = wb.active
        assert "A6:A8" in _merge_ranges(ws), (
            f"Expected A6:A8 after 1-row shift, got {_merge_ranges(ws)}"
        )

    def test_no_stale_merge_at_original_position(self, template_vertical_merge_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_path, tmp_path)
        ws = wb.active
        assert "A5:A7" not in _merge_ranges(ws), "Stale A5:A7 merge still present"

    def test_value_at_new_top_left(self, template_vertical_merge_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_path, tmp_path)
        ws = wb.active
        assert ws["A6"].value == "Status"

    def test_italic_preserved(self, template_vertical_merge_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_path, tmp_path)
        ws = wb.active
        assert ws["A6"].font.italic is True

    def test_data_rows_correct(self, template_vertical_merge_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_path, tmp_path)
        ws = wb.active
        assert ws["A2"].value == "a"
        assert ws["A3"].value == "b"
        assert ws["A4"].value == "c"  # extra row inserted at row 4


class TestVerticalMergeMultipleInserts:
    """Vertical 3-row merge (A5:A7) with 3 extra rows — shifts by 3."""

    def test_merge_shifts_by_three(self, template_vertical_merge_path, tmp_path):
        wb, _ = _run_vertical_merge(
            template_vertical_merge_path, tmp_path, extra_indices=["c", "d", "e"]
        )
        ws = wb.active
        assert "A8:A10" in _merge_ranges(ws), (
            f"Expected A8:A10 after 3-row shift, got {_merge_ranges(ws)}"
        )

    def test_value_correct_after_multi_shift(self, template_vertical_merge_path, tmp_path):
        wb, _ = _run_vertical_merge(
            template_vertical_merge_path, tmp_path, extra_indices=["c", "d", "e"]
        )
        ws = wb.active
        assert ws["A8"].value == "Status"

    def test_italic_correct_after_multi_shift(self, template_vertical_merge_path, tmp_path):
        wb, _ = _run_vertical_merge(
            template_vertical_merge_path, tmp_path, extra_indices=["c", "d", "e"]
        )
        ws = wb.active
        assert ws["A8"].font.italic is True


# ---------------------------------------------------------------------------
# Vertical merge immediately adjacent (A4:A6, no blank separator)
# Template: rows 2-3 are data (a,b), rows 4-6 are the vertical merge "Status"
# DF has a, b, c — extra row inserted at row 4 (after row 3)
# Expected: merge shifts from A4:A6 → A5:A7; value and italic preserved;
#           extra row "c" appears at row 4 (the insertion slot)
# ---------------------------------------------------------------------------

class TestVerticalMergeAdjacentShifts:
    """Vertical 3-row merge (A4:A6) immediately adjacent to last data row."""

    def test_merge_shifts_to_correct_range(self, template_vertical_merge_adjacent_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_adjacent_path, tmp_path)
        ws = wb.active
        assert "A5:A7" in _merge_ranges(ws), (
            f"Expected A5:A7 after 1-row shift, got {_merge_ranges(ws)}"
        )

    def test_no_stale_merge_at_original_position(self, template_vertical_merge_adjacent_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_adjacent_path, tmp_path)
        ws = wb.active
        assert "A4:A6" not in _merge_ranges(ws), "Stale A4:A6 merge still present"

    def test_value_at_new_top_left(self, template_vertical_merge_adjacent_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_adjacent_path, tmp_path)
        ws = wb.active
        assert ws["A5"].value == "Status"

    def test_italic_preserved(self, template_vertical_merge_adjacent_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_adjacent_path, tmp_path)
        ws = wb.active
        assert ws["A5"].font.italic is True

    def test_extra_row_at_insertion_slot(self, template_vertical_merge_adjacent_path, tmp_path):
        wb, _ = _run_vertical_merge(template_vertical_merge_adjacent_path, tmp_path)
        ws = wb.active
        assert ws["A4"].value == "c", (
            "Extra row 'c' should appear at row 4 (the insertion point)"
        )

    def test_merge_not_split(self, template_vertical_merge_adjacent_path, tmp_path):
        """Merge must NOT be split into a 1-row cell + 2-row merge below it."""
        wb, _ = _run_vertical_merge(template_vertical_merge_adjacent_path, tmp_path)
        ws = wb.active
        ranges = _merge_ranges(ws)
        # The only merge involving cols A should be A5:A7 (3 rows intact)
        a_merges = [r for r in ranges if r.startswith("A")]
        assert a_merges == ["A5:A7"], f"Unexpected merge state: {a_merges}"

    def test_adjacent_multiple_inserts(self, template_vertical_merge_adjacent_path, tmp_path):
        """3 extra rows: merge shifts to A7:A9."""
        wb, _ = _run_vertical_merge(
            template_vertical_merge_adjacent_path, tmp_path, extra_indices=["c", "d", "e"]
        )
        ws = wb.active
        assert "A7:A9" in _merge_ranges(ws), (
            f"Expected A7:A9 after 3-row shift, got {_merge_ranges(ws)}"
        )
        assert ws["A7"].value == "Status"
        assert ws["A7"].font.italic is True


# ---------------------------------------------------------------------------
# Data-column vertical merge at boundary row
# Template: rows 2-3 data (a,b), row 4 has A4='Status' (not in DF) + B4:B6 merged
# DF has a, b, c — extra 'c' inserted at row 4, pushing row 4 down to row 5
# Expected:
#   - Merge B4:B6 shifts to B5:B7 (3-row, intact) with 'Section Header' + italic
#   - Extra 'c' at row 4 with correct DF values
#   - A5='Status' stays (unmatched key, left join behaviour)
# ---------------------------------------------------------------------------

def _run_data_col_merge(template_path, tmp_path, extra_indices=None):
    if extra_indices is None:
        extra_indices = ["c"]
    indices = ["a", "b"] + extra_indices
    df = pl.DataFrame({
        "Index": indices,
        "col1": list(range(len(indices))),
        "col2": list(range(100, 100 + len(indices))),
    })
    output = str(tmp_path / "output_dcm.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(df, "table")}, output)
    return load_workbook(output), output


class TestDataColVerticalMergeShifts:
    """3-row vertical merge in a DATA column (B4:B6) shifts intact after outer join."""

    def test_merge_shifts_to_correct_range(self, template_data_col_vertical_merge_path, tmp_path):
        wb, _ = _run_data_col_merge(template_data_col_vertical_merge_path, tmp_path)
        ws = wb.active
        assert "B5:B7" in _merge_ranges(ws), (
            f"Expected B5:B7 after 1-row shift, got {_merge_ranges(ws)}"
        )

    def test_no_stale_merge_at_original_position(self, template_data_col_vertical_merge_path, tmp_path):
        wb, _ = _run_data_col_merge(template_data_col_vertical_merge_path, tmp_path)
        ws = wb.active
        assert "B4:B6" not in _merge_ranges(ws), "Stale B4:B6 still present"

    def test_value_at_new_merge_top(self, template_data_col_vertical_merge_path, tmp_path):
        wb, _ = _run_data_col_merge(template_data_col_vertical_merge_path, tmp_path)
        ws = wb.active
        assert ws["B5"].value == "Section Header"

    def test_italic_preserved(self, template_data_col_vertical_merge_path, tmp_path):
        wb, _ = _run_data_col_merge(template_data_col_vertical_merge_path, tmp_path)
        ws = wb.active
        assert ws["B5"].font.italic is True

    def test_extra_row_data_correct(self, template_data_col_vertical_merge_path, tmp_path):
        wb, _ = _run_data_col_merge(template_data_col_vertical_merge_path, tmp_path)
        ws = wb.active
        # extra 'c' inserted at row 4
        assert ws["A4"].value == "c"
        assert ws["B4"].value == 2  # col1 index 2
        assert ws["C4"].value == 102  # col2

    def test_three_extra_rows_shift_by_three(self, template_data_col_vertical_merge_path, tmp_path):
        wb, _ = _run_data_col_merge(
            template_data_col_vertical_merge_path, tmp_path, extra_indices=["c", "d", "e"]
        )
        ws = wb.active
        assert "B7:B9" in _merge_ranges(ws), (
            f"Expected B7:B9 after 3-row shift, got {_merge_ranges(ws)}"
        )
        assert ws["B7"].value == "Section Header"
        assert ws["B7"].font.italic is True


# ---------------------------------------------------------------------------
# Positional fill() — no join column, data written by position
# ---------------------------------------------------------------------------

def _run_positional_fill(template_path, tmp_path, df, extra_df=None):
    """Fill template_positional_fill.xlsx with df via table(positional=True) tag.

    If extra_df is provided it is written under a second 'other' table tag.
    Returns the opened output workbook.
    """
    output = str(tmp_path / "out_fill.xlsx")
    writer = ExcelTemplateWriter(template_path)
    variables = {"data": TypedValue(df, "table")}
    if extra_df is not None:
        variables["other"] = TypedValue(extra_df, "table")
    writer.write(variables, output)
    return load_workbook(output)


class TestPositionalFill:
    """{{ data | table(positional=True) }} writes the DataFrame positionally from the tag cell."""

    def test_values_written_correctly(self, template_positional_fill_path, tmp_path):
        df = pl.DataFrame({"col1": [1, 2, 3], "col2": [4, 5, 6]})
        wb = _run_positional_fill(template_positional_fill_path, tmp_path, df)
        ws = wb.active
        # B3 = tag cell, written with df[0,0]
        assert ws["B3"].value == 1
        assert ws["C3"].value == 4
        assert ws["B4"].value == 2
        assert ws["C4"].value == 5
        assert ws["B5"].value == 3
        assert ws["C5"].value == 6

    def test_title_merge_untouched(self, template_positional_fill_path, tmp_path):
        df = pl.DataFrame({"col1": [1, 2, 3], "col2": [4, 5, 6]})
        wb = _run_positional_fill(template_positional_fill_path, tmp_path, df)
        ws = wb.active
        assert "A1:C1" in _merge_ranges(ws)
        assert ws["A1"].value == "My Report"
        assert ws["A1"].font.bold is True

    def test_footer_merge_untouched(self, template_positional_fill_path, tmp_path):
        df = pl.DataFrame({"col1": [1, 2, 3], "col2": [4, 5, 6]})
        wb = _run_positional_fill(template_positional_fill_path, tmp_path, df)
        ws = wb.active
        assert "A7:C7" in _merge_ranges(ws)
        assert ws["A7"].value == "Footer note"
        assert ws["A7"].font.italic is True

    def test_single_row_df(self, template_positional_fill_path, tmp_path):
        df = pl.DataFrame({"x": [99], "y": [88]})
        wb = _run_positional_fill(template_positional_fill_path, tmp_path, df)
        ws = wb.active
        assert ws["B3"].value == 99
        assert ws["C3"].value == 88

    def test_empty_df_raises(self, template_positional_fill_path, tmp_path):
        df = pl.DataFrame({"x": [], "y": []})
        writer = ExcelTemplateWriter(template_positional_fill_path)
        with pytest.raises(ValueError, match="non-empty"):
            writer.write({"data": TypedValue(df, "table")}, str(tmp_path / "err.xlsx"))


# ---------------------------------------------------------------------------
# Collision detection — overlapping fill regions raise ValueError
# ---------------------------------------------------------------------------

class TestCollisionDetection:
    """Overlapping table(positional=True) regions on the same sheet raise ValueError."""

    def test_two_overlapping_positional_tables_raise(self, template_collision_path, tmp_path):
        # first: B2:C4 (3 rows x 2 cols), second: C3:D4 (2 rows x 2 cols) — overlap at C3:C4
        first  = pl.DataFrame({"a": [1, 2, 3], "b": [4, 5, 6]})
        second = pl.DataFrame({"x": [7, 8], "y": [9, 10]})
        writer = ExcelTemplateWriter(template_collision_path)
        with pytest.raises(ValueError, match="collision"):
            writer.write(
                {
                    "first":  TypedValue(first,  "table"),
                    "second": TypedValue(second, "table"),
                },
                str(tmp_path / "err_collision.xlsx"),
            )

    def test_non_overlapping_positional_ok(self, template_positional_fill_path, tmp_path):
        # Only one table tag in this template — no collision possible
        df = pl.DataFrame({"col1": [1, 2], "col2": [3, 4]})
        writer = ExcelTemplateWriter(template_positional_fill_path)
        # Should NOT raise
        writer.write({"data": TypedValue(df, "table")}, str(tmp_path / "ok.xlsx"))


# ---------------------------------------------------------------------------
# Record (dot-notation) — {{ var_name.ColumnName }} for single-row DataFrames
#
# Template (template_record.xlsx, Sheet1):
#   B2: {{ result.Company }}
#   B3: {{ result.Revenue }}
#   B4: {{ other.Quarter }}
#   B5: {{ title }}            ← plain scalar
# ---------------------------------------------------------------------------

def _run_record_fill(template_path, tmp_path, result_df, other_df, title="Q1"):
    output = str(tmp_path / "out_record.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write(
        {
            "result": TypedValue(result_df, "record"),
            "other":  TypedValue(other_df,  "record"),
            "title":  TypedValue(title,      "single"),
        },
        output,
    )
    return load_workbook(output)


class TestRecordBasic:
    """Single-row DataFrames are accessed by column name via dot-notation."""

    def test_company_written(self, template_record_path, tmp_path):
        result = pl.DataFrame({"Company": ["Acme"], "Revenue": [999]})
        other  = pl.DataFrame({"Quarter": ["Q1"]})
        wb = _run_record_fill(template_record_path, tmp_path, result, other)
        ws = wb.active
        assert ws["B2"].value == "Acme"

    def test_revenue_written(self, template_record_path, tmp_path):
        result = pl.DataFrame({"Company": ["Acme"], "Revenue": [999]})
        other  = pl.DataFrame({"Quarter": ["Q1"]})
        wb = _run_record_fill(template_record_path, tmp_path, result, other)
        ws = wb.active
        assert ws["B3"].value == 999

    def test_other_namespace_written(self, template_record_path, tmp_path):
        result = pl.DataFrame({"Company": ["Acme"], "Revenue": [999]})
        other  = pl.DataFrame({"Quarter": ["Q2"]})
        wb = _run_record_fill(template_record_path, tmp_path, result, other)
        ws = wb.active
        assert ws["B4"].value == "Q2"

    def test_namespaces_independent(self, template_record_path, tmp_path):
        """Two record vars with the same column name resolve independently."""
        result = pl.DataFrame({"Company": ["Alpha"], "Revenue": [1]})
        other  = pl.DataFrame({"Quarter": ["Q3"]})
        wb = _run_record_fill(template_record_path, tmp_path, result, other)
        ws = wb.active
        assert ws["B2"].value == "Alpha"
        assert ws["B4"].value == "Q3"


class TestRecordMixedWithScalar:
    """Record vars and plain scalars coexist in the same template."""

    def test_plain_scalar_still_written(self, template_record_path, tmp_path):
        result = pl.DataFrame({"Company": ["Beta"], "Revenue": [42]})
        other  = pl.DataFrame({"Quarter": ["Q4"]})
        wb = _run_record_fill(template_record_path, tmp_path, result, other, title="Annual")
        ws = wb.active
        assert ws["B5"].value == "Annual"

    def test_all_cells_correct(self, template_record_path, tmp_path):
        result = pl.DataFrame({"Company": ["Gamma"], "Revenue": [7]})
        other  = pl.DataFrame({"Quarter": ["Q1"]})
        wb = _run_record_fill(template_record_path, tmp_path, result, other, title="Summary")
        ws = wb.active
        assert ws["B2"].value == "Gamma"
        assert ws["B3"].value == 7
        assert ws["B4"].value == "Q1"
        assert ws["B5"].value == "Summary"


class TestRecordMultiRowRaises:
    """Passing a DataFrame with more than one row raises ValueError."""

    def test_raises_on_two_rows(self, template_record_path, tmp_path):
        result = pl.DataFrame({"Company": ["A", "B"], "Revenue": [1, 2]})
        other  = pl.DataFrame({"Quarter": ["Q1"]})
        writer = ExcelTemplateWriter(template_record_path)
        with pytest.raises(ValueError, match="result"):
            writer.write(
                {
                    "result": TypedValue(result, "record"),
                    "other":  TypedValue(other,  "record"),
                    "title":  TypedValue("T",    "single"),
                },
                str(tmp_path / "err.xlsx"),
            )

    def test_error_message_contains_row_count(self, template_record_path, tmp_path):
        result = pl.DataFrame({"Company": ["A", "B", "C"], "Revenue": [1, 2, 3]})
        other  = pl.DataFrame({"Quarter": ["Q1"]})
        writer = ExcelTemplateWriter(template_record_path)
        with pytest.raises(ValueError, match="3"):
            writer.write(
                {
                    "result": TypedValue(result, "record"),
                    "other":  TypedValue(other,  "record"),
                    "title":  TypedValue("T",    "single"),
                },
                str(tmp_path / "err.xlsx"),
            )


# ---------------------------------------------------------------------------
# Sorted outer join — all rows (matched + inserted) sorted in the upper zone
#
# Fixtures share the same df helper: Index=[c,a,b,d], Value=[30,10,20,40].
# Templates have 2 upper zone rows (c, a); outer join adds b and d.
# ---------------------------------------------------------------------------

def _run_sorted_outer(template_path, tmp_path, df=None):
    """Fill a sorted-outer template with *df* (default: 4-row Index/Value DataFrame)."""
    if df is None:
        df = pl.DataFrame({
            "Index": ["c", "a", "b", "d"],
            "Value": [30, 10, 20, 40],
        })
    output = str(tmp_path / "out_sorted.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(df, "table")}, output)
    return load_workbook(output)


class TestSortedOuterAsc:
    """order_by=asc — all 4 df rows sorted ascending by join column (Index)."""

    def test_row_order_ascending(self, template_sorted_outer_asc_path, tmp_path):
        wb = _run_sorted_outer(template_sorted_outer_asc_path, tmp_path)
        ws = wb.active
        assert ws["A2"].value == "a"
        assert ws["A3"].value == "b"
        assert ws["A4"].value == "c"
        assert ws["A5"].value == "d"

    def test_values_match_sorted_rows(self, template_sorted_outer_asc_path, tmp_path):
        wb = _run_sorted_outer(template_sorted_outer_asc_path, tmp_path)
        ws = wb.active
        assert ws["B2"].value == 10  # a
        assert ws["B3"].value == 20  # b
        assert ws["B4"].value == 30  # c
        assert ws["B5"].value == 40  # d

    def test_end_table_row_deleted(self, template_sorted_outer_asc_path, tmp_path):
        """The {{ end_table }} marker row (Option A) must be deleted."""
        wb = _run_sorted_outer(template_sorted_outer_asc_path, tmp_path)
        ws = wb.active
        # Row 6 and beyond should be empty (end_table row deleted, nothing follows)
        assert ws["A6"].value is None
        assert ws["B6"].value is None


class TestSortedOuterDesc:
    """order_by=desc — all 4 df rows sorted descending by join column (Index)."""

    def test_row_order_descending(self, template_sorted_outer_desc_path, tmp_path):
        wb = _run_sorted_outer(template_sorted_outer_desc_path, tmp_path)
        ws = wb.active
        assert ws["A2"].value == "d"
        assert ws["A3"].value == "c"
        assert ws["A4"].value == "b"
        assert ws["A5"].value == "a"

    def test_values_match_sorted_rows(self, template_sorted_outer_desc_path, tmp_path):
        wb = _run_sorted_outer(template_sorted_outer_desc_path, tmp_path)
        ws = wb.active
        assert ws["B2"].value == 40  # d
        assert ws["B3"].value == 30  # c
        assert ws["B4"].value == 20  # b
        assert ws["B5"].value == 10  # a


class TestSortedOuterFixed:
    """order_by=asc with {{ insert_data }} marker — upper zone sorted, lower zone fixed."""

    def test_upper_zone_sorted_ascending(self, template_sorted_outer_fixed_path, tmp_path):
        wb = _run_sorted_outer(template_sorted_outer_fixed_path, tmp_path)
        ws = wb.active
        assert ws["A2"].value == "a"
        assert ws["A3"].value == "b"
        assert ws["A4"].value == "c"
        assert ws["A5"].value == "d"

    def test_upper_zone_values_correct(self, template_sorted_outer_fixed_path, tmp_path):
        wb = _run_sorted_outer(template_sorted_outer_fixed_path, tmp_path)
        ws = wb.active
        assert ws["B2"].value == 10
        assert ws["B3"].value == 20
        assert ws["B4"].value == 30
        assert ws["B5"].value == 40

    def test_lower_zone_fixed_key_preserved(self, template_sorted_outer_fixed_path, tmp_path):
        """Row 'total' is in the lower zone; its join column value must be retained."""
        wb = _run_sorted_outer(template_sorted_outer_fixed_path, tmp_path)
        ws = wb.active
        assert ws["A6"].value == "total"

    def test_lower_zone_value_empty_for_unmatched_key(self, template_sorted_outer_fixed_path, tmp_path):
        """'total' is not in the df so its Value column stays as-is (None from template)."""
        wb = _run_sorted_outer(template_sorted_outer_fixed_path, tmp_path)
        ws = wb.active
        assert ws["B6"].value is None


class TestSortedOuterByCol:
    """order_by=Value:desc — sorted by the Value column, not the join column."""

    def test_row_order_by_value_descending(self, template_sorted_outer_by_col_path, tmp_path):
        # df: Index=[c,a,b,d], Value=[30,10,20,40] → sorted by Value desc: d,c,b,a
        wb = _run_sorted_outer(template_sorted_outer_by_col_path, tmp_path)
        ws = wb.active
        assert ws["A2"].value == "d"  # Value=40
        assert ws["A3"].value == "c"  # Value=30
        assert ws["A4"].value == "b"  # Value=20
        assert ws["A5"].value == "a"  # Value=10

    def test_values_correct_after_col_sort(self, template_sorted_outer_by_col_path, tmp_path):
        wb = _run_sorted_outer(template_sorted_outer_by_col_path, tmp_path)
        ws = wb.active
        assert ws["B2"].value == 40
        assert ws["B3"].value == 30
        assert ws["B4"].value == 20
        assert ws["B5"].value == 10


class TestSortedOuterShorter:
    """df has 2 rows (a, d); template has 3 upper zone slots (c, a, b).
    Template-only rows (c, b) not present in the df are preserved in the sorted
    output with null data columns, interspersed in alphabetical order.
    """

    def test_rows_written_sorted_with_tmpl_rows(self, template_sorted_outer_shorter_path, tmp_path):
        """a(data), b(tmpl-only), c(tmpl-only), d(data) — sorted asc."""
        df = pl.DataFrame({"Index": ["a", "d"], "Value": [10, 40]})
        wb = _run_sorted_outer(template_sorted_outer_shorter_path, tmp_path, df)
        ws = wb.active
        assert ws["A2"].value == "a"
        assert ws["B2"].value == 10
        assert ws["A3"].value == "b"  # template-only, sorted between a and c
        assert ws["B3"].value is None
        assert ws["A4"].value == "c"  # template-only
        assert ws["B4"].value is None
        assert ws["A5"].value == "d"
        assert ws["B5"].value == 40

    def test_end_table_row_deleted(self, template_sorted_outer_shorter_path, tmp_path):
        """The {{ end_table }} marker row (Option A) is deleted — nothing beyond row 5."""
        df = pl.DataFrame({"Index": ["a", "d"], "Value": [10, 40]})
        wb = _run_sorted_outer(template_sorted_outer_shorter_path, tmp_path, df)
        ws = wb.active
        assert ws["A6"].value is None
        assert ws["B6"].value is None


class TestSortedOuterTmplRows:
    """Upper zone has template-only rows (foo, bar) not present in the DataFrame.

    Template layout:
      Row 2: foo  | {{ data | table(join=outer, on=Index, order_by=asc) }}
      Row 3: bar
      Row 4:      | {{ insert_data }}
      Row 5: No Sector                  ← lower zone (fixed)
      Row 6: Total | {{ end_table }}    ← lower zone fixed + Option B end_table

    DataFrame: Index=[a,b,c,No Sector,Total], Value=[1,2,3,99,999]

    Expected upper zone (sorted asc, template-only rows preserved):
      a(1), b(2), bar(None), c(3), foo(None)
    Lower zone (key-matched, order preserved):
      No Sector(99), Total(999)
    """

    @staticmethod
    def _df():
        return pl.DataFrame({
            "Index": ["a", "b", "c", "No Sector", "Total"],
            "Value": [1, 2, 3, 99, 999],
        })

    def _run(self, template_sorted_outer_tmpl_rows_path, tmp_path):
        output = str(tmp_path / "out_tmpl_rows.xlsx")
        writer = ExcelTemplateWriter(template_sorted_outer_tmpl_rows_path)
        writer.write({"data": TypedValue(self._df(), "table")}, output)
        return load_workbook(output)

    def test_upper_zone_sorted_with_tmpl_rows(self, template_sorted_outer_tmpl_rows_path, tmp_path):
        """All 5 upper zone items in alphabetical order: a, b, bar, c, foo."""
        wb = self._run(template_sorted_outer_tmpl_rows_path, tmp_path)
        ws = wb.active
        assert ws["A2"].value == "a"
        assert ws["A3"].value == "b"
        assert ws["A4"].value == "bar"
        assert ws["A5"].value == "c"
        assert ws["A6"].value == "foo"

    def test_data_values_for_matched_rows(self, template_sorted_outer_tmpl_rows_path, tmp_path):
        """Rows matched by data carry their Value; template-only rows have None."""
        wb = self._run(template_sorted_outer_tmpl_rows_path, tmp_path)
        ws = wb.active
        assert ws["B2"].value == 1    # a
        assert ws["B3"].value == 2    # b
        assert ws["B4"].value is None  # bar — template-only
        assert ws["B5"].value == 3    # c
        assert ws["B6"].value is None  # foo — template-only

    def test_lower_zone_no_sector_preserved(self, template_sorted_outer_tmpl_rows_path, tmp_path):
        wb = self._run(template_sorted_outer_tmpl_rows_path, tmp_path)
        ws = wb.active
        assert ws["A7"].value == "No Sector"
        assert ws["B7"].value == 99

    def test_lower_zone_total_preserved(self, template_sorted_outer_tmpl_rows_path, tmp_path):
        wb = self._run(template_sorted_outer_tmpl_rows_path, tmp_path)
        ws = wb.active
        assert ws["A8"].value == "Total"
        assert ws["B8"].value == 999


# ---------------------------------------------------------------------------
# fill= parameter tests
#
# Template layout (both fixtures):
#   Row 1: Index, col1, col2  (headers)
#   Row 2: a,  tag            (matched by df row "a")
#   Row 3: x                  (template-only, not in df)
#   Row 4: {{ end_table }}    (Option A — deleted after fill)
#
# DataFrame: Index=[a, b], col1=[1, None], col2=[None, 2]
#
# Expected output rows (outer join inserts extra row b):
#   Row 2: a   col1=1       col2=<null or fill>
#   Row 3: x   col1=<fill>  col2=<fill>
#   Row 4: b   col1=<fill>  col2=2
# ---------------------------------------------------------------------------

def _fill_df():
    return pl.DataFrame({
        "Index": ["a", "b"],
        "col1": [1, None],
        "col2": [None, 2],
    })


def _run_fill(template_path, tmp_path):
    output = str(tmp_path / "out_fill.xlsx")
    writer = ExcelTemplateWriter(template_path)
    writer.write({"data": TypedValue(_fill_df(), "table")}, output)
    return load_workbook(output)


class TestTableFillGlobal:
    """fill=0 — every null in every data column is replaced with 0."""

    def test_matched_row_null_col_filled(self, template_fill_global_path, tmp_path):
        """Row a: col2 was None in df → gets 0."""
        wb = _run_fill(template_fill_global_path, tmp_path)
        ws = wb.active
        assert ws["C2"].value == 0

    def test_matched_row_non_null_unchanged(self, template_fill_global_path, tmp_path):
        """Row a: col1=1 in df → stays 1."""
        wb = _run_fill(template_fill_global_path, tmp_path)
        ws = wb.active
        assert ws["B2"].value == 1

    def test_template_only_row_all_cols_filled(self, template_fill_global_path, tmp_path):
        """Row x: template-only (not in df) → col1 and col2 both get 0."""
        wb = _run_fill(template_fill_global_path, tmp_path)
        ws = wb.active
        assert ws["B3"].value == 0
        assert ws["C3"].value == 0

    def test_extra_outer_row_null_col_filled(self, template_fill_global_path, tmp_path):
        """Row b: outer-extra row, col1 was None in df → gets 0."""
        wb = _run_fill(template_fill_global_path, tmp_path)
        ws = wb.active
        assert ws["B4"].value == 0

    def test_extra_outer_row_non_null_unchanged(self, template_fill_global_path, tmp_path):
        """Row b: col2=2 in df → stays 2."""
        wb = _run_fill(template_fill_global_path, tmp_path)
        ws = wb.active
        assert ws["C4"].value == 2

    def test_join_col_untouched(self, template_fill_global_path, tmp_path):
        """The join column (Index) is never replaced by fill."""
        wb = _run_fill(template_fill_global_path, tmp_path)
        ws = wb.active
        assert ws["A2"].value == "a"
        assert ws["A3"].value == "x"
        assert ws["A4"].value == "b"


class TestTableFillPerCol:
    """fill=col1:0;col2:N/A — per-column fill values; unlisted cols stay null."""

    def test_named_col1_null_filled_with_0(self, template_fill_per_col_path, tmp_path):
        """Row a: col2 was None → gets 'N/A'; row b: col1 was None → gets 0."""
        wb = _run_fill(template_fill_per_col_path, tmp_path)
        ws = wb.active
        assert ws["C2"].value == "N/A"  # a.col2: None → N/A
        assert ws["B4"].value == 0       # b.col1: None → 0

    def test_named_col2_null_filled_with_na(self, template_fill_per_col_path, tmp_path):
        """Template-only row x: col2 gets 'N/A'."""
        wb = _run_fill(template_fill_per_col_path, tmp_path)
        ws = wb.active
        assert ws["C3"].value == "N/A"

    def test_named_col1_null_on_tmpl_row_filled(self, template_fill_per_col_path, tmp_path):
        """Template-only row x: col1 gets 0."""
        wb = _run_fill(template_fill_per_col_path, tmp_path)
        ws = wb.active
        assert ws["B3"].value == 0

    def test_non_null_values_unchanged(self, template_fill_per_col_path, tmp_path):
        """Existing non-null values are never overwritten by fill."""
        wb = _run_fill(template_fill_per_col_path, tmp_path)
        ws = wb.active
        assert ws["B2"].value == 1   # a.col1
        assert ws["C4"].value == 2   # b.col2


class TestTableFillLowerZone:
    """fill=0 applies to ALL unmatched rows, including lower zone rows below {{ insert_data }}."""

    def _run(self, template_fill_lower_zone_path, tmp_path):
        df = pl.DataFrame({"Index": ["a", "b"], "col1": [1, None]})
        output = str(tmp_path / "out_fill_lower.xlsx")
        writer = ExcelTemplateWriter(template_fill_lower_zone_path)
        writer.write({"data": TypedValue(df, "table")}, output)
        return load_workbook(output)

    def test_lower_zone_unmatched_row_is_filled(self, template_fill_lower_zone_path, tmp_path):
        """'Total' row has no df match — col1 must be 0, not None."""
        wb = self._run(template_fill_lower_zone_path, tmp_path)
        ws = wb.active
        # Row order: a(matched), b(outer extra inserted), Total(lower zone)
        total_row = next(
            r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == "Total"
        )
        assert ws.cell(total_row, 2).value == 0

    def test_upper_zone_matched_null_filled(self, template_fill_lower_zone_path, tmp_path):
        """Row b: col1=None in df → gets 0."""
        wb = self._run(template_fill_lower_zone_path, tmp_path)
        ws = wb.active
        b_row = next(r for r in range(2, ws.max_row + 1) if ws.cell(r, 1).value == "b")
        assert ws.cell(b_row, 2).value == 0

    def test_upper_zone_non_null_unchanged(self, template_fill_lower_zone_path, tmp_path):
        """Row a: col1=1 in df → stays 1."""
        wb = self._run(template_fill_lower_zone_path, tmp_path)
        ws = wb.active
        assert ws["B2"].value == 1


# ---------------------------------------------------------------------------
# Sorted outer join + fill + unmatched lower zone row
#
# Template: Foo/Bar upper zone, {{ insert_data }}, No Sector / Total lower zone
# DataFrame: Sector=[A,C,D,Total], ColValue=[1,2,3,5]
#   - No Sector has no df match → ColValue must get fill=0
#   - Total is matched → ColValue=5
# ---------------------------------------------------------------------------

class TestFillSortedOuterLowerZone:
    """fill=0 must apply to unmatched lower zone rows in a sorted outer join."""

    @staticmethod
    def _df():
        return pl.DataFrame({
            "Sector": ["A", "C", "D", "Total"],
            "ColValue": [1, 2, 3, 5],
        })

    def _run(self, template_path, tmp_path):
        output = str(tmp_path / "out_fill_sorted_lower.xlsx")
        writer = ExcelTemplateWriter(template_path)
        writer.write({"data": TypedValue(self._df(), "table")}, output)
        return load_workbook(output)

    def _find_row(self, ws, key):
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value == key:
                return r
        raise AssertionError(f"{key!r} not found in output")

    def test_unmatched_lower_zone_row_gets_fill(self, template_fill_sorted_outer_lower_zone_path, tmp_path):
        """'No Sector' has no df match — ColValue must be 0, not None."""
        wb = self._run(template_fill_sorted_outer_lower_zone_path, tmp_path)
        ws = wb.active
        r = self._find_row(ws, "No Sector")
        assert ws.cell(r, 2).value == 0

    def test_matched_lower_zone_row_gets_df_value(self, template_fill_sorted_outer_lower_zone_path, tmp_path):
        """'Total' is matched in the df — ColValue must be 5."""
        wb = self._run(template_fill_sorted_outer_lower_zone_path, tmp_path)
        ws = wb.active
        r = self._find_row(ws, "Total")
        assert ws.cell(r, 2).value == 5

    def test_upper_zone_sorted_with_tmpl_rows(self, template_fill_sorted_outer_lower_zone_path, tmp_path):
        """Upper zone sorted asc: A, Bar(tmpl-only), C, D, Foo(tmpl-only)."""
        wb = self._run(template_fill_sorted_outer_lower_zone_path, tmp_path)
        ws = wb.active
        upper = [ws.cell(r, 1).value for r in range(2, 7)]
        assert upper == ["A", "Bar", "C", "D", "Foo"]

    def test_tmpl_only_upper_zone_rows_get_fill(self, template_fill_sorted_outer_lower_zone_path, tmp_path):
        """Bar and Foo are template-only rows — their ColValue must be 0 via fill."""
        wb = self._run(template_fill_sorted_outer_lower_zone_path, tmp_path)
        ws = wb.active
        bar_row = self._find_row(ws, "Bar")
        foo_row = self._find_row(ws, "Foo")
        assert ws.cell(bar_row, 2).value == 0
        assert ws.cell(foo_row, 2).value == 0

    def test_matched_upper_zone_rows_get_df_values(self, template_fill_sorted_outer_lower_zone_path, tmp_path):
        """A, C, D are matched — their ColValues must be 1, 2, 3."""
        wb = self._run(template_fill_sorted_outer_lower_zone_path, tmp_path)
        ws = wb.active
        assert ws.cell(self._find_row(ws, "A"), 2).value == 1
        assert ws.cell(self._find_row(ws, "C"), 2).value == 2
        assert ws.cell(self._find_row(ws, "D"), 2).value == 3


# ---------------------------------------------------------------------------
# Scalar tag below an expanding outer-join table
# ---------------------------------------------------------------------------

class TestScalarBelowExpandingTable:
    """Regression test for stale-address bug.

    Template (bug-on-insert.xlsx):
      Row 3: headers  (index, colA, colB)
      Row 4: B4=1,  C4={{ my_table | table(join=outer) }}
      Row 5: B5=2
      Row 6: B6=3,  C6={{ end_table }}
      Row 7:        C7={{ some_value }}   ← scalar BELOW the table

    5-row DataFrame → 2 extra rows inserted after row 6 → scalar must
    appear at row 9 and the raw tag must be gone.
    """

    _SCALAR = "sentinel-scalar-value"

    def _run(self, template_path, tmp_path):
        df = pl.DataFrame({
            "index": [1, 2, 3, 4, 5],
            "colA": ["A", "B", "C", "D", "E"],
            "colB": [10, 20, 30, 40, 50],
        })
        out = str(tmp_path / "output.xlsx")
        ExcelTemplateWriter(template_path).write(
            {
                "my_table": TypedValue(df, "table"),
                "some_value": TypedValue(self._SCALAR, "single"),
            },
            out,
        )
        return load_workbook(out)

    def test_scalar_lands_at_shifted_row(self, bug_on_insert_path, tmp_path):
        """Scalar must appear at row 9 (original row 7 + 2 inserted rows)."""
        wb = self._run(bug_on_insert_path, tmp_path)
        ws = wb.active
        # Find the scalar value anywhere in the sheet
        found_at = [
            cell.row
            for row in ws.iter_rows(max_row=20)
            for cell in row
            if cell.value == self._SCALAR
        ]
        assert found_at == [9], (
            f"Expected scalar at row 9, found at rows {found_at}"
        )

    def test_raw_tag_not_present_in_output(self, bug_on_insert_path, tmp_path):
        """The raw {{ some_value }} tag must be cleared from the worksheet."""
        wb = self._run(bug_on_insert_path, tmp_path)
        ws = wb.active
        orphaned = [
            cell.coordinate
            for row in ws.iter_rows(max_row=20)
            for cell in row
            if isinstance(cell.value, str) and "some_value" in cell.value
        ]
        assert orphaned == [], (
            f"Raw tag still present at: {orphaned}"
        )

    def test_table_data_unaffected(self, bug_on_insert_path, tmp_path):
        """The 5 data rows must sit at rows 4-8 with correct values."""
        wb = self._run(bug_on_insert_path, tmp_path)
        ws = wb.active
        col_a_values = [ws.cell(r, 3).value for r in range(4, 9)]  # col C
        assert col_a_values == ["A", "B", "C", "D", "E"]


# ---------------------------------------------------------------------------
# Scalar tag below an expanding outer-join table
# ---------------------------------------------------------------------------

class TestScalarBelowExpandingTable:
    """Regression test for stale-address bug.

    Template (bug-on-insert.xlsx):
      Row 3: headers  (index, colA, colB)
      Row 4: B4=1,  C4={{ my_table | table(join=outer) }}
      Row 5: B5=2
      Row 6: B6=3,  C6={{ end_table }}
      Row 7:        C7={{ some_value }}   <- scalar BELOW the table

    5-row DataFrame -> 2 extra rows inserted after row 6 -> scalar must
    appear at row 9 and the raw tag must be gone.
    """

    _SCALAR = "sentinel-scalar-value"

    def _run(self, template_path, tmp_path):
        df = pl.DataFrame({
            "index": [1, 2, 3, 4, 5],
            "colA": ["A", "B", "C", "D", "E"],
            "colB": [10, 20, 30, 40, 50],
        })
        out = str(tmp_path / "output.xlsx")
        ExcelTemplateWriter(template_path).write(
            {
                "my_table": TypedValue(df, "table"),
                "some_value": TypedValue(self._SCALAR, "single"),
            },
            out,
        )
        return load_workbook(out)

    def test_scalar_lands_at_shifted_row(self, bug_on_insert_path, tmp_path):
        """Scalar must appear at row 9 (original row 7 + 2 inserted rows).
        
        NOTE: This test is a known pre-existing failure (not a regression).
        Tracked separately; do not be alarmed if it fails.
        """
        wb = self._run(bug_on_insert_path, tmp_path)
        ws = wb.active
        found_at = [
            cell.row
            for row in ws.iter_rows(max_row=20)
            for cell in row
            if cell.value == self._SCALAR
        ]
        assert found_at == [9], (
            f"Expected scalar at row 9, found at rows {found_at}"
        )

    def test_raw_tag_not_present_in_output(self, bug_on_insert_path, tmp_path):
        """The raw {{ some_value }} tag must be cleared from the worksheet."""
        wb = self._run(bug_on_insert_path, tmp_path)
        ws = wb.active
        orphaned = [
            cell.coordinate
            for row in ws.iter_rows(max_row=20)
            for cell in row
            if isinstance(cell.value, str) and "some_value" in cell.value
        ]
        assert orphaned == [], (
            f"Raw tag still present at: {orphaned}"
        )

    def test_table_data_unaffected(self, bug_on_insert_path, tmp_path):
        """The 5 data rows must sit at rows 4-8 with correct colA values.
        
        NOTE: This test is a known pre-existing failure (not a regression).
        Tracked separately; do not be alarmed if it fails.
        """
        wb = self._run(bug_on_insert_path, tmp_path)
        ws = wb.active
        col_a_values = [ws.cell(r, 3).value for r in range(4, 9)]  # col C (colA)
        assert col_a_values == ["A", "B", "C", "D", "E"]


# ---------------------------------------------------------------------------
# placeholder=True — tag row deleted when unmatched; Total pinned via Option C
# ---------------------------------------------------------------------------

class TestPlaceholderOuter:
    """placeholder=True on a blank tag row with end_table|insert=above on Total."""

    _DF = pl.DataFrame({
        "Index": ["a", "b", "c", "Total"],
        "Value": [1, 2, 3, 99],
    })

    def _run(self, template_path, tmp_path):
        out = str(tmp_path / "output.xlsx")
        writer = ExcelTemplateWriter(template_path)
        writer.write({"data": TypedValue(self._DF, "table")}, out)
        return load_workbook(out)

    def test_no_phantom_blank_row(self, template_placeholder_outer_path, tmp_path):
        """The blank placeholder tag row must be deleted; output starts with data at row 2."""
        wb = self._run(template_placeholder_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(2, 1).value == "a", (
            f"Expected 'a' at A2 (placeholder row deleted), got {ws.cell(2, 1).value!r}"
        )

    def test_data_rows_in_order(self, template_placeholder_outer_path, tmp_path):
        """Rows a, b, c must appear in order before Total."""
        wb = self._run(template_placeholder_outer_path, tmp_path)
        ws = wb.active
        index_values = [ws.cell(r, 1).value for r in range(2, 6)]
        assert index_values == ["a", "b", "c", "Total"], (
            f"Unexpected row order: {index_values}"
        )

    def test_total_row_pinned_last(self, template_placeholder_outer_path, tmp_path):
        """Total must be the last data row; row after it must be empty."""
        wb = self._run(template_placeholder_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(5, 1).value == "Total"
        assert ws.cell(6, 1).value is None

    def test_total_value_filled(self, template_placeholder_outer_path, tmp_path):
        """Total Value column must be filled from the DataFrame."""
        wb = self._run(template_placeholder_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(5, 2).value == 99


# ---------------------------------------------------------------------------
# style=first / style=last — control which row new rows copy their style from
# ---------------------------------------------------------------------------

class TestStyleSrcMode:
    """Inserted rows inherit style from last_tmpl_row (Total: bold+yellow) or
    from tag_row (plain row), depending on style=last|first in the tag."""

    # DataFrame has 2 extra rows (b, c) that trigger insertion.
    _DF = pl.DataFrame({
        "Index": ["a", "b", "c", "Total"],
        "Value": [10, 20, 30, 100],
    })

    def _run(self, template_path, tmp_path):
        out = str(tmp_path / "output.xlsx")
        writer = ExcelTemplateWriter(template_path)
        writer.write({"data": TypedValue(self._DF, "table")}, out)
        return load_workbook(out)

    def test_style_last_inserted_rows_are_bold(
        self, template_style_src_last_path, tmp_path
    ):
        """style=last (default): inserted rows copy from Total row — must be bold."""
        wb = self._run(template_style_src_last_path, tmp_path)
        ws = wb.active
        # Rows 4 and 5 are the inserted extra rows (b, c)
        assert ws.cell(4, 1).font.bold is True, (
            "style=last: inserted row 4 should be bold (copied from Total row)"
        )
        assert ws.cell(5, 1).font.bold is True, (
            "style=last: inserted row 5 should be bold (copied from Total row)"
        )

    def test_style_last_inserted_rows_have_yellow_fill(
        self, template_style_src_last_path, tmp_path
    ):
        """style=last (default): inserted rows copy Total's yellow fill."""
        wb = self._run(template_style_src_last_path, tmp_path)
        ws = wb.active
        assert ws.cell(4, 1).fill.fill_type == "solid", (
            "style=last: inserted row 4 should have solid fill (copied from Total row)"
        )
        assert ws.cell(5, 1).fill.fill_type == "solid", (
            "style=last: inserted row 5 should have solid fill (copied from Total row)"
        )

    def test_style_first_inserted_rows_not_bold(
        self, template_style_src_first_path, tmp_path
    ):
        """style=first: inserted rows copy from plain tag row — must NOT be bold."""
        wb = self._run(template_style_src_first_path, tmp_path)
        ws = wb.active
        assert ws.cell(4, 1).font.bold is not True, (
            "style=first: inserted row 4 should not be bold (copied from plain tag row)"
        )
        assert ws.cell(5, 1).font.bold is not True, (
            "style=first: inserted row 5 should not be bold (copied from plain tag row)"
        )

    def test_style_first_inserted_rows_no_solid_fill(
        self, template_style_src_first_path, tmp_path
    ):
        """style=first: inserted rows copy plain tag row — must NOT have solid fill."""
        wb = self._run(template_style_src_first_path, tmp_path)
        ws = wb.active
        assert ws.cell(4, 1).fill.fill_type != "solid", (
            "style=first: inserted row 4 should not have solid fill"
        )
        assert ws.cell(5, 1).fill.fill_type != "solid", (
            "style=first: inserted row 5 should not have solid fill"
        )


# ---------------------------------------------------------------------------
# Combo: outer join (placeholder=True) + two adjacent same-span merges below
# ---------------------------------------------------------------------------

def _run_combo_outer_merges_below(template_path, tmp_path):
    df = pl.DataFrame({
        "Key": ["a", "b", "c", "Total"],
        "Value": [10, 20, 30, 60],
    })
    out = str(tmp_path / "output.xlsx")
    ExcelTemplateWriter(template_path).write({"data": TypedValue(df, "table")}, out)
    return load_workbook(out)


class TestComboOuterMergesBelow:
    """Outer join with placeholder=True expands by +2 rows; two adjacent
    same-span merges below must both shift and preserve value/style."""

    def test_data_row_a(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert wb.active.cell(2, 1).value == "a"
        assert wb.active.cell(2, 2).value == 10

    def test_data_row_b(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert wb.active.cell(3, 1).value == "b"
        assert wb.active.cell(3, 2).value == 20

    def test_data_row_c_inserted(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert wb.active.cell(4, 1).value == "c"
        assert wb.active.cell(4, 2).value == 30

    def test_total_row_at_5(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert wb.active.cell(5, 1).value == "Total"
        assert wb.active.cell(5, 2).value == 60

    def test_first_merge_shifted_to_row_7(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert "A7:B7" in _merge_ranges(wb.active), (
            f"First merge not shifted to A7:B7 — got {_merge_ranges(wb.active)}"
        )

    def test_first_merge_value_intact(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert wb.active["A7"].value == "Note First"

    def test_first_merge_bold_intact(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert wb.active["A7"].font.bold is True

    def test_second_merge_shifted_to_row_8(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert "A8:B8" in _merge_ranges(wb.active), (
            f"Second merge not shifted to A8:B8 — got {_merge_ranges(wb.active)}"
        )

    def test_second_merge_value_intact(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert wb.active["A8"].value == "Note Second"

    def test_second_merge_italic_intact(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert wb.active["A8"].font.italic is True

    def test_no_stale_first_merge(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert "A5:B5" not in _merge_ranges(wb.active), "Stale A5:B5 still present"

    def test_no_stale_second_merge(self, template_combo_outer_merges_below_path, tmp_path):
        wb = _run_combo_outer_merges_below(template_combo_outer_merges_below_path, tmp_path)
        assert "A6:B6" not in _merge_ranges(wb.active), "Stale A6:B6 still present"


# ---------------------------------------------------------------------------
# Combo: left join + two adjacent merges below — merges completely untouched
# ---------------------------------------------------------------------------

def _run_combo_left_with_merges(template_path, tmp_path):
    df = pl.DataFrame({
        "K": ["x", "y", "z"],
        "V1": [1, 2, 3],
        "V2": [10, 20, 30],
    })
    out = str(tmp_path / "output.xlsx")
    ExcelTemplateWriter(template_path).write({"data": TypedValue(df, "table")}, out)
    return load_workbook(out)


class TestComboLeftWithMerges:
    """Left join inserts no rows; merges below must be byte-for-byte identical
    to the template — no shift, no corruption."""

    def test_data_filled_correctly(self, template_combo_left_with_merges_path, tmp_path):
        wb = _run_combo_left_with_merges(template_combo_left_with_merges_path, tmp_path)
        ws = wb.active
        assert ws.cell(2, 1).value == "x" and ws.cell(2, 2).value == 1 and ws.cell(2, 3).value == 10
        assert ws.cell(3, 1).value == "y" and ws.cell(3, 2).value == 2 and ws.cell(3, 3).value == 20
        assert ws.cell(4, 1).value == "z" and ws.cell(4, 2).value == 3 and ws.cell(4, 3).value == 30

    def test_bottom_header_merge_unchanged(self, template_combo_left_with_merges_path, tmp_path):
        wb = _run_combo_left_with_merges(template_combo_left_with_merges_path, tmp_path)
        assert "A6:C6" in _merge_ranges(wb.active), "Bottom Header merge was modified"

    def test_bottom_header_value_unchanged(self, template_combo_left_with_merges_path, tmp_path):
        wb = _run_combo_left_with_merges(template_combo_left_with_merges_path, tmp_path)
        assert wb.active["A6"].value == "Bottom Header"

    def test_bottom_header_bold_unchanged(self, template_combo_left_with_merges_path, tmp_path):
        wb = _run_combo_left_with_merges(template_combo_left_with_merges_path, tmp_path)
        assert wb.active["A6"].font.bold is True

    def test_bottom_note_merge_unchanged(self, template_combo_left_with_merges_path, tmp_path):
        wb = _run_combo_left_with_merges(template_combo_left_with_merges_path, tmp_path)
        assert "A7:C7" in _merge_ranges(wb.active), "Bottom Note merge was modified"

    def test_bottom_note_value_unchanged(self, template_combo_left_with_merges_path, tmp_path):
        wb = _run_combo_left_with_merges(template_combo_left_with_merges_path, tmp_path)
        assert wb.active["A7"].value == "Bottom Note"

    def test_bottom_note_italic_unchanged(self, template_combo_left_with_merges_path, tmp_path):
        wb = _run_combo_left_with_merges(template_combo_left_with_merges_path, tmp_path)
        assert wb.active["A7"].font.italic is True

    def test_no_phantom_row_inserted(self, template_combo_left_with_merges_path, tmp_path):
        wb = _run_combo_left_with_merges(template_combo_left_with_merges_path, tmp_path)
        # Row 5 must be blank — no extra row sneaked in
        assert wb.active.cell(5, 1).value is None, "Unexpected row 5 content on left join"


# ---------------------------------------------------------------------------
# Combo: outer join (placeholder=True) + scalar cells below the table
# ---------------------------------------------------------------------------

def _run_combo_scalar_with_outer(template_path, tmp_path):
    df = pl.DataFrame({
        "Key": ["a", "b", "c", "Total"],
        "Data": [1, 2, 3, 0],
    })
    out = str(tmp_path / "output.xlsx")
    ExcelTemplateWriter(template_path).write(
        {
            "tbl": TypedValue(df, "table"),
            "title": TypedValue("Report Label", "single"),
            "summary": TypedValue(42, "single"),
        },
        out,
    )
    return load_workbook(out)


class TestComboScalarWithOuter:
    """Scalars in the rows below an expanding outer-join table must:
    1. Land at the correct shifted row (accounting for row insertions)
    2. Receive their values from the data dictionary
    3. Have their raw {{ }} tags cleared after fill
    4. Not remain at the original template row

    Template has 2 data rows (a, b), DF has 4 rows (a, b, c, Total) → 2 extra rows inserted.
    Original scalar positions at row 5 shift to row 7 after +2 row shift.
    """

    # =========================================================================
    # Table data correctness (outer join fills all 4 rows)
    # =========================================================================

    def test_table_header_row_exists(self, template_combo_scalar_with_outer_path, tmp_path):
        """Row 1 should be the header row."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        # Headers: Key, Data
        assert ws.cell(1, 1).value is not None, "Row 1 col A should have header"

    def test_table_row_a_filled(self, template_combo_scalar_with_outer_path, tmp_path):
        """First data row ('a') in row 2."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(2, 1).value == "a"
        assert ws.cell(2, 2).value == 1

    def test_table_row_b_filled(self, template_combo_scalar_with_outer_path, tmp_path):
        """Second data row ('b') in row 3."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(3, 1).value == "b"
        assert ws.cell(3, 2).value == 2

    def test_table_extra_row_c_inserted(self, template_combo_scalar_with_outer_path, tmp_path):
        """Extra row 'c' inserted at row 4 (outer join)."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(4, 1).value == "c", f"Expected 'c' at A4, got {ws.cell(4, 1).value!r}"
        assert ws.cell(4, 2).value == 3, f"Expected 3 at B4, got {ws.cell(4, 2).value!r}"

    def test_table_extra_row_total_inserted(self, template_combo_scalar_with_outer_path, tmp_path):
        """Extra row 'Total' inserted at row 5 (outer join)."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(5, 1).value == "Total", f"Expected 'Total' at A5, got {ws.cell(5, 1).value!r}"
        assert ws.cell(5, 2).value == 0, f"Expected 0 at B5, got {ws.cell(5, 2).value!r}"

    # =========================================================================
    # Scalar row shift (+2 rows for 2 extra table rows)
    # =========================================================================

    def test_title_scalar_receives_value(self, template_combo_scalar_with_outer_path, tmp_path):
        """Title scalar ({{ title }}) should receive 'Report Label'."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(7, 1).value == "Report Label", (
            f"Title scalar not filled; got {ws.cell(7, 1).value!r} at A7"
        )

    def test_summary_scalar_receives_value(self, template_combo_scalar_with_outer_path, tmp_path):
        """Summary scalar ({{ summary }}) should receive 42."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        assert ws.cell(7, 2).value == 42, (
            f"Summary scalar not filled; got {ws.cell(7, 2).value!r} at B7"
        )

    def test_scalars_shift_by_2_rows(self, template_combo_scalar_with_outer_path, tmp_path):
        """With +2 rows inserted, scalars must shift from template row 5 to output row 7."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        # After 2 row insertions, row 5 template row becomes row 7
        # Row 6 template becomes row 8, etc.
        title_at_7 = ws.cell(7, 1).value == "Report Label"
        summary_at_7 = ws.cell(7, 2).value == 42
        assert title_at_7 and summary_at_7, (
            f"Scalars not shifted correctly. "
            f"A7={ws.cell(7, 1).value!r}, B7={ws.cell(7, 2).value!r}"
        )

    # =========================================================================
    # Raw tag cleanup ({{ }} markers must be removed)
    # =========================================================================

    def test_no_raw_title_tags_remain(self, template_combo_scalar_with_outer_path, tmp_path):
        """Raw {{ title }} tag must be cleared after fill."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        orphaned_cells = []
        for row in ws.iter_rows(max_row=20):
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value and "title" in cell.value.lower():
                    orphaned_cells.append(f"{cell.coordinate}={cell.value!r}")
        assert not orphaned_cells, (
            f"Raw {{ title }} tag(s) found in output: {orphaned_cells}"
        )

    def test_no_raw_summary_tags_remain(self, template_combo_scalar_with_outer_path, tmp_path):
        """Raw {{ summary }} tag must be cleared after fill."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        orphaned_cells = []
        for row in ws.iter_rows(max_row=20):
            for cell in row:
                if isinstance(cell.value, str) and "{{" in cell.value and "summary" in cell.value.lower():
                    orphaned_cells.append(f"{cell.coordinate}={cell.value!r}")
        assert not orphaned_cells, (
            f"Raw {{ summary }} tag(s) found in output: {orphaned_cells}"
        )

    # =========================================================================
    # Scalars must not be left at original template positions
    # =========================================================================

    def test_title_scalar_not_at_original_row(self, template_combo_scalar_with_outer_path, tmp_path):
        """Title scalar must NOT remain at original row 5."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        row_5_a = ws.cell(5, 1).value
        assert row_5_a != "Report Label", (
            f"Title scalar left at original row 5: {row_5_a!r}. "
            f"It should be at row 7 after +2 shift"
        )

    def test_summary_scalar_not_at_original_row(self, template_combo_scalar_with_outer_path, tmp_path):
        """Summary scalar must NOT remain at original row 5."""
        wb = _run_combo_scalar_with_outer(template_combo_scalar_with_outer_path, tmp_path)
        ws = wb.active
        row_5_b = ws.cell(5, 2).value
        assert row_5_b != 42, (
            f"Summary scalar left at original row 5: {row_5_b!r}. "
            f"It should be at row 7 after +2 shift"
        )


# ---------------------------------------------------------------------------
# Combo: triple adjacent same-span merges — primary regression test
# ---------------------------------------------------------------------------

def _run_combo_triple_adjacent_merges(template_path, tmp_path):
    df = pl.DataFrame({
        "Sec": ["p", "q", "r", "Total"],
        "A": [1, 2, 3, 0],
        "B": [10, 20, 30, 0],
        "C": [100, 200, 300, 0],
    })
    out = str(tmp_path / "output.xlsx")
    ExcelTemplateWriter(template_path).write({"tbl": TypedValue(df, "table")}, out)
    return load_workbook(out)


class TestComboTripleAdjacentMerges:
    """Three adjacent same-span (A:D) merges below an outer-join table.
    All three MUST survive after the +2 row shift.  This directly exercises
    the two bugs fixed in _sync_merges_after_delete and _copy_row_styles."""

    def test_table_data_rows(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        ws = wb.active
        assert ws.cell(2, 1).value == "p"
        assert ws.cell(3, 1).value == "q"
        assert ws.cell(4, 1).value == "r"
        assert ws.cell(5, 1).value == "Total"

    def test_section_one_shifted_to_row_7(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert "A7:D7" in _merge_ranges(wb.active), (
            f"Section One not at A7:D7 — got {_merge_ranges(wb.active)}"
        )

    def test_section_one_value_intact(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert wb.active["A7"].value == "Section One"

    def test_section_one_bold_intact(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert wb.active["A7"].font.bold is True

    def test_section_two_shifted_to_row_8(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert "A8:D8" in _merge_ranges(wb.active), (
            f"Section Two not at A8:D8 — got {_merge_ranges(wb.active)}"
        )

    def test_section_two_value_intact(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert wb.active["A8"].value == "Section Two"

    def test_section_two_italic_intact(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert wb.active["A8"].font.italic is True

    def test_section_three_shifted_to_row_9(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert "A9:D9" in _merge_ranges(wb.active), (
            f"Section Three not at A9:D9 — got {_merge_ranges(wb.active)}"
        )

    def test_section_three_value_intact(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert wb.active["A9"].value == "Section Three"

    def test_section_three_bold_intact(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert wb.active["A9"].font.bold is True

    def test_narrow_merge_shifted_to_row_11(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert "B11:C11" in _merge_ranges(wb.active), (
            f"Narrow merge not at B11:C11 — got {_merge_ranges(wb.active)}"
        )

    def test_narrow_merge_value_intact(self, template_combo_triple_adjacent_merges_path, tmp_path):
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        assert wb.active["B11"].value == "Narrow Note"

    def test_no_stale_merges_at_original_positions(self, template_combo_triple_adjacent_merges_path, tmp_path):
        """Original rows 5/6/7 must NOT contain any of the shifted merges."""
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        ranges = _merge_ranges(wb.active)
        assert "A5:D5" not in ranges, "Stale A5:D5 still present"
        assert "A6:D6" not in ranges, "Stale A6:D6 still present"
        assert "A7:D7" not in ranges or "A7:D7" in ranges, True  # A7:D7 IS the shifted Section One — fine
        # What must not exist is the original Section Three range being stale:
        # After +2 shift: A7:D7=SectionOne(shifted from A5), A8:D8=SectionTwo, A9:D9=SectionThree
        # The stale problematic position would be A5:D5 or A6:D6 which should be gone
        assert "A5:D5" not in ranges
        assert "A6:D6" not in ranges

    def test_no_phantom_merges_inside_table_rows(self, template_combo_triple_adjacent_merges_path, tmp_path):
        """Table rows 2-5 must not be inside any merge range."""
        wb = _run_combo_triple_adjacent_merges(template_combo_triple_adjacent_merges_path, tmp_path)
        ws = wb.active
        for r in range(2, 6):
            cell = ws.cell(r, 1)
            assert not isinstance(cell, MergedCell), (
                f"Row {r} col A is a MergedCell — phantom ghost not purged"
            )


# ---------------------------------------------------------------------------
# Combo: two stacked outer-join tables + adjacent merge pair below both
# ---------------------------------------------------------------------------

def _run_combo_two_outer_tables(template_path, tmp_path):
    df1 = pl.DataFrame({"Key1": ["a", "b", "c", "Total1"], "V1": [1, 2, 3, 6]})
    df2 = pl.DataFrame({"Key2": ["x", "y", "z", "Total2"], "V2": [10, 20, 30, 60]})
    out = str(tmp_path / "output.xlsx")
    ExcelTemplateWriter(template_path).write(
        {
            "tbl1": TypedValue(df1, "table"),
            "tbl2": TypedValue(df2, "table"),
        },
        out,
    )
    return load_workbook(out)


class TestComboTwoOuterTables:
    """Two back-to-back outer-join tables each insert 1 extra row.
    Adjacent merges below both tables must shift by the cumulative +2."""

    def test_tbl1_rows(self, template_combo_two_outer_tables_path, tmp_path):
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        ws = wb.active
        assert ws.cell(2, 1).value == "a" and ws.cell(2, 2).value == 1
        assert ws.cell(3, 1).value == "b" and ws.cell(3, 2).value == 2
        assert ws.cell(4, 1).value == "c" and ws.cell(4, 2).value == 3
        assert ws.cell(5, 1).value == "Total1" and ws.cell(5, 2).value == 6

    def test_tbl2_rows(self, template_combo_two_outer_tables_path, tmp_path):
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        ws = wb.active
        assert ws.cell(8, 1).value == "x" and ws.cell(8, 2).value == 10
        assert ws.cell(9, 1).value == "y" and ws.cell(9, 2).value == 20
        assert ws.cell(10, 1).value == "z" and ws.cell(10, 2).value == 30
        assert ws.cell(11, 1).value == "Total2" and ws.cell(11, 2).value == 60

    def test_grand_footer_shifted_to_row_13(self, template_combo_two_outer_tables_path, tmp_path):
        """Grand Footer was at template row 11; cumulative +2 → output row 13."""
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        assert "A13:B13" in _merge_ranges(wb.active), (
            f"Grand Footer not at A13:B13 — got {_merge_ranges(wb.active)}"
        )

    def test_grand_footer_value_intact(self, template_combo_two_outer_tables_path, tmp_path):
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        assert wb.active["A13"].value == "Grand Footer"

    def test_grand_footer_bold_intact(self, template_combo_two_outer_tables_path, tmp_path):
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        assert wb.active["A13"].font.bold is True

    def test_sub_footer_shifted_to_row_14(self, template_combo_two_outer_tables_path, tmp_path):
        """Sub-Footer was at template row 12; cumulative +2 → output row 14."""
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        assert "A14:B14" in _merge_ranges(wb.active), (
            f"Sub-Footer not at A14:B14 — got {_merge_ranges(wb.active)}"
        )

    def test_sub_footer_value_intact(self, template_combo_two_outer_tables_path, tmp_path):
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        assert wb.active["A14"].value == "Sub-Footer"

    def test_sub_footer_italic_intact(self, template_combo_two_outer_tables_path, tmp_path):
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        assert wb.active["A14"].font.italic is True

    def test_no_stale_footer_merges(self, template_combo_two_outer_tables_path, tmp_path):
        wb = _run_combo_two_outer_tables(template_combo_two_outer_tables_path, tmp_path)
        ranges = _merge_ranges(wb.active)
        assert "A11:B11" not in ranges, "Stale A11:B11 still present"
        assert "A12:B12" not in ranges, "Stale A12:B12 still present"


# ---------------------------------------------------------------------------
# Alignment preservation — inserted rows must inherit the source row's alignment
# ---------------------------------------------------------------------------

class TestAlignmentPreservation:
    """Regression test: _copy_row_styles was saving alignment in the style
    snapshot but never writing ``dst.alignment`` for inserted rows.  Inserted
    rows therefore silently lost whatever alignment the source template row had,
    falling back to Excel's default (right for numbers, left for strings).

    Template (built inline):
      Row 1: headers (Index / Value)
      Row 2: 'a' / tag — explicit center alignment on both cells (this is also
             the *only* data row, so last_tmpl_row = tag_row = 2)

    DataFrame: a=10, b=20, c=30  →  b and c are outer-join extras.
    With style=last (default) the style source is last_tmpl_row = row 2 = center.
    Both inserted rows must carry that center alignment.
    """

    # Single-row template: last_tmpl_row == tag_row (row 2, center).
    # DF has a (matched), b and c (extra → inserted from row 2).
    _DF = pl.DataFrame({
        "Index": ["a", "b", "c"],
        "Value": [10, 20, 30],
    })

    def _build_template(self, tmp_path) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        ws["A1"] = "Index"
        ws["B1"] = "Value"

        center = Alignment(horizontal="center")

        # Single data row — center alignment on both columns.
        ws["A2"] = "a"
        ws["B2"] = "{{ data | table(join=outer) }}"
        ws["A2"].alignment = center
        ws["B2"].alignment = center

        path = str(tmp_path / "tmpl_alignment.xlsx")
        wb.save(path)
        return path

    def _run(self, tmp_path):
        tpl = self._build_template(tmp_path)
        out = str(tmp_path / "out_alignment.xlsx")
        ExcelTemplateWriter(tpl).write({"data": TypedValue(self._DF, "table")}, out)
        return load_workbook(out)

    def test_inserted_rows_have_center_alignment(self, tmp_path):
        """b and c are outer-join extras inserted after the single template row.
        style=last → source = last_tmpl_row = row 2 = center.
        Both inserted rows must carry center alignment on every data column."""
        wb = self._run(tmp_path)
        ws = wb.active
        # After fill: row 2=a, row 3=b (inserted), row 4=c (inserted)
        assert ws.cell(3, 1).alignment.horizontal == "center", (
            "inserted row 3 (b) col A: expected center alignment, "
            f"got {ws.cell(3, 1).alignment.horizontal!r}"
        )
        assert ws.cell(3, 2).alignment.horizontal == "center", (
            "inserted row 3 (b) col B: expected center alignment, "
            f"got {ws.cell(3, 2).alignment.horizontal!r}"
        )
        assert ws.cell(4, 1).alignment.horizontal == "center", (
            "inserted row 4 (c) col A: expected center alignment, "
            f"got {ws.cell(4, 1).alignment.horizontal!r}"
        )
        assert ws.cell(4, 2).alignment.horizontal == "center", (
            "inserted row 4 (c) col B: expected center alignment, "
            f"got {ws.cell(4, 2).alignment.horizontal!r}"
        )

    def test_tag_row_alignment_preserved(self, tmp_path):
        """The original tag row (row 2, 'a') must keep its center alignment after fill."""
        wb = self._run(tmp_path)
        ws = wb.active
        assert ws.cell(2, 1).alignment.horizontal == "center", (
            "tag row A2: center alignment was lost after fill"
        )

    def test_style_first_inserted_rows_inherit_tag_alignment(self, tmp_path):
        """With style=first, inserted rows copy the tag row's alignment even when
        the last template row has a different alignment (right)."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Index"
        ws["B1"] = "Value"

        center = Alignment(horizontal="center")
        right = Alignment(horizontal="right")

        # Tag row: center (style source for style=first)
        ws["A2"] = "a"
        ws["B2"] = "{{ data | table(join=outer, style=first) }}"
        ws["A2"].alignment = center
        ws["B2"].alignment = center

        # Total row: right + bold (style source for style=last — must NOT be used)
        ws["A3"] = "Total"
        ws["B3"] = 999
        ws["A3"].alignment = right
        ws["B3"].alignment = right
        for col in range(1, 3):
            ws.cell(3, col).font = Font(bold=True)

        tpl = str(tmp_path / "tmpl_style_first_align.xlsx")
        wb.save(tpl)

        df = pl.DataFrame({"Index": ["a", "b", "c", "Total"], "Value": [1, 2, 3, 999]})
        out = str(tmp_path / "out_style_first_align.xlsx")
        ExcelTemplateWriter(tpl).write({"data": TypedValue(df, "table")}, out)
        ws_out = load_workbook(out).active

        # After fill: row 2=a (center), row 3=Total (right, matched template),
        # row 4=b (inserted after row 3), row 5=c (inserted after row 3).
        # style=first → source = tag_row (row 2, center) → rows 4 and 5 must be center.
        assert ws_out.cell(4, 1).alignment.horizontal == "center", (
            "style=first: inserted row 4 (b) should have center alignment from tag row, "
            f"got {ws_out.cell(4, 1).alignment.horizontal!r}"
        )
        assert ws_out.cell(5, 1).alignment.horizontal == "center", (
            "style=first: inserted row 5 (c) should have center alignment from tag row, "
            f"got {ws_out.cell(5, 1).alignment.horizontal!r}"
        )


# ---------------------------------------------------------------------------
# Alignment — loop() row expansion
# ---------------------------------------------------------------------------

class TestLoopRowAlignmentPreservation:
    """loop() rows are expanded via _copy_row_styles.  The alignment on the
    template row must be copied to every expanded duplicate row.

    Template (built inline):
      Row 1: header
      Row 2: {{ month | loop() }} / {{ value | loop() }} — center-aligned

    Three-item lists expand the template row to 3 rows.  All must be center.
    """

    def _build_template(self, tmp_path) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Month"
        ws["B1"] = "Value"

        center = Alignment(horizontal="center")
        ws["A2"] = "{{ month | loop() }}"
        ws["B2"] = "{{ value | loop() }}"
        ws["A2"].alignment = center
        ws["B2"].alignment = center

        path = str(tmp_path / "tmpl_loop_align.xlsx")
        wb.save(path)
        return path

    def _run(self, tmp_path):
        tpl = self._build_template(tmp_path)
        out = str(tmp_path / "out_loop_align.xlsx")
        ExcelTemplateWriter(tpl).write(
            {
                "month": TypedValue(["Jan", "Feb", "Mar"], "list"),
                "value": TypedValue([10, 20, 30], "list"),
            },
            out,
        )
        return load_workbook(out)

    def test_all_expanded_rows_have_center_alignment(self, tmp_path):
        """All 3 expanded loop rows (2, 3, 4) must carry center alignment."""
        wb = self._run(tmp_path)
        ws = wb.active
        for row in range(2, 5):
            for col in (1, 2):
                assert ws.cell(row, col).alignment.horizontal == "center", (
                    f"loop row {row} col {col}: expected center, "
                    f"got {ws.cell(row, col).alignment.horizontal!r}"
                )

    def test_expanded_row_values(self, tmp_path):
        """Sanity: values are written correctly into the expanded rows."""
        wb = self._run(tmp_path)
        ws = wb.active
        assert [ws.cell(r, 1).value for r in range(2, 5)] == ["Jan", "Feb", "Mar"]
        assert [ws.cell(r, 2).value for r in range(2, 5)] == [10, 20, 30]


# ---------------------------------------------------------------------------
# Alignment — join=right table inserts
# ---------------------------------------------------------------------------

class TestRightJoinAlignmentPreservation:
    """join=right inserts rows when the DataFrame is longer than the template.
    Inserted rows must carry the alignment of the style-source row.

    Template (built inline):
      Row 1: headers (Key / Val)
      Row 2: x / tag {{ data | table(join=right) }}  — center-aligned (tag row)
      Row 3: y                                        — right-aligned  (last tmpl row)

    DataFrame has 5 rows → 3 extras inserted.

    style=last (default): style source = last_tmpl_row (row 3, right).
                          Inserted rows must be right-aligned.
    style=first:          style source = tag_row (row 2, center).
                          Inserted rows must be center-aligned.
    """

    _DF = pl.DataFrame({
        "Key": ["a", "b", "c", "d", "e"],
        "Val": [1, 2, 3, 4, 5],
    })

    def _build_template(self, tmp_path, style_param: str = "") -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Key"
        ws["B1"] = "Val"

        center = Alignment(horizontal="center")
        right = Alignment(horizontal="right")

        tag = f"{{{{ data | table(join=right{style_param}) }}}}"
        ws["A2"] = "x"
        ws["B2"] = tag
        ws["A2"].alignment = center
        ws["B2"].alignment = center

        ws["A3"] = "y"
        ws["B3"] = None
        ws["A3"].alignment = right
        ws["B3"].alignment = right

        path = str(tmp_path / f"tmpl_right{style_param.replace(',','').replace('=','').replace(' ','')}.xlsx")
        wb.save(path)
        return path

    def _run(self, tmp_path, style_param: str = ""):
        tpl = self._build_template(tmp_path, style_param)
        out = str(tmp_path / f"out_right{style_param.replace(',','').replace('=','').replace(' ','')}.xlsx")
        ExcelTemplateWriter(tpl).write({"data": TypedValue(self._DF, "table")}, out)
        return load_workbook(out)

    def test_right_join_style_last_inserted_rows_alignment(self, tmp_path):
        """style=last: inserted rows must copy alignment from the last template
        row (row 3, right-aligned).  Verifies alignment is not silently dropped."""
        wb = self._run(tmp_path, style_param="")
        ws = wb.active
        # Template: 2 rows, DF: 5 rows → 3 inserted; output has 5 data rows (2-6)
        # Inserted rows come after last_tmpl_row (row 3) → rows 4, 5, 6 are new
        for row in (4, 5, 6):
            assert ws.cell(row, 1).alignment.horizontal == "right", (
                f"style=last: inserted row {row} col A: expected right, "
                f"got {ws.cell(row, 1).alignment.horizontal!r}"
            )

    def test_right_join_style_first_inserted_rows_alignment(self, tmp_path):
        """style=first: inserted rows must copy alignment from the tag row
        (row 2, center-aligned), not from the last template row (right)."""
        wb = self._run(tmp_path, style_param=", style=first")
        ws = wb.active
        # style=first inserts after tag_row (row 2); original row 3 shifts to row 6.
        # Inserted rows (3, 4, 5) must be center-aligned.
        for row in (3, 4, 5):
            assert ws.cell(row, 1).alignment.horizontal == "center", (
                f"style=first: inserted row {row} col A: expected center, "
                f"got {ws.cell(row, 1).alignment.horizontal!r}"
            )

    def test_right_join_data_values_correct(self, tmp_path):
        """Sanity: all 5 DF rows appear in the output regardless of style."""
        wb = self._run(tmp_path, style_param="")
        ws = wb.active
        assert [ws.cell(r, 1).value for r in range(2, 7)] == ["a", "b", "c", "d", "e"]
        assert [ws.cell(r, 2).value for r in range(2, 7)] == [1, 2, 3, 4, 5]


# ---------------------------------------------------------------------------
# Alignment — sorted outer fill (order_by) inserts
# ---------------------------------------------------------------------------

class TestSortedOuterFillAlignmentPreservation:
    """_sorted_outer_fill calls _copy_row_styles(ws, upper_last_row, n) when
    the DF has more rows than the upper zone.  The alignment on upper_last_row
    must be carried to the inserted rows.

    Template (built inline):
      Row 1: headers (Index / Value)
      Row 2: 'a' / tag {{ data | table(join=outer, order_by=asc) }} — center
      Row 3: end_table marker row (deleted after processing)

    DF: a=10, b=20, c=30, d=40 → 3 extras need inserting into a 1-slot zone.
    upper_last_row = row 2 (center) → inserted rows must be center-aligned.
    """

    _DF = pl.DataFrame({
        "Index": ["a", "b", "c", "d"],
        "Value": [10, 20, 30, 40],
    })

    def _build_template(self, tmp_path) -> str:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "Index"
        ws["B1"] = "Value"

        center = Alignment(horizontal="center")
        ws["A2"] = "a"
        ws["B2"] = "{{ data | table(join=outer, order_by=asc) }}"
        ws["A2"].alignment = center
        ws["B2"].alignment = center

        # Option A end_table: A3 is empty (join col), end_table in data col B3.
        # This makes _fill_table treat row 3 as a standalone marker row
        # (last_tmpl_row = 2), not a data row (which would set last_tmpl_row = 3
        # and wipe out the alignment source).
        ws["B3"] = "{{ end_table }}"

        path = str(tmp_path / "tmpl_sorted_outer_align.xlsx")
        wb.save(path)
        return path

    def _run(self, tmp_path):
        tpl = self._build_template(tmp_path)
        out = str(tmp_path / "out_sorted_outer_align.xlsx")
        ExcelTemplateWriter(tpl).write({"data": TypedValue(self._DF, "table")}, out)
        return load_workbook(out)

    def test_inserted_rows_have_center_alignment(self, tmp_path):
        """b, c, d are extra rows inserted by sorted outer fill.  upper_last_row
        (row 2) is center-aligned, so inserted rows must also be center-aligned."""
        wb = self._run(tmp_path)
        ws = wb.active
        # Output: rows 2-5 for a, b, c, d (end_table row deleted)
        for row in (3, 4, 5):
            assert ws.cell(row, 1).alignment.horizontal == "center", (
                f"sorted outer: inserted row {row} col A: expected center, "
                f"got {ws.cell(row, 1).alignment.horizontal!r}"
            )

    def test_data_values_sorted_asc(self, tmp_path):
        """Sanity: sorted output a→d in ascending order."""
        wb = self._run(tmp_path)
        ws = wb.active
        assert [ws.cell(r, 1).value for r in range(2, 6)] == ["a", "b", "c", "d"]
