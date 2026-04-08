import os
from typing import Any
from types import TracebackType
from openpyxl.utils import coordinate_to_tuple
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel.exceptions import ExcelError, ExcelSheetNotFoundError, KeywordNotFoundError, MultipleKeywordsFoundError
from excel._utils import load_excel_workbook


class ExcelCellReader:
    """Read individual cell values from an Excel file."""

    def __init__(self, filepath: str | os.PathLike[str] | bytes):
        self.filepath = filepath
        self._wb: Workbook | None = None

    @property
    def wb(self) -> Workbook:
        if self._wb is None:
            raise ExcelError(
                "Workbook not loaded. Use ExcelCellReader as context manager:\n"
                "    with ExcelCellReader('file.xlsx') as reader:\n"
                "        value = reader.get('Sheet1!B5')"
            )
        return self._wb

    def __enter__(self) -> "ExcelCellReader":
        self._wb = load_excel_workbook(self.filepath, data_only=True)
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: TracebackType | None,
    ) -> None:
        if self._wb:
            self._wb.close()

    def get(self, cell_ref: str) -> Any:
        """
        Get the value of a single cell.

        Args:
            cell_ref: Cell reference, either "Sheet1!B5" or "B5" (uses active sheet)

        Returns:
            Cell value (str, int, float, datetime, or None)

        Example:
            >>> with ExcelCellReader('config.xlsx') as reader:
            ...     rate = reader.get('Config!B5')
            ...     name = reader.get('A1')  # active sheet
        """
        sheet, cell_addr = self._parse_ref(cell_ref)
        row, col = coordinate_to_tuple(cell_addr)
        return sheet.cell(row, col).value

    def get_many(self, cell_refs: list[str]) -> dict[str, Any]:
        """
        Get values for multiple cell references.

        Args:
            cell_refs: List of cell references (e.g., ["Sheet1!B5", "Sheet2!C10"])

        Returns:
            Dict mapping each reference to its value

        Example:
            >>> with ExcelCellReader('config.xlsx') as reader:
            ...     values = reader.get_many(['Config!B5', 'Config!B6'])
            ...     # {'Config!B5': 0.15, 'Config!B6': 'USD'}
        """
        return {ref: self.get(ref) for ref in cell_refs}

    def get_relative(
        self,
        cell_ref: str | None = None,
        keyword: str | None = None,
        *,
        sheet: str | None = None,
        right: int = 0,
        left: int = 0,
        down: int = 0,
        up: int = 0,
    ) -> Any:
        """
        Get a cell value anchored by a cell address or a keyword label, with
        directional offset applied.

        Exactly one of ``cell_ref`` or ``keyword`` must be provided.

        Args:
            cell_ref: Anchor cell address, e.g. ``"Sheet1!B5"`` or ``"B5"``.
            keyword: Text to search for in the workbook. Case-insensitive exact
                match (after stripping whitespace).
            sheet: Restrict keyword search to this sheet. Ignored when using
                ``cell_ref``.
            right: Steps to move right from the anchor.
            left: Steps to move left from the anchor.
            down: Steps to move down from the anchor.
            up: Steps to move up from the anchor.

        Returns:
            Cell value at the computed offset position.

        Raises:
            ValueError: If both or neither of ``cell_ref``/``keyword`` are given.
            KeywordNotFoundError: Keyword not found in any searched sheet.
            MultipleKeywordsFoundError: Keyword found in more than one cell.

        Example:
            >>> with ExcelCellReader('report.xlsx') as reader:
            ...     value = reader.get_relative(keyword='Revenue', right=1)
            ...     value = reader.get_relative(cell_ref='Sheet1!A2', down=1)
        """
        if cell_ref is None and keyword is None:
            raise ValueError("Specify either cell_ref= or keyword=")
        if cell_ref is not None and keyword is not None:
            raise ValueError("Cannot specify both cell_ref= and keyword=")

        if keyword is not None:
            anchor_sheet, row, col = self._find_keyword(keyword, sheet)
        else:
            anchor_sheet, cell_addr = self._parse_ref(cell_ref)  # type: ignore[arg-type]
            row, col = coordinate_to_tuple(cell_addr)

        return self._value_at_offset(anchor_sheet, row, col, right=right, left=left, down=down, up=up)

    def get_many_relative(
        self,
        cell_ref: str | None = None,
        keyword: str | None = None,
        offsets: dict[str, dict[str, int]] | None = None,
        *,
        sheet: str | None = None,
    ) -> dict[str, Any]:
        """
        Get multiple values at different offsets from a single anchor.

        Exactly one of ``cell_ref`` or ``keyword`` must be provided.

        Args:
            cell_ref: Anchor cell address, e.g. ``"Sheet1!B5"`` or ``"B5"``.
            keyword: Text to search for in the workbook. Case-insensitive exact
                match (after stripping whitespace).
            offsets: Mapping of result key → offset kwargs dict.
                Each value is a dict of any combination of ``right``, ``left``,
                ``down``, ``up`` (all default to 0 if omitted).
            sheet: Restrict keyword search to this sheet. Ignored when using
                ``cell_ref``.

        Returns:
            Dict mapping each key from ``offsets`` to the value at that offset.

        Raises:
            ValueError: If both or neither of ``cell_ref``/``keyword`` are given.
            KeywordNotFoundError: Keyword not found in any searched sheet.
            MultipleKeywordsFoundError: Keyword found in more than one cell.

        Example:
            >>> with ExcelCellReader('report.xlsx') as reader:
            ...     values = reader.get_many_relative(
            ...         keyword='Revenue',
            ...         offsets={'value': {'right': 1}, 'label': {'right': 2}},
            ...     )
        """
        if cell_ref is None and keyword is None:
            raise ValueError("Specify either cell_ref= or keyword=")
        if cell_ref is not None and keyword is not None:
            raise ValueError("Cannot specify both cell_ref= and keyword=")

        if keyword is not None:
            anchor_sheet, row, col = self._find_keyword(keyword, sheet)
        else:
            anchor_sheet, cell_addr = self._parse_ref(cell_ref)  # type: ignore[arg-type]
            row, col = coordinate_to_tuple(cell_addr)

        return {
            name: self._value_at_offset(anchor_sheet, row, col, **offset)
            for name, offset in (offsets or {}).items()
        }

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    def _find_keyword(
        self, keyword: str, sheet_name: str | None
    ) -> tuple[Worksheet, int, int]:
        """Scan for a cell whose value exactly matches *keyword* (case-insensitive).

        Returns (worksheet, row, col) of the matching cell.

        Raises:
            KeywordNotFoundError: No matching cell found.
            MultipleKeywordsFoundError: More than one matching cell found.
        """
        normalised = keyword.strip().lower()
        sheets: list[Worksheet]
        if sheet_name is not None:
            if sheet_name not in self.wb.sheetnames:
                raise ExcelSheetNotFoundError(
                    f"Sheet '{sheet_name}' not found in {self.filepath}. "
                    f"Available sheets: {self.wb.sheetnames}"
                )
            sheets = [self.wb[sheet_name]]
        else:
            sheets = list(self.wb.worksheets)

        matches: list[tuple[Worksheet, int, int]] = []
        for ws in sheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and str(cell.value).strip().lower() == normalised:
                        matches.append((ws, cell.row, cell.column))

        if len(matches) == 0:
            scope = f"sheet '{sheet_name}'" if sheet_name else "any sheet"
            raise KeywordNotFoundError(
                f"Keyword '{keyword}' not found in {scope} of {self.filepath}."
            )
        if len(matches) > 1:
            found_in = [
                f"{ws.title}!{ws.cell(r, c).coordinate}"
                for ws, r, c in matches
            ]
            raise MultipleKeywordsFoundError(
                f"Keyword '{keyword}' found in multiple cells: {found_in}. "
                "Use sheet= to restrict the search.",
                found_in=found_in,
            )

        return matches[0]

    def _value_at_offset(
        self,
        ws: Worksheet,
        row: int,
        col: int,
        *,
        right: int = 0,
        left: int = 0,
        down: int = 0,
        up: int = 0,
    ) -> Any:
        """Return the value at (row + down - up, col + right - left)."""
        return ws.cell(row + down - up, col + right - left).value

    def _parse_ref(self, cell_ref: str):
        """Parse 'Sheet1!B5' into (sheet, 'B5'), falling back to active sheet."""
        if "!" in cell_ref:
            sheet_name, cell_addr = cell_ref.split("!", 1)
            if sheet_name not in self.wb.sheetnames:
                raise ExcelSheetNotFoundError(
                    f"Sheet '{sheet_name}' not found in {self.filepath}. "
                    f"Available sheets: {self.wb.sheetnames}"
                )
            return self.wb[sheet_name], cell_addr
        else:
            active = self.wb.active
            if active is None:
                raise ExcelError(
                    f"No active sheet found in {self.filepath}."
                )
            return active, cell_ref
