import os
from openpyxl.utils import coordinate_to_tuple, range_boundaries
import polars as pl
from typing import cast
from types import TracebackType
from excel.exceptions import (
    ExcelTableReaderError,
    TableNotFoundError,
    MultipleTablesFoundError,
    ColumnNamesMismatchError,
)
from excel._utils import load_excel_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell


class ExcelTableReader:
    """Extract tabular data from Excel workbooks as Polars DataFrames.

    Must be used as a context manager::

        with ExcelTableReader('file.xlsx') as reader:
            df = reader.extract_table_by_column_names(['Col1', 'Col2'])
    """

    def __init__(self, filepath: str | os.PathLike[str] | bytes):
        self.filepath = filepath
        self._wb: Workbook | None = None

    @property
    def wb(self) -> Workbook:
        """
        Get loaded workbook.

        Returns:
            Workbook object

        Raises:
            ExcelTableReaderError: If workbook not loaded
        """
        if self._wb is None:
            raise ExcelTableReaderError(
                "Workbook not loaded. Use ExcelTableReader as context manager:\n"
                "    with ExcelTableReader('file.xlsx') as reader:\n"
                "        df = reader.extract_table_by_column_names([...])"
            )
        return self._wb

    def __enter__(self) -> "ExcelTableReader":
        self._wb = load_excel_workbook(self.filepath, data_only=True)
        return self

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_val: BaseException | None,
        exc_tb: TracebackType | None,
    ) -> None:
        if self._wb:
            self._wb.close()

    def extract_table_by_column_names(
        self,
        column_names: list[str],
        exact_columns: bool = False,
        unmerge_cells: bool = True,
        fill_forward: bool = True,
        skip_empty_cols: bool = False,
    ) -> pl.DataFrame:
        """Search all sheets and return the first table containing *column_names*.

        *column_names* are used as a mandatory subset to locate the table — all
        listed columns must be present in the same header row.  By default
        (``exact_columns=False``) the returned DataFrame contains every column
        in the detected range, which may include additional columns beyond those
        listed.  Set ``exact_columns=True`` to require the table to have no
        extra columns.

        Args:
            column_names: Column names that must all appear together in a
                header row.
            exact_columns: If ``True``, return only the columns in *column_names*,
                in the order given, discarding any extra columns in the detected
                range.  If ``False`` (default), all detected columns are returned.
            unmerge_cells: Expand merged cells before extraction.
            fill_forward: Forward-fill null values after extraction.

        Returns:
            Polars DataFrame of the matched table.

        Raises:
            TableNotFoundError: No sheet contains a row with all *column_names*.
            MultipleTablesFoundError: More than one row matches across all sheets.
        """
        for sheet_name in self.wb.sheetnames:
            try:
                df = self.extract_table_by_column_names_from_sheet(
                    column_names,
                    sheet_name,
                    exact_columns,
                    unmerge_cells,
                    fill_forward,
                    skip_empty_cols,
                )
            except TableNotFoundError:
                continue
            else:
                return df

        raise TableNotFoundError(
            f"Could not find table with columns {column_names} "
            f"in any sheet of {self.filepath}"
        )

    def extract_table_by_column_names_from_sheet(
        self,
        column_names: list[str],
        sheet_name: str,
        exact_columns: bool = False,
        unmerge_cells: bool = True,
        fill_forward: bool = True,
        skip_empty_cols: bool = False,
    ) -> pl.DataFrame:
        """Extract from a specific sheet by matching column names.

        *column_names* are used as a mandatory subset to locate the table — all
        listed columns must be present in the same header row.  By default
        (``exact_columns=False``) the returned DataFrame contains every column
        in the detected range.  Set ``exact_columns=True`` to return only the
        columns in *column_names*, discarding any extras.

        Args:
            column_names: Column names that must all appear in the header row.
            sheet_name: Sheet to search.
            exact_columns: If ``True``, return only the columns in *column_names*,
                in the order given.  If ``False`` (default), all detected columns
                are returned.
            unmerge_cells: Expand merged cells before extraction.
            fill_forward: Forward-fill null values after extraction.

        Returns:
            Polars DataFrame of the matched table.

        Raises:
            TableNotFoundError: Sheet does not contain all *column_names*.
            MultipleTablesFoundError: More than one header row matches.
            ExcelTableReaderError: Sheet does not exist, or a column in
                *column_names* appears more than once in the header row.
        """
        sheet = self._get_sheet(sheet_name)

        if unmerge_cells:
            self._unmerge_and_fill_sheet(sheet)

        header_row = self._find_header_row_in_sheet(
            sheet,
            column_names,
        )

        # Find the leftmost column among the matched column_names in the header row.
        # Use first occurrence to avoid picking a duplicate table further right.
        col_positions: dict[str, int] = {}
        for cell in sheet[header_row]:
            if (
                cell.value in set(column_names)
                and cell.column is not None
                and cell.value not in col_positions
            ):
                col_positions[cell.value] = cell.column
        start_col = min(col_positions.values())
        boundaries = self._detect_boundaries_in_sheet(
            sheet, header_row, start_col, has_headers=True, skip_empty_cols=skip_empty_cols
        )
        # Extend max_col to cover any requested column beyond the auto-detected
        # right boundary (handles non-contiguous headers separated by gaps).
        rightmost_requested = max(col_positions.values())
        min_row, min_col, max_row, max_col = boundaries
        boundaries = (min_row, min_col, max_row, max(max_col, rightmost_requested))

        self._check_no_duplicate_columns(
            sheet, header_row, column_names, boundaries[1], boundaries[3]
        )

        df = self._extract_range_from_sheet(
            sheet,
            *boundaries,
            has_headers=True,
            column_names=None,
        )

        if fill_forward and len(df) > 0:
            df = df.with_columns([pl.col(col).forward_fill() for col in df.columns])

        if exact_columns:
            df = df.select(column_names)

        return df

    def extract_table_by_range(
        self,
        range_str: str,
        sheet: str,
        has_headers: bool = True,
        column_names: list[str] | None = None,
        dynamic: bool = False,
        unmerge_cells: bool = True,
        fill_forward: bool = True,
        max_empty_rows: int = 2,
        stop_at: str | None = None,
        stop_before: str | None = None,
        skip_empty_cols: bool = False,
    ) -> pl.DataFrame:
        """Extract a table from an Excel range string.

        Two modes are supported:

        * **Exact** (``dynamic=False``, default) — the range is treated as a
          fixed rectangle.  Both column and row boundaries come from
          *range_str*.
        * **Dynamic** (``dynamic=True``) — only the column span is taken from
          *range_str*; the row boundary is auto-detected downward from the
          start row.  Useful when the column layout is fixed but the number of
          rows varies.  Pass a single-row range (e.g. ``"A1:D1"``) to pin just
          the header row and let the data rows grow freely.  ``max_empty_rows``,
          ``stop_at``, and ``stop_before`` are respected only in this mode.

        Args:
            range_str: Excel range, e.g. ``"A5:C20"`` or ``"A1:D1"``.
            sheet: Sheet name (required).
            has_headers: Whether the first row of the range is a header row.
            column_names: Manual column names when *has_headers* is ``False``;
                auto-generates ``col_0``, ``col_1``… if omitted.
            dynamic: If ``True``, fix the column span from *range_str* but
                auto-detect the bottom row boundary.
            unmerge_cells: Expand merged cells before extraction.
            fill_forward: Forward-fill null values after extraction.
            max_empty_rows: (dynamic only) Stop after this many consecutive
                empty rows (ignored when *stop_at* or *stop_before* is set).
            stop_at: (dynamic only) Stop when any cell in the row equals this
                value, including that row in the result.
            stop_before: (dynamic only) Stop when any cell in the row equals
                this value, excluding that row from the result.

        Returns:
            Polars DataFrame.

        Raises:
            ValueError: Both *stop_at* and *stop_before* are provided.
            ExcelTableReaderError: Sheet does not exist.

        Examples:
            Exact range::

                df = reader.extract_table_by_range("A5:C20", sheet="Data")

            Dynamic columns — pin header row, auto-detect data rows::

                df = reader.extract_table_by_range(
                    "A1:D1", sheet="Sheet1", dynamic=True
                )

            Dynamic with stop marker::

                df = reader.extract_table_by_range(
                    "A1:D1", sheet="Sheet1", dynamic=True, stop_before="Total"
                )
        """
        if stop_at is not None and stop_before is not None:
            raise ValueError("Provide either stop_at or stop_before, not both.")

        sheet_obj = self._get_sheet(sheet)

        if unmerge_cells:
            self._unmerge_and_fill_sheet(sheet_obj)

        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        assert min_row is not None
        assert min_col is not None
        assert max_row is not None
        assert max_col is not None

        if dynamic:
            # Fix the column span; detect the row boundary dynamically.
            _, _, detected_max_row, _ = self._detect_boundaries_in_sheet(
                sheet_obj,
                min_row,
                min_col,
                has_headers=has_headers,
                max_empty_rows=max_empty_rows,
                stop_at=stop_at,
                stop_before=stop_before,
                skip_empty_cols=skip_empty_cols,
            )
            max_row = detected_max_row

        df = self._extract_range_from_sheet(
            sheet_obj,
            min_row,
            min_col,
            max_row,
            max_col,
            has_headers=has_headers,
            column_names=column_names,
        )

        if fill_forward and len(df) > 0:
            df = df.with_columns([pl.col(col).forward_fill() for col in df.columns])

        return df

    def extract_table_from_cell(
        self,
        start_cell: str,
        sheet: str,
        has_headers: bool = True,
        column_names: list[str] | None = None,
        unmerge_cells: bool = True,
        fill_forward: bool = True,
        max_empty_rows: int = 2,
        stop_at: str | None = None,
        stop_before: str | None = None,
        skip_empty_cols: bool = False,
    ) -> pl.DataFrame:
        """
        Extract table starting from cell, auto-detecting boundaries.

        Args:
            start_cell: Starting cell (e.g., "A5")
            sheet: Sheet name (REQUIRED)
            has_headers: First row is headers
            column_names: Manual column names
            unmerge_cells: Unmerge cells
            fill_forward: Fill forward
            max_empty_rows: Stop after this many consecutive empty rows
                (ignored when stop_at or stop_before is set)
            stop_at: Stop when any cell in the row equals this value,
                including that row in the result.
            stop_before: Stop when any cell in the row equals this value,
                excluding that row from the result.

        Returns:
            Polars DataFrame

        Raises:
            ValueError: If both stop_at and stop_before are provided.

        Example:
            >>> df = extractor.extract_table_from_cell('A5', sheet='Data')
            >>> df = extractor.extract_table_from_cell(
            ...     'A1', sheet='Sheet1', stop_at='Total'
            ... )
        """
        if stop_at is not None and stop_before is not None:
            raise ValueError("Provide either stop_at or stop_before, not both.")

        sheet_obj = self._get_sheet(sheet)

        if unmerge_cells:
            self._unmerge_and_fill_sheet(sheet_obj)

        start_row, start_col = coordinate_to_tuple(start_cell)

        boundaries = self._detect_boundaries_in_sheet(
            sheet_obj,
            start_row,
            start_col,
            has_headers,
            max_empty_rows,
            stop_at=stop_at,
            stop_before=stop_before,
            skip_empty_cols=skip_empty_cols,
        )

        df = self._extract_range_from_sheet(
            sheet_obj,
            *boundaries,
            has_headers=has_headers,
            column_names=column_names,
        )

        if fill_forward and len(df) > 0:
            df = df.with_columns([pl.col(col).forward_fill() for col in df.columns])

        return df

    def extract_table_near(
        self,
        column_names: list[str],
        sheet: str,
        ref_cell: str | None = None,
        keyword: str | None = None,
        exact_columns: bool = False,
        unmerge_cells: bool = True,
        fill_forward: bool = True,
        stop_at: str | None = None,
        stop_before: str | None = None,
        skip_empty_cols: bool = False,
    ) -> pl.DataFrame:
        """Find and extract a table by scanning from an anchor for matching column headers.

        Locates the anchor via *ref_cell* or *keyword*, then scans downward from
        that row to find the first row containing all *column_names* as headers.
        Exactly one of *ref_cell* or *keyword* must be provided.

        * ``ref_cell`` — start scanning from this A1 cell address.
        * ``keyword`` — locate the anchor by finding a cell whose value exactly
          equals *keyword* (case-sensitive). Raises if 0 or >1 matches.

        Args:
            column_names: Column names to locate the header row.
            sheet: Sheet name (REQUIRED).
            ref_cell: A1 address to start scanning from (e.g. ``"B2"``).
            keyword: Exact cell value to use as the anchor (e.g. ``"Sales"``).
            exact_columns: If ``True``, return only the columns in *column_names*,
                in the order given.  If ``False`` (default), all detected columns
                are returned.
            unmerge_cells: Unmerge cells before extraction.
            fill_forward: Forward-fill null values after extraction.
            stop_at: Stop when any cell in the row equals this value,
                including that row in the result.
            stop_before: Stop when any cell in the row equals this value,
                excluding that row from the result.

        Returns:
            Polars DataFrame.

        Raises:
            ValueError: If both or neither of *ref_cell* / *keyword* are given,
                or if both *stop_at* and *stop_before* are given.
            TableNotFoundError: If the anchor or the column headers are not found.
            MultipleTablesFoundError: If *keyword* or the header row matches more
                than once.
        """
        if (ref_cell is None) == (keyword is None):
            raise ValueError("Provide exactly one of ref_cell or keyword.")
        if stop_at is not None and stop_before is not None:
            raise ValueError("Provide either stop_at or stop_before, not both.")

        sheet_obj = self._get_sheet(sheet)

        if unmerge_cells:
            self._unmerge_and_fill_sheet(sheet_obj)

        if ref_cell is not None:
            anchor_row, anchor_col = coordinate_to_tuple(ref_cell)
        else:
            matches: list[tuple[int, int]] = [
                (cell.row, cell.column)
                for row in sheet_obj.iter_rows()
                for cell in row
                if cell.value == keyword
                and cell.row is not None
                and cell.column is not None
            ]
            if len(matches) == 0:
                raise TableNotFoundError(
                    f"No cell with value {keyword!r} found in sheet '{sheet}' "
                    f"of {self.filepath}"
                )
            if len(matches) > 1:
                locations = [
                    f"{sheet}!{sheet_obj.cell(r, c).coordinate}" for r, c in matches
                ]
                raise MultipleTablesFoundError(
                    f"Multiple cells with value {keyword!r} found in sheet "
                    f"'{sheet}': {locations}",
                    found_in=locations,
                )
            anchor_row, anchor_col = matches[0]

        header_row = self._find_header_row_in_sheet(
            sheet_obj, column_names, start_row=anchor_row
        )

        # Use first occurrence of each name at or after anchor_col — this ensures
        # we pick the correct table when duplicate column names exist further left.
        col_positions: dict[str, int] = {}
        for cell in sheet_obj[header_row]:
            if (
                cell.value in set(column_names)
                and cell.column is not None
                and cell.column >= anchor_col
                and cell.value not in col_positions
            ):
                col_positions[cell.value] = cell.column
        start_col = min(col_positions.values())
        boundaries = self._detect_boundaries_in_sheet(
            sheet_obj,
            header_row,
            start_col,
            has_headers=True,
            stop_at=stop_at,
            stop_before=stop_before,
            skip_empty_cols=skip_empty_cols,
        )
        # Extend max_col to cover any requested column beyond the auto-detected
        # right boundary (handles non-contiguous headers separated by gaps).
        rightmost_requested = max(col_positions.values())
        min_row, min_col, max_row, max_col = boundaries
        boundaries = (min_row, min_col, max_row, max(max_col, rightmost_requested))

        self._check_no_duplicate_columns(
            sheet_obj, header_row, column_names, boundaries[1], boundaries[3]
        )

        df = self._extract_range_from_sheet(
            sheet_obj,
            *boundaries,
            has_headers=True,
            column_names=None,
        )

        if fill_forward and len(df) > 0:
            df = df.with_columns([pl.col(col).forward_fill() for col in df.columns])

        if exact_columns:
            df = df.select(column_names)

        return df

    # ==================== Utility Methods (INTERNAL) ====================

    def _check_no_duplicate_columns(
        self,
        sheet: Worksheet,
        header_row: int,
        column_names: list[str],
        min_col: int,
        max_col: int,
    ) -> None:
        """Raise if any column in *column_names* appears more than once within [min_col, max_col]."""
        row_values = [
            sheet.cell(header_row, col).value for col in range(min_col, max_col + 1)
        ]
        duplicates = [col for col in column_names if row_values.count(col) > 1]
        if duplicates:
            raise ExcelTableReaderError(
                f"Columns {duplicates} appear more than once in row {header_row}. "
                f"Cannot determine which column to use."
            )

    def _get_sheet(self, sheet_name: str) -> Worksheet:
        """
        Get sheet, with validation.

        Args:
            sheet_name: Sheet name (if None, uses active sheet)

        Returns:
            Worksheet object

        Raises:
            ExcelDataExtractorError: If sheet not found
        """

        if sheet_name not in self.wb.sheetnames:
            raise ExcelTableReaderError(
                f"Sheet '{sheet_name}' not found in {self.filepath}. "
                f"Available sheets: {self.wb.sheetnames}"
            )

        return self.wb[sheet_name]

    def _unmerge_and_fill_sheet(self, sheet: Worksheet) -> None:
        """
        Unmerge all merged cells in the sheet and fill values forward.

        Args:
            sheet: Worksheet object
        """
        merged_ranges = list(sheet.merged_cells.ranges)

        for merged_range in merged_ranges:
            # Get bounds of merged range
            min_col, min_row, max_col, max_row = merged_range.bounds

            # Get value from top-left cell (where value is stored)
            value = sheet.cell(min_row, min_col).value

            # Unmerge
            sheet.unmerge_cells(str(merged_range))

            # Fill value to all cells in the range
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    cast(Cell, sheet.cell(row, col)).value = value

    def _find_table_start(
        self,
        sheet: Worksheet,
        ref_row: int,
        ref_col: int,
    ) -> tuple[int, int] | None:
        """
        Scan right and downward from (ref_row, ref_col) to find the first non-empty cell.

        Args:
            sheet: Worksheet object
            ref_row: Starting row (1-indexed)
            ref_col: Starting column (1-indexed)

        Returns:
            (row, col) of first non-empty cell, or None if not found
        """
        for row in range(ref_row, sheet.max_row + 1):
            for col in range(ref_col, sheet.max_column + 1):
                value = sheet.cell(row, col).value
                if value is not None and value != "":
                    return (row, col)
        return None

    def _detect_boundaries_in_sheet(
        self,
        sheet: Worksheet,
        start_row: int,
        start_col: int,
        has_headers: bool = True,
        max_empty_rows: int = 2,
        stop_at: str | None = None,
        stop_before: str | None = None,
        skip_empty_cols: bool = False,
    ) -> tuple[int, int, int, int]:
        """
        Auto-detect table boundaries from a starting point.

        Args:
            sheet: Worksheet object
            start_row: Starting row number (1-indexed)
            start_col: Starting column number (1-indexed)
            has_headers: Whether first row is headers
            max_empty_rows: Stop after this many consecutive empty rows
                (ignored when stop_at or stop_before is set)
            stop_at: Stop when any cell in the row matches this value,
                including that row.
            stop_before: Stop when any cell in the row matches this value,
                excluding that row.
            skip_empty_cols: If ``True``, empty cells in the header row do not
                stop the column scan — scanning continues past them.

        Returns:
            Tuple of (min_row, min_col, max_row, max_col)
        """
        # Detect right boundary (columns)
        max_col = start_col
        for col in range(start_col, min(start_col + 100, sheet.max_column + 1)):
            cell_value = sheet.cell(start_row, col).value
            if cell_value is None or cell_value == "":
                if not skip_empty_cols:
                    break
            else:
                max_col = col

        # Detect bottom boundary (rows)
        data_start_row = start_row + 1 if has_headers else start_row
        max_row = start_row
        consecutive_empty = 0
        anchor = stop_at or stop_before
        use_anchor = anchor is not None

        for row in range(data_start_row, sheet.max_row + 1):
            row_values = [
                sheet.cell(row, col).value for col in range(start_col, max_col + 1)
            ]

            if use_anchor:
                if anchor in row_values:
                    if stop_at is not None:
                        max_row = row
                    # stop_before: do not update max_row
                    break
                max_row = row
            else:
                has_data = any(v is not None and v != "" for v in row_values)
                if has_data:
                    max_row = row
                    consecutive_empty = 0
                else:
                    consecutive_empty += 1
                    if consecutive_empty >= max_empty_rows:
                        break

        return (start_row, start_col, max_row, max_col)

    def _extract_range_from_sheet(
        self,
        sheet: Worksheet,
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
        has_headers: bool = True,
        column_names: list[str] | None = None,
    ) -> pl.DataFrame:
        """
        Extract data from specific cell range.

        Args:
            sheet: Worksheet object
            min_row, min_col: Top-left corner (1-indexed)
            max_row, max_col: Bottom-right corner (1-indexed)
            has_headers: Whether first row contains headers
            column_names: Manual column names if has_headers=False

        Returns:
            Polars DataFrame
        """
        # Step 1: Get headers
        headers: list[str]
        if has_headers:
            headers = [
                str(sheet.cell(min_row, col).value)
                for col in range(min_col, max_col + 1)
            ]
            data_start = min_row + 1
        else:
            n_cols = max_col - min_col + 1
            # Use provided column names or generate default
            if column_names:
                if len(column_names) != n_cols:
                    raise ColumnNamesMismatchError(
                        f"column_names has {len(column_names)} entries but the range "
                        f"spans {n_cols} columns. Provide exactly {n_cols} names.",
                        expected=n_cols,
                        got=len(column_names),
                    )
                headers = column_names
            else:
                headers = [f"col_{i}" for i in range(n_cols)]
            data_start = min_row

        # Step 2: Get data rows
        data = []
        for row in range(data_start, max_row + 1):
            row_data = [
                sheet.cell(row, col).value for col in range(min_col, max_col + 1)
            ]
            data.append(row_data)

        # Step 3: Create DataFrame
        if not data:
            # Empty table - return empty DataFrame with schema
            return pl.DataFrame(schema={h: pl.Utf8 for h in headers})

        return pl.DataFrame(data, schema=headers, orient="row")

    def _find_header_row_in_sheet(
        self,
        sheet: Worksheet,
        column_names: list[str],
        start_row: int = 1,
    ) -> int:
        """
        Find the first row (at or after *start_row*) containing all *column_names*.

        Args:
            sheet: Worksheet object.
            column_names: Column names that must all appear in the row.
            start_row: Row to start searching from (1-indexed).

        Returns:
            Row number (1-indexed).

        Raises:
            TableNotFoundError: No row contains all *column_names*.
            MultipleTablesFoundError: More than one row matches.
        """
        matching_rows: list[int] = []

        for row_idx in range(start_row, sheet.max_row + 1):
            row_values = [
                cell.value for cell in sheet[row_idx] if cell.value is not None
            ]
            if all(col in row_values for col in column_names):
                matching_rows.append(row_idx)

        if not matching_rows:
            raise TableNotFoundError(
                f"Could not find table with columns: {column_names} "
                f"in sheet '{sheet.title}'"
            )

        if len(matching_rows) > 1:
            raise MultipleTablesFoundError(
                f"Column names {column_names} were found in multiple rows: {matching_rows}. "
                f"Cannot determine which table to extract.",
                found_in=[str(r) for r in matching_rows],
            )

        return matching_rows[0]
