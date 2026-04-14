from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from os import PathLike
from typing import Any, Literal, cast

import polars as pl
from openpyxl.cell.cell import MergedCell as _MC
from openpyxl.utils import coordinate_to_tuple
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from excel._utils import load_excel_workbook
from excel.protocols import TypedValue
from excel.template_reader import ExcelTemplateReader, MarkedCell
from excel._types import (
    JoinMode,
    InsertMode,
    StyleSource,
    OrderBy,
    FillSpec,
    TableMeta,
    ImageMeta,
    _EndTableMarker,
    _apply_fill,
    _is_loop,
    _is_table,
    _is_image,
    _FILL_MISSING,
    _VALID_JOIN_MODES,
    _VALID_STYLE_SOURCES,
    _TemplateRegex,
)


class ExcelTemplateWriter:
    """Fill an Excel template with data and write the output file.

    Template syntax:

    * ``{{ variable }}`` — scalar placeholder; replaced with a single value.
    * ``{{ variable | loop() }}`` — loop row tag; marks cells for row expansion.
      All loop-tagged cells in the same row must reference list variables of the
      same length.  The template row is duplicated N times (one per list element).
    * ``{{ variable | table(...) }}`` — table tag; fills a structured data region.

      Table parameters:

      - ``join=[left|inner|outer|right]`` — join strategy (default ``left``).
      - ``on=ColName`` — override the join column (default: inferred from the header
        to the left of the tag cell).
      - ``order_by=[asc|desc|ColName[:asc|:desc]]`` — sort the outer-join upper zone
        by column (``outer`` only).
      - ``fill=[value|col1:<v1>;col2:<v2>]`` — fill ``None`` values globally or
        per-column.
      - ``positional=True`` — fill by position, no join column or header matching.
      - ``placeholder=true`` — tag row is used as a style source for inserted rows,
        then deleted from the output if its join value is absent from the DataFrame.
        Applies to ``join=outer`` only.
      - ``style=first|last`` (default ``last``) — which template row to copy styles
        from when inserting new rows.  ``first`` uses the tag row; ``last`` uses the
        last template row.

    * ``{{ end_table }}`` — marks the table boundary; inserted rows go after.
    * ``{{ end_table | insert=above }}`` — marks the table boundary; inserted rows
      go before this row.
    * ``{{ insert_data }}`` — marks the insertion point for outer-join extras.

    Record access (scalar only):

    * ``{{ record_var.column_name }}`` — extract a single cell from a 1-row DataFrame.

    Usage::

        writer = ExcelTemplateWriter("template.xlsx")
        writer.write(
            {
                "title": TypedValue("Q1 Report", "single"),
                "month": TypedValue(["Jan", "Feb", "Mar"], "list"),
                "sales": TypedValue(sales_df, "table"),
            },
            "output.xlsx",
        )
    """

    def __init__(self, template: str | PathLike[str] | bytes):
        self._template = template

    def write(self, vars: dict[str, TypedValue], file: str | PathLike[str]) -> None:
        """Fill the template and save to *file*.

        Args:
            vars: Variable name → TypedValue.
            file: Output path for the filled workbook.

        Raises:
            KeyError: A template tag references a variable not present in *vars*.
            ValueError: Loop cells in the same row have lists of different lengths.
        """
        wb = load_excel_workbook(self._template)
        structure = ExcelTemplateReader().read(self._template)

        loop_rows, tables, scalars, images = self._categorize_template_cells(structure)
        self._fill_scalar_cells(wb, scalars, vars)
        self._fill_image_cells(wb, images, vars)
        self._fill_loop_rows(wb, loop_rows, vars)
        self._check_table_collisions(wb, tables, vars)
        self._fill_table_cells(wb, tables, vars)

        wb.save(file)

    def _categorize_template_cells(
        self, structure: dict[str, list[MarkedCell]]
    ) -> tuple[
        dict[tuple[str, int], list[MarkedCell]],
        list[MarkedCell],
        list[MarkedCell],
        list[MarkedCell],
    ]:
        """Split tagged cells into loop participants, tables, scalars, and images.
        
        Returns:
            (loop_rows, tables, scalars, images) where loop_rows is keyed by (sheet_name, row_number).
        """
        loop_rows: dict[tuple[str, int], list[MarkedCell]] = {}
        tables: list[MarkedCell] = []
        scalars: list[MarkedCell] = []
        images: list[MarkedCell] = []

        for sheet_name, cells in structure.items():
            for cell in cells:
                if cell.name in ("end_table", "insert_data"):
                    continue  # structural marker, not a variable
                if _is_table(cell):
                    tables.append(cell)
                elif _is_loop(cell):
                    row_n = coordinate_to_tuple(cell.cell_addr)[0]
                    loop_rows.setdefault((sheet_name, row_n), []).append(cell)
                elif _is_image(cell):
                    images.append(cell)
                else:
                    scalars.append(cell)

        return loop_rows, tables, scalars, images

    def _fill_scalar_cells(
        self, wb: Workbook, scalars: list[MarkedCell], vars: dict[str, TypedValue]
    ) -> None:
        """Write scalar cells to the worksheet.
        
        Scalars are written FIRST—before any row insertions from loop or table
        processing. openpyxl shifts all cells below an insert_rows() call, so values
        already written here will ride along to their correct final positions.
        """
        for mc in scalars:
            ws = wb[mc.sheet]
            row_n, col_n = coordinate_to_tuple(mc.cell_addr)
            if "." in mc.name:
                var_name, col_name = mc.name.split(".", 1)
                tv = vars[var_name]
                if tv.value.height != 1:
                    raise ValueError(
                        f"Record variable '{var_name}' has {tv.value.height} rows;"
                        " expected exactly 1"
                    )
                ws.cell(row_n, col_n).value = tv.value[col_name][0]
            else:
                ws.cell(row_n, col_n).value = vars[mc.name].value

    def _fill_image_cells(
        self, wb: Workbook, images: list[MarkedCell], vars: dict[str, TypedValue]
    ) -> None:
        """Anchor images at their tagged cells, clearing the placeholder text.

        The variable value must be a file path (``str`` or ``PathLike``) or a
        ``PIL.Image.Image`` instance accepted by openpyxl.

        Dimensions from the tag (``width=`` / ``height=`` in pixels) override the
        image's natural size.  Omitting a dimension keeps the natural value.
        """
        from openpyxl.drawing.image import Image as _OxlImage

        for mc in images:
            ws = wb[mc.sheet]
            row_n, col_n = coordinate_to_tuple(mc.cell_addr)
            ws.cell(row_n, col_n).value = None

            img_meta = ImageMeta.from_cell(mc)
            img = _OxlImage(vars[mc.name].value)
            if img_meta.width is not None:
                img.width = img_meta.width
            if img_meta.height is not None:
                img.height = img_meta.height
            ws.add_image(img, mc.cell_addr)

    def _fill_loop_rows(
        self,
        wb: Workbook,
        loop_rows: dict[tuple[str, int], list[MarkedCell]],
        vars: dict[str, TypedValue],
    ) -> None:
        """Expand and duplicate loop rows based on list variable lengths.
        
        Processes each sheet's rows bottom-up so row inserts don't invalidate
        later row indices within the same sheet.
        """
        by_sheet: dict[str, list[tuple[int, list[MarkedCell]]]] = {}
        for (sheet_name, row_n), cells in loop_rows.items():
            by_sheet.setdefault(sheet_name, []).append((row_n, cells))

        for sheet_name, rows in by_sheet.items():
            ws = wb[sheet_name]
            for row_n, cells in sorted(rows, key=lambda x: x[0], reverse=True):
                n = len(vars[cells[0].name].value)

                for mc in cells[1:]:
                    actual = len(vars[mc.name].value)
                    if actual != n:
                        raise ValueError(
                            f"Loop variables in row {row_n} of '{sheet_name}' have "
                            f"different lengths: '{cells[0].name}'={n}, "
                            f"'{mc.name}'={actual}"
                        )

                if n == 0:
                    ws.delete_rows(row_n)
                    continue

                if n > 1:
                    _copy_row_styles(ws, row_n, n - 1)

                for i in range(n):
                    for mc in cells:
                        _, col_n = coordinate_to_tuple(mc.cell_addr)
                        ws.cell(row_n + i, col_n).value = vars[mc.name].value[i]

    def _check_table_collisions(
        self, wb: Workbook, tables: list[MarkedCell], vars: dict[str, TypedValue]
    ) -> None:
        """Validate that no two table operations on the same sheet overlap."""
        regions_by_sheet: dict[str, list[tuple[str, tuple[int, int, int, int]]]] = {}
        for mc in tables:
            ws = wb[mc.sheet]
            df = vars[mc.name].value
            region = _compute_table_region(ws, mc, df)
            regions_by_sheet.setdefault(mc.sheet, []).append((f"table({mc.name})", region))
        for sheet_name, regions in regions_by_sheet.items():
            _check_region_collisions(regions)

    def _fill_table_cells(
        self, wb: Workbook, tables: list[MarkedCell], vars: dict[str, TypedValue]
    ) -> None:
        """Fill all table regions with data.
        
        Processes tables bottom-up per sheet so outer row insertions don't corrupt
        later table positions in the same sheet.
        """
        tables_by_sheet: dict[str, list[MarkedCell]] = {}
        for mc in tables:
            tables_by_sheet.setdefault(mc.sheet, []).append(mc)

        for sheet_name, mcs in tables_by_sheet.items():
            ws = wb[sheet_name]
            for mc in sorted(
                mcs, key=lambda c: coordinate_to_tuple(c.cell_addr)[0], reverse=True
            ):
                tm = TableMeta.from_cell(mc)
                if tm.positional:
                    _fill_positional(ws, mc, vars[mc.name].value)
                else:
                    _fill_table(ws, mc, vars[mc.name].value)



# ── 1. openpyxl cell and merge helpers ─────────────────────────────────────


def _get_style_source(ws: Worksheet, row: int, col: int):
    """Return the cell that owns the style for *(row, col)*.

    Inside a merged range only the top-left cell carries style data;
    every other cell is a MergedCell proxy with ``has_style == False``.
    This helper resolves to the top-left cell when needed.
    """
    cell = ws.cell(row, col)
    if cell.has_style:
        return cell
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return ws.cell(m.min_row, m.min_col)
    return cell


def _get_merged_cell_value(ws: Worksheet, row: int, col: int) -> Any:
    """Return the effective value of a cell, resolving merged ranges.

    In normal (non-read-only) mode openpyxl returns ``None`` for every cell
    inside a merged range except the top-left corner.  This helper checks
    whether *(row, col)* belongs to a merged range and, if so, returns the
    value stored in the top-left cell of that range.
    """
    val = ws.cell(row, col).value
    if val is not None:
        return val
    for m in ws.merged_cells.ranges:
        if m.min_row <= row <= m.max_row and m.min_col <= col <= m.max_col:
            return ws.cell(m.min_row, m.min_col).value
    return None


def _safe_remove_merge(ws: Worksheet, merge_range: Any) -> None:
    """Remove a merged range from the registry without crashing on missing cells.

    openpyxl's ws.unmerge_cells() deletes all cell objects inside the range
    from ws._cells.  After insert_rows() the cell objects may have already been
    moved or may never have existed as real cells, causing a KeyError.  This
    helper removes the range from the registry directly and then purges only the
    MergedCell proxy objects that actually exist in ws._cells, so the caller can
    safely follow up with ws.merge_cells() or write values.
    """
    ws.merged_cells.ranges.discard(merge_range)
    for r in range(merge_range.min_row, merge_range.max_row + 1):
        for c in range(merge_range.min_col, merge_range.max_col + 1):
            cell = ws._cells.get((r, c))
            if isinstance(cell, _MC):
                del ws._cells[(r, c)]


def _sync_merges_after_delete(ws: Worksheet, deleted_row: int) -> None:
    """Correct the merged-cells registry after ``delete_rows()``.

    openpyxl 3.x shifts cell *data* when rows are deleted but does **not**
    update the merged-cells registry.  Every merge whose ``min_row`` is
    greater than ``deleted_row`` ends up pointing one row too high.  This
    helper removes each stale entry and re-registers it at the correct
    post-delete position so that subsequent operations (especially
    ``_copy_row_styles``) see accurate row numbers.
    """
    updated: list[tuple[int, int, int, int]] = []
    # Non-top-left border data keyed by (new_min_r, min_c, new_max_r, max_c)
    # so it can be restored after ws.merge_cells() re-creates the MergedCell
    # proxies with blank style.  delete_rows() has already physically shifted
    # the cell objects to their new positions before this function runs.
    nontopleft_border_restore: dict[tuple[int, int, int, int], dict[tuple[int, int], Any]] = {}
    for m in list(ws.merged_cells.ranges):
        if m.max_row < deleted_row:
            continue  # fully above the deleted row — no change needed
        if m.min_row > deleted_row:
            # Fully below: shift both bounds up by 1.
            # IMPORTANT: delete_rows() has already shifted ws._cells, so the
            # ghost cells that belonged to this merge now live at the *shifted*
            # positions (min_row-1 … max_row-1), not at the stale registry
            # positions (min_row … max_row).  Using _safe_remove_merge (which
            # iterates the stale positions) would purge interior ghosts of the
            # adjacent merge sitting just below this one, and leave this
            # merge's own ghosts uncleaned — causing the re-registered range to
            # be silently dropped on save when adjacent same-column-span merges
            # are present.  Purge at the correct shifted positions instead.
            new_min_r = m.min_row - 1
            new_max_r = m.max_row - 1
            # Snapshot non-TL cell borders at their already-shifted positions
            # before the purge loop destroys them.
            nontl: dict[tuple[int, int], Any] = {}
            for r in range(new_min_r, new_max_r + 1):
                for c in range(m.min_col, m.max_col + 1):
                    if r == new_min_r and c == m.min_col:
                        continue  # top-left: real Cell, not purged
                    cell = ws._cells.get((r, c))
                    if isinstance(cell, _MC) and cell.has_style:
                        nontl[(r - new_min_r, c - m.min_col)] = copy(cell.border)
            if nontl:
                nontopleft_border_restore[(new_min_r, m.min_col, new_max_r, m.max_col)] = nontl
            ws.merged_cells.ranges.discard(m)
            for r in range(new_min_r, new_max_r + 1):
                for c in range(m.min_col, m.max_col + 1):
                    if isinstance(ws._cells.get((r, c)), _MC):
                        del ws._cells[(r, c)]
            updated.append((new_min_r, m.min_col, new_max_r, m.max_col))
        elif m.max_row >= deleted_row:
            # Spans the deleted row: shrink max_row by 1.
            # Rows above deleted_row are unshifted so _safe_remove_merge's
            # stale-coord ghost purge is correct for those rows.
            _safe_remove_merge(ws, m)
            new_max = m.max_row - 1
            if new_max >= m.min_row:
                updated.append((m.min_row, m.min_col, new_max, m.max_col))
            # else: merge collapses to nothing — just discard
    for min_r, min_c, max_r, max_c in updated:
        ws.merge_cells(
            start_row=min_r, start_column=min_c,
            end_row=max_r, end_column=max_c,
        )
        # Restore non-TL borders after merge_cells created fresh proxies.
        nontl = nontopleft_border_restore.get((min_r, min_c, max_r, max_c))
        if nontl:
            for (dr, dc), saved_border in nontl.items():
                target_r, target_c = min_r + dr, min_c + dc
                mc = ws._cells.get((target_r, target_c))
                if mc is not None:
                    mc.border = copy(saved_border)



def _copy_row_styles(
    ws: Worksheet, source_row: int, count: int, *, style_from: int | None = None
) -> None:
    """Insert *count* rows below *source_row*, copying values and styles from it.

    If *style_from* is provided, inserted rows inherit their styles from that row
    instead of from *source_row*.  The insertion position is always below *source_row*.
    This decouples the 'where to insert' from the 'which row to style after', which
    is required for ``style=first`` (insert after last template row but style from tag).

    Handles merged cells carefully: snapshots merges and styles before insert_rows(),
    then reconstructs/splits merges that span the insertion point. openpyxl auto-extends
    spanning merges, which is incorrect for data rows — this function prevents that by
    splitting them into top and bottom portions with unmerged insertion space between.

    MergedCell proxies are purged from inserted rows to ensure all cells are writable.
    """
    # Snapshot merged ranges before insert — openpyxl's insert_rows() auto-extends
    # any merged range that spans the insertion point, which would incorrectly
    # merge the newly inserted data rows.
    saved_merges = [
        (m.min_row, m.min_col, m.max_row, m.max_col)
        for m in ws.merged_cells.ranges
    ]

    max_col = ws.max_column

    # Snapshot source-row styles BEFORE insert_rows() to protect against
    # MergedCell proxy issues: openpyxl may create temporary proxies that later
    # get purged, leaving cells without their original styles. Save now, restore after.
    saved_styles: dict[int, dict[str, Any]] = {}
    for col in range(1, max_col + 1):
        style_cell = _get_style_source(ws, source_row, col)
        if style_cell.has_style:
            saved_styles[col] = {
                "font": copy(style_cell.font),
                "border": copy(style_cell.border),
                "fill": copy(style_cell.fill),
                "number_format": style_cell.number_format,
                "protection": copy(style_cell.protection),
                "alignment": copy(style_cell.alignment),
            }

    # Snapshot styles for inserted rows — from style_from if provided, else source_row.
    insert_styles: dict[int, dict[str, Any]]
    if style_from is not None:
        insert_styles = {}
        for col in range(1, max_col + 1):
            style_cell = _get_style_source(ws, style_from, col)
            if style_cell.has_style:
                insert_styles[col] = {
                    "font": copy(style_cell.font),
                    "border": copy(style_cell.border),
                    "fill": copy(style_cell.fill),
                    "number_format": style_cell.number_format,
                    "protection": copy(style_cell.protection),
                    "alignment": copy(style_cell.alignment),
                }
    else:
        insert_styles = saved_styles

    # Snapshot values from merged cells BELOW source_row before insert_rows().
    # After shifting, the top-left Cell of a merge may be replaced by a MergedCell
    # ghost with read-only .value; save the real top-left value now to restore later.
    merged_cell_styles: dict[tuple[int, int, int, int], dict[str, Any]] = {}
    merged_cell_values: dict[tuple[int, int, int, int], Any] = {}
    for m in ws.merged_cells.ranges:
        if m.min_row >= source_row:  # >= captures merges starting exactly at source_row
            key = (m.min_row, m.min_col, m.max_row, m.max_col)
            merged_cell_values[key] = ws.cell(m.min_row, m.min_col).value
            style_cell = ws.cell(m.min_row, m.min_col)
            if style_cell.has_style:
                merged_cell_styles[key] = {
                    "font": copy(style_cell.font),
                    "border": copy(style_cell.border),
                    "fill": copy(style_cell.fill),
                    "number_format": style_cell.number_format,
                    "protection": copy(style_cell.protection),
                    "alignment": copy(style_cell.alignment),
                }

    # Snapshot border data for every non-top-left cell of fully-below merged
    # ranges.  Excel-created files store per-cell border info on every cell
    # within a merge so that each edge of the merged rectangle renders
    # correctly (e.g. the right border of a 2-column merge lives on the
    # rightmost column's cell, not on the top-left).  _safe_remove_merge purges
    # these MergedCell objects, destroying that data.  Capture it here—keyed by
    # (dr, dc) offsets from the top-left—so it can be rewritten onto the
    # freshly created MergedCell proxies after ws.merge_cells() below.
    # Key: (min_r, min_c, max_r, max_c); value: {(dr, dc): Border}
    merged_nontopleft_borders: dict[tuple[int, int, int, int], dict[tuple[int, int], Any]] = {}
    for min_r, min_c, max_r, max_c in saved_merges:
        if min_r <= source_row:
            continue  # spanning or at-source merges — skip
        per_cell: dict[tuple[int, int], Any] = {}
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                if r == min_r and c == min_c:
                    continue  # top-left handled by merged_cell_styles
                cell = ws._cells.get((r, c))
                if cell is not None and cell.has_style:
                    per_cell[(r - min_r, c - min_c)] = copy(cell.border)
        if per_cell:
            merged_nontopleft_borders[(min_r, min_c, max_r, max_c)] = per_cell

    ws.insert_rows(source_row + 1, count)

    # insert_rows() correctly shifts merges fully above or below the insertion
    # point.  The only merges we must fix manually are those that SPAN
    # source_row (min_row <= source_row < max_row): openpyxl auto-extends their
    # max_row by `count`, which is wrong.  We split them into a top portion
    # (min_row…source_row) and a bottom portion (source_row+count+1…max_row+count),
    # leaving the inserted rows unmerged.
    inserted_start = source_row + 1
    inserted_end = source_row + count

    # Process saved_merges in DESCENDING min_r order for fully-below merges.
    # When two same-column-span merges are adjacent (e.g. A26:N26 and A27:N34),
    # openpyxl silently rejects ws.merge_cells(A30:N30) because the stale A27:N34
    # range (rows 27-34) is still in the registry and row 30 falls inside it.
    # Processing the higher-row merge first removes that stale entry before we
    # try to register the shifted lower-row range.  Spanning merges (min_r <=
    # source_row) are unaffected by this ordering change.
    for min_r, min_c, max_r, max_c in sorted(saved_merges, key=lambda x: x[0], reverse=True):
        if min_r > source_row:
            # Merge fully BELOW — use the pre-saved value because the top-left
            # may now be a MergedCell ghost returning None.
            key = (min_r, min_c, max_r, max_c)
            saved_value = merged_cell_values.get(key)
            # openpyxl may keep the original stale range, shift it, or both.
            # Unmerge any range for these columns at either the original or
            # the shifted row positions to ensure a clean slate.
            for existing_m in list(ws.merged_cells.ranges):
                if (existing_m.min_col == min_c
                        and existing_m.max_col == max_c
                        and existing_m.min_row in (min_r, min_r + count)
                        and existing_m.max_row in (max_r, max_r + count)):
                    _safe_remove_merge(ws, existing_m)
            # After registry unmerge, clean up any MergedCell ghosts in ws._cells
            # so ws.cell() returns a fresh writable Cell instead of a proxy.
            ws._cells.pop((min_r + count, min_c), None)
            # Set value BEFORE re-merging (after merge, top-left becomes a
            # MergedCell with read-only .value)
            ws.cell(min_r + count, min_c).value = saved_value
            ws.merge_cells(
                start_row=min_r + count, start_column=min_c,
                end_row=max_r + count, end_column=max_c,
            )

            saved_style = merged_cell_styles.get((min_r, min_c, max_r, max_c))
            if saved_style:
                new_top_left = ws.cell(min_r + count, min_c)
                new_top_left.font = copy(saved_style["font"])
                new_top_left.border = copy(saved_style["border"])
                new_top_left.fill = copy(saved_style["fill"])
                new_top_left.number_format = saved_style["number_format"]
                new_top_left.protection = copy(saved_style["protection"])
                new_top_left.alignment = copy(saved_style["alignment"])
            # Restore per-cell border data for non-top-left cells.  After
            # ws.merge_cells() above these positions hold fresh MergedCell
            # proxies; writing .border to them preserves the edge-rendering
            # borders that were present in the original Excel-created file.
            non_tl = merged_nontopleft_borders.get((min_r, min_c, max_r, max_c))
            if non_tl:
                for (dr, dc), saved_border in non_tl.items():
                    target_r, target_c = min_r + count + dr, min_c + dc
                    mc = ws._cells.get((target_r, target_c))
                    if mc is not None:
                        mc.border = copy(saved_border)
            continue

        # Only spanning merges need manual correction.
        if not (min_r <= source_row < max_r):
            continue

        # Use pre-saved snapshot in case insert_rows() left MergedCell ghosts.
        key = (min_r, min_c, max_r, max_c)
        saved_value = merged_cell_values.get(key) or ws.cell(min_r, min_c).value

        # Unmerge the auto-extended range openpyxl created after insert_rows.
        for existing_m in list(ws.merged_cells.ranges):
            if (existing_m.min_row == min_r and existing_m.min_col == min_c
                    and existing_m.max_row == max_r + count
                    and existing_m.max_col == max_c):
                _safe_remove_merge(ws, existing_m)
                break

        # Top portion: min_r … source_row (only if multi-cell range)
        if source_row > min_r or max_c > min_c:
            ws.cell(min_r, min_c).value = saved_value
            ws.merge_cells(
                start_row=min_r, start_column=min_c,
                end_row=source_row, end_column=max_c,
            )
        # Bottom portion: rows originally below source_row, shifted by count.
        if max_r > source_row:
            bottom_start = source_row + count + 1
            bottom_end = max_r + count
            if bottom_end >= bottom_start:
                ws.merge_cells(
                    start_row=bottom_start, start_column=min_c,
                    end_row=bottom_end, end_column=max_c,
                )

    # Safety: strip any merge that still overlaps the inserted rows.
    for m in list(ws.merged_cells.ranges):
        if m.min_row <= inserted_end and m.max_row >= inserted_start:
            _safe_remove_merge(ws, m)

    # Purge MergedCell ghost objects from the inserted rows only.
    # Ghosts in source_row belong to legitimate top-portion merges and must
    # NOT be deleted.  Only the freshly inserted rows need clean Cell objects.
    for r in range(inserted_start, inserted_end + 1):
        for c in list(ws._cells):
            if isinstance(ws._cells.get(c), _MC) and c[0] == r:
                del ws._cells[c]

    # Restore source_row styles (may have been wiped by ghost purge) and
    # copy styles + values to each inserted row.
    for col in range(1, max_col + 1):
        src_cell = ws.cell(source_row, col)
        style = saved_styles.get(col)
        if style:
            src_cell.font = copy(style["font"])
            src_cell.border = copy(style["border"])
            src_cell.fill = copy(style["fill"])
            src_cell.number_format = style["number_format"]
            src_cell.protection = copy(style["protection"])
            src_cell.alignment = copy(style["alignment"])

    for offset in range(1, count + 1):
        for col in range(1, max_col + 1):
            val_src = ws.cell(source_row, col)
            dst = ws.cell(source_row + offset, col)
            dst.value = val_src.value
            style = insert_styles.get(col)
            if style:
                dst.font = copy(style["font"])
                dst.border = copy(style["border"])
                dst.fill = copy(style["fill"])
                dst.number_format = style["number_format"]
                dst.protection = copy(style["protection"])
                dst.alignment = copy(style["alignment"])


# ── 2. Table boundary scanning ─────────────────────────────────────────────

def _read_headers(ws: Worksheet, header_row: int, start_col: int) -> list[tuple[str, int]]:
    """Read non-empty header names rightward from *start_col*.

    Returns a list of ``(column_name, col_index)`` pairs.
    """
    headers: list[tuple[str, int]] = []
    col = start_col
    while True:
        val = _get_merged_cell_value(ws, header_row, col)
        if val is None or str(val).strip() == "":
            break
        headers.append((str(val), col))
        col += 1
    return headers


def _find_last_data_row(ws: Worksheet, start_row: int, join_col: int) -> int:
    """Return the last row in join_col with a non-empty value, starting from start_row.

    Stops before a multi-row vertical merge in the join column: such merges
    are section labels or footers, not individual data rows.  Scanning into
    the interior of a merge (MergedCell proxies) would inflate last_tmpl_row
    and cause row insertion to happen after the merge instead of before it.
    """
    last_row = start_row
    row = start_row
    while True:
        # MergedCell proxy → we are inside a multi-row merge that started
        # above; the table ended at the previous row.
        if isinstance(ws._cells.get((row, join_col)), _MC):
            break
        val = ws.cell(row, join_col).value
        if val is None or str(val).strip() == "":
            break
        # If this row is the TOP of ANY multi-row vertical merge (in any
        # column), treat it as a structural boundary (section label, footer,
        # etc.) and stop before including it.  Checking only the join column
        # was insufficient: a merge in a data column at the same row also
        # causes source_row == merge.min_row, triggering the spanning split
        # and losing the merge value / corrupting the merge structure.
        if any(m.min_row == row and m.max_row > row
               for m in ws.merged_cells.ranges):
            return last_row
        last_row = row
        row += 1
    return last_row


def _find_end_table_row(ws: Worksheet, start_row: int) -> _EndTableMarker | None:
    """Scan rows from *start_row* downward for a ``{{ end_table }}`` marker.

    Returns an :class:`_EndTableMarker` describing the marker cell, or ``None``
    if not found.  Stops at the first fully-blank row (all columns empty).
    """
    max_col = ws.max_column or 1
    row = start_row
    while True:
        all_empty = True
        for col in range(1, max_col + 1):
            val = ws.cell(row, col).value
            if val is not None and str(val).strip():
                all_empty = False
                if isinstance(val, str):
                    m = _TemplateRegex.END_TABLE.search(val)
                    if m:
                        insert_mode: InsertMode = "below"
                        meta = (m.group("meta") or "").strip()
                        if meta:
                            for part in meta.split(","):
                                k, _, v = part.partition("=")
                                if k.strip().lower() == "insert":
                                    insert_mode = cast(InsertMode, v.strip().lower())
                        return _EndTableMarker(row=row, col=col, insert_mode=insert_mode)
        if all_empty:
            break
        row += 1
    return None


def _find_insert_data_row(ws: Worksheet, start_row: int) -> tuple[int, int] | None:
    """Scan rows from *start_row* downward for a ``{{ insert_data }}`` marker.

    Returns ``(row, col)`` of the marker cell, or ``None`` if not found.
    Stops at the first fully-blank row.
    """
    max_col = ws.max_column or 1
    row = start_row
    while True:
        all_empty = True
        for col in range(1, max_col + 1):
            val = ws.cell(row, col).value
            if val is not None and str(val).strip():
                all_empty = False
                if isinstance(val, str) and _TemplateRegex.INSERT_DATA.search(val):
                    return (row, col)
        if all_empty:
            break
        row += 1
    return None


def _compute_table_region(
    ws: Worksheet, mc: MarkedCell, df: pl.DataFrame
) -> tuple[int, int, int, int]:
    """Return (min_row, min_col, max_row, max_col) of the region a table() tag would write.

    Includes all template rows plus any potential outer-join extra rows.
    Column span is join_col through the last header column.
    For ``positional=True`` tables, the region is tag_cell to tag_cell + df shape.
    """
    tm = TableMeta.from_cell(mc)
    if tm.positional:
        start_row, start_col = coordinate_to_tuple(mc.cell_addr)
        return start_row, start_col, start_row + df.height - 1, start_col + df.width - 1
    tag_row, tag_col = coordinate_to_tuple(mc.cell_addr)
    join_col = tag_col - 1

    headers = _read_headers(ws, tag_row - 1, tag_col)
    last_col = headers[-1][1] if headers else tag_col
    last_row = _find_last_data_row(ws, tag_row, join_col)

    # Collision detection uses the template-only boundary, not a post-insertion
    # projection.  Vertically stacked tables are processed bottom-up and
    # insert_rows() shifts the lower table down safely — inflating last_row
    # here caused false-positive collisions when an outer/right table had a
    # large DataFrame sitting above a neighbour table.

    return tag_row - 1, join_col, last_row, last_col  # include header row


def _check_region_collisions(
    regions: list[tuple[str, tuple[int, int, int, int]]]
) -> None:
    """Raise ValueError if any two (sheet, region) pairs overlap.

    Each entry is (description, (min_row, min_col, max_row, max_col)).
    All entries are assumed to be on the same worksheet.
    """
    for i in range(len(regions)):
        for j in range(i + 1, len(regions)):
            desc_a, (r1a, c1a, r2a, c2a) = regions[i]
            desc_b, (r1b, c1b, r2b, c2b) = regions[j]
            row_overlap = r1a <= r2b and r1b <= r2a
            col_overlap = c1a <= c2b and c1b <= c2a
            if row_overlap and col_overlap:
                raise ValueError(
                    f"Fill collision on sheet: '{desc_a}' "
                    f"(rows {r1a}-{r2a}, cols {c1a}-{c2a}) overlaps "
                    f"'{desc_b}' (rows {r1b}-{r2b}, cols {c1b}-{c2b})"
                )


# ── 3. Table fill algorithms ───────────────────────────────────────────────

def _fill_positional(ws: Worksheet, mc: MarkedCell, df: pl.DataFrame) -> None:
    """Fill a DataFrame positionally starting at the tag cell.

    The tag cell is the top-left corner of the written region.
    Columns are written left-to-right in DataFrame column order.
    Rows are written top-to-bottom in DataFrame row order.
    No join column, no header matching.  The tag cell value is cleared
    and then overwritten with df[0, 0].

    Raises:
        ValueError: If the DataFrame has zero rows or zero columns.
    """
    if df.height == 0 or df.width == 0:
        raise ValueError(
            f"positional table(positional=True) for '{mc.name}' requires a non-empty DataFrame "
            f"(got {df.height} rows × {df.width} cols)"
        )

    start_row, start_col = coordinate_to_tuple(mc.cell_addr)
    ws.cell(start_row, start_col).value = None  # clear the tag

    columns = df.columns
    for r_offset, row in enumerate(df.iter_rows(named=True)):
        for c_offset, col_name in enumerate(columns):
            ws.cell(start_row + r_offset, start_col + c_offset).value = row[col_name]


def _sorted_outer_fill(
    ws: Worksheet,
    df: pl.DataFrame,
    tag_row: int,
    join_tmpl_col: int,
    join_df_col: str,
    headers: list[tuple[str, int]],
    saved_join: dict[int, Any],
    last_tmpl_row: int,
    insert_before_row: int | None,
    order_by: OrderBy,
    fill: FillSpec | None = None,
) -> int:
    """Write a sorted outer join into the upper zone of a table region.

    All rows from *tag_row* up to (but not including) *insert_before_row* are
    written top-down in the order given by *order_by*.  Template rows in the
    upper zone whose join key is absent from the DataFrame are preserved in the
    sorted output with null data-column values (template-only rows).  Rows at
    or below *insert_before_row* (the lower zone) keep their template positions
    and are filled by key-match.

    Returns the number of rows inserted into the worksheet.
    """
    sort_col = order_by.col if order_by.col is not None else join_df_col
    sort_asc = order_by.ascending

    # Lower zone: join values of all rows at/below insert_before_row.
    # These are key-matched separately and excluded from the sorted upper zone.
    lower_vals: set[Any] = set()
    if insert_before_row is not None:
        for r in range(insert_before_row, last_tmpl_row + 1):
            lower_vals.add(saved_join.get(r))

    # Upper zone boundary before any insertion.
    upper_last_row = (
        insert_before_row - 1 if insert_before_row is not None else last_tmpl_row
    )

    # All data rows for the upper zone (not reserved for the lower zone).
    upper_items: list[dict[str, Any]] = [
        row
        for row in df.iter_rows(named=True)
        if row[join_df_col] not in lower_vals
    ]

    # Template-only upper zone rows: join key present in the template but absent
    # from the DataFrame.  Preserved in the sort output with null data columns.
    df_join_vals: set[Any] = set(df[join_df_col].to_list())
    for r in range(tag_row, upper_last_row + 1):
        v = saved_join.get(r)
        if v is not None and v not in df_join_vals and v not in lower_vals:
            tmpl_row: dict[str, Any] = {join_df_col: v}
            for col_name, _ in headers:
                tmpl_row[col_name] = None
            upper_items.append(tmpl_row)

    # Sort combined upper zone: non-null sort-key values first (asc or desc),
    # null sort-key values always last.
    non_null = [r for r in upper_items if r.get(sort_col) is not None]
    null_rows = [r for r in upper_items if r.get(sort_col) is None]
    non_null.sort(key=lambda r: r[sort_col], reverse=not sort_asc)
    upper_items = non_null + null_rows

    n_upper_tmpl = upper_last_row - tag_row + 1
    n_upper_items = len(upper_items)
    rows_inserted = 0

    # Insert rows if the sorted upper zone has more items than available slots.
    if n_upper_items > n_upper_tmpl:
        rows_inserted = n_upper_items - n_upper_tmpl
        _copy_row_styles(ws, upper_last_row, rows_inserted)
        upper_last_row += rows_inserted
        if insert_before_row is not None:
            insert_before_row += rows_inserted
            last_tmpl_row += rows_inserted
        else:
            last_tmpl_row += rows_inserted

    # Write the sorted upper zone rows top-down.
    for i, row in enumerate(upper_items):
        ws_row = tag_row + i
        ws.cell(ws_row, join_tmpl_col).value = row[join_df_col]
        for col_name, col_idx in headers:
            ws.cell(ws_row, col_idx).value = _apply_fill(row.get(col_name), col_name, fill)

    # Clear leftover upper zone slots when fewer items than slots.
    for r in range(tag_row + n_upper_items, upper_last_row + 1):
        ws.cell(r, join_tmpl_col).value = None
        for _, col_idx in headers:
            ws.cell(r, col_idx).value = None

    # Write lower zone rows by key-match (template order preserved).
    if insert_before_row is not None:
        orig_insert = insert_before_row - rows_inserted
        df_lookup: dict[Any, dict[str, Any]] = {
            row[join_df_col]: row for row in df.iter_rows(named=True)
        }
        for orig_r, tmpl_val in sorted(saved_join.items()):
            if orig_r < orig_insert:
                continue  # upper zone row — already written
            new_r = orig_r + rows_inserted
            if tmpl_val in df_lookup:
                row = df_lookup[tmpl_val]
                for col_name, col_idx in headers:
                    ws.cell(new_r, col_idx).value = _apply_fill(row.get(col_name), col_name, fill)
            elif fill is not None:
                # Unmatched lower zone row — apply fill to existing cell values.
                for col_name, col_idx in headers:
                    current = ws.cell(new_r, col_idx).value
                    ws.cell(new_r, col_idx).value = _apply_fill(current, col_name, fill)

    return rows_inserted


def _fill_table(ws: Worksheet, mc: MarkedCell, df: pl.DataFrame) -> None:
    """Fill *ws* with *df* data using the join semantics declared in *mc*.

    Join modes:

    * ``left``  — fill matched rows; leave unmatched template rows as-is.  No inserts.
    * ``inner`` — fill matched rows; clear data columns on unmatched rows.  No inserts.
    * ``outer`` — fill matched rows; append unmatched DataFrame rows (pushes content down).
      With ``order_by``: uses :func:`_sorted_outer_fill` to sort the upper zone.
    * ``right`` — overwrite rows top-down in DataFrame order; insert if DataFrame is longer;
      clear remaining template rows if DataFrame is shorter.

    Structural markers (optional):

    * ``{{ end_table }}`` on its own row:  table ends at the row above; marker row deleted.
    * ``{{ end_table }}`` on a data row:   marker cell cleared; that row is the last data row.
    * ``{{ end_table | insert=above }}`` on a data row:  as above, but extras inserted before it.
    * ``{{ insert_data }}`` — marks the insertion point for outer-join extras; takes precedence
      over ``{{ end_table }}`` for insertion placement.
    """
    tm = TableMeta.from_cell(mc)
    tag_row, tag_col = coordinate_to_tuple(mc.cell_addr)
    header_row = tag_row - 1
    join_tmpl_col = tag_col - 1

    tmpl_join_header = _get_merged_cell_value(ws, header_row, join_tmpl_col)
    join_df_col: str = tm.on if tm.on is not None else str(tmpl_join_header)

    headers = _read_headers(ws, header_row, tag_col)
    # Handle {{ insert_data }} marker — delete its row early so the boundary
    # scans below see a contiguous table.  The deleted row's position becomes
    # the insertion point for extra rows (outer join).
    insert_data_marker = _find_insert_data_row(ws, tag_row)
    insert_data_before: int | None = None
    if insert_data_marker is not None:
        id_row, _id_col = insert_data_marker
        ws.delete_rows(id_row)
        # openpyxl shifts cell data but not the merge registry on delete_rows.
        # Correct the registry so _copy_row_styles sees accurate positions.
        _sync_merges_after_delete(ws, id_row)
        # After deletion, the row that was below now sits at id_row.
        # Extra rows should be inserted above that row.
        insert_data_before = id_row

    # Determine table boundary.  An explicit {{ end_table }} marker takes
    # precedence over the heuristic scan in _find_last_data_row.
    #
    # Three modes:
    #   Option A: {{ end_table }} on its own row (no join value) — table ends
    #             at the row above; marker row is deleted after writes.
    #   Option B: {{ end_table }} on a data row — marker cell cleared; that
    #             row is the last table row.  Extras go after it.
    #   Option C: {{ end_table | insert=above }} on a data row — marker cell
    #             cleared; that row is part of the table AND gets matched,
    #             but extras are inserted BEFORE it (insert_before_row).
    end_marker = _find_end_table_row(ws, tag_row)
    delete_end_row: int | None = None
    insert_before_row: int | None = None  # Option C: extras go before this row

    if end_marker is not None:
        end_join_val = _get_merged_cell_value(ws, end_marker.row, join_tmpl_col)
        if end_join_val is not None and str(end_join_val).strip():
            # Data row with {{ end_table }}.  Clear the marker cell.
            ws.cell(end_marker.row, end_marker.col).value = None
            if end_marker.insert_mode == "above":
                # Option C: match all rows through end_row, but insert
                # extras above end_row (between the data rows and this row).
                last_tmpl_row = end_marker.row
                insert_before_row = end_marker.row
            else:
                # Option B: extras go after last_tmpl_row (default).
                last_tmpl_row = end_marker.row
        else:
            # Option A: {{ end_table }} on its own row.
            last_tmpl_row = end_marker.row - 1
            delete_end_row = end_marker.row
    else:
        last_tmpl_row = _find_last_data_row(ws, tag_row, join_tmpl_col)

    # {{ insert_data }} takes precedence over {{ end_table }} for the
    # insertion point.  The table boundary (last_tmpl_row) is unaffected —
    # matching continues through all rows.
    if insert_data_before is not None:
        insert_before_row = insert_data_before

    n_tmpl = last_tmpl_row - tag_row + 1
    rows_inserted = 0

    # Clear the tag placeholder before any data writes so the data loop
    # can overwrite (tag_row, tag_col) with the correct value.
    ws.cell(tag_row, tag_col).value = None

    if tm.join == "right":
        df_list = df.to_dicts()
        n_df = len(df_list)
        if n_df > n_tmpl:
            rows_inserted = n_df - n_tmpl
            style_src = tag_row if tm.style == "first" else last_tmpl_row
            _copy_row_styles(ws, style_src, rows_inserted)
        for i, row in enumerate(df_list):
            ws_row = tag_row + i
            ws.cell(ws_row, join_tmpl_col).value = row.get(join_df_col)
            for col_name, col_idx in headers:
                ws.cell(ws_row, col_idx).value = _apply_fill(row.get(col_name), col_name, tm.fill)
        # Clear extra template rows if DataFrame has fewer rows.
        for r in range(tag_row + n_df, last_tmpl_row + 1):
            ws.cell(r, join_tmpl_col).value = None
            for _, col_idx in headers:
                ws.cell(r, col_idx).value = None
    else:
        df_lookup: dict[Any, dict[str, Any]] = {}
        for row in df.iter_rows(named=True):
            df_lookup[row[join_df_col]] = row

        # Snapshot per-row join values BEFORE any row insertion.  _copy_row_styles
        # may purge MergedCell ghosts on source_row, wiping the join column
        # value and causing the matched-row lookup to miss that row entirely.
        saved_join: dict[int, Any] = {
            r: _get_merged_cell_value(ws, r, join_tmpl_col)
            for r in range(tag_row, last_tmpl_row + 1)
        }

        # Sorted outer join: write all upper-zone rows in declared sort order.
        # This path replaces the standard outer join logic below.
        if tm.join == "outer" and tm.order_by is not None:
            rows_inserted = _sorted_outer_fill(
                ws, df, tag_row, join_tmpl_col, join_df_col,
                headers, saved_join, last_tmpl_row, insert_before_row,
                tm.order_by, tm.fill,
            )
        else:
            # outer: insert extra rows BEFORE writing any data.  insert_rows()
            # can create phantom MergedCell proxies on source_row, silently
            # destroying values already written there.  By inserting first and
            # writing matched rows afterwards, all cells are real Cell objects.
            extra: list[dict[str, Any]] = []
            if tm.join == "outer":
                tmpl_vals = set(saved_join.values())
                extra = [
                    row
                    for row in df.iter_rows(named=True)
                    if row[join_df_col] not in tmpl_vals
                ]
                if extra:
                    rows_inserted = len(extra)
                    if insert_before_row is not None:
                        # Option C: insert extras BEFORE the insert_before_row.
                        # Always insert just above the marker row; style=first copies
                        # styles from tag_row instead of from the row immediately
                        # above the marker.
                        insert_after = insert_before_row - 1
                        if insert_after < tag_row:
                            insert_after = tag_row
                        if tm.style == "first":
                            _copy_row_styles(ws, insert_after, rows_inserted, style_from=tag_row)
                        else:
                            _copy_row_styles(ws, insert_after, rows_inserted)
                        # Rows were inserted after insert_after, so
                        # insert_before_row (and everything below) shifted down.
                        insert_before_row += rows_inserted
                        last_tmpl_row += rows_inserted
                    else:
                        # Always insert after last_tmpl_row; style=first copies styles
                        # from tag_row instead of from last_tmpl_row.
                        if tm.style == "first":
                            _copy_row_styles(ws, last_tmpl_row, rows_inserted, style_from=tag_row)
                        else:
                            _copy_row_styles(ws, last_tmpl_row, rows_inserted)

            # After row insertion, saved_join keys may no longer match worksheet
            # row numbers.  Rebuild the mapping: original template rows that were
            # at or above the insertion point keep their row number; rows at or
            # below insert_before_row (Option C) have shifted down by rows_inserted.
            if rows_inserted and insert_before_row is not None:
                shifted_join: dict[int, Any] = {}
                orig_insert_row = insert_before_row - rows_inserted
                for orig_r, val in saved_join.items():
                    if orig_r >= orig_insert_row:
                        shifted_join[orig_r + rows_inserted] = val
                    else:
                        shifted_join[orig_r] = val
                saved_join = shifted_join

            # Fill / clear matched template rows (safe now — any row insertion
            # and ghost purge has already happened above).  Use saved_join
            # instead of re-reading from the worksheet, because the ghost purge
            # inside _copy_row_styles may have wiped the join column cell.
            for r in sorted(saved_join):
                tmpl_val = saved_join[r]
                if tmpl_val in df_lookup:
                    row = df_lookup[tmpl_val]
                    for col_name, col_idx in headers:
                        ws.cell(r, col_idx).value = _apply_fill(row.get(col_name), col_name, tm.fill)
                elif tm.join == "inner":
                    for _, col_idx in headers:
                        ws.cell(r, col_idx).value = None

            # Apply fill to unmatched template rows in left/outer joins.
            if tm.fill is not None and tm.join in ("left", "outer"):
                for r in sorted(saved_join):
                    tmpl_val = saved_join[r]
                    if tmpl_val not in df_lookup:
                        for col_name, col_idx in headers:
                            current = ws.cell(r, col_idx).value
                            ws.cell(r, col_idx).value = _apply_fill(current, col_name, tm.fill)

            if extra:
                if insert_before_row is not None:
                    # Option C: extras occupy the rows just above insert_before_row.
                    extra_start = insert_before_row - rows_inserted
                else:
                    extra_start = last_tmpl_row + 1
                for i, row in enumerate(extra):
                    ws_row = extra_start + i
                    ws.cell(ws_row, join_tmpl_col).value = row[join_df_col]
                    for col_name, col_idx in headers:
                        ws.cell(ws_row, col_idx).value = _apply_fill(row.get(col_name), col_name, tm.fill)

    # Delete the {{ end_table }} marker row (Option A).  Row indices may have
    # shifted if extra rows were inserted above, so adjust accordingly.
    if delete_end_row is not None:
        ws.delete_rows(delete_end_row + rows_inserted)
        _sync_merges_after_delete(ws, delete_end_row + rows_inserted)

    # Delete placeholder row (outer join, non-sorted only).  The tag row is used
    # as the style source for inserted rows, then removed if its join value is
    # absent from the DataFrame.  Deletion is done last so the row is available
    # as a style source throughout, and after delete_end_row so that the end_table
    # row (which is always below the tag row) is gone first — keeping index stable.
    if tm.placeholder and tm.join == "outer" and tm.order_by is None:
        tag_join_val = saved_join.get(tag_row)
        if tag_join_val not in df_lookup:
            ws.delete_rows(tag_row)
            _sync_merges_after_delete(ws, tag_row)



