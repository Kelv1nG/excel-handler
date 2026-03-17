from openpyxl.utils import coordinate_to_tuple, range_boundaries
import polars as pl
from typing import Optional, Tuple, List, Any
from excel.exceptions import ExcelDataExtractorError, TableNotFoundError
from excel.utils import load_excel_workbook
from openpyxl.workbook.workbook import Workbook


class ExcelDataExtractor:
    """Universal Excel data extraction utility."""

    def __init__(self, filepath: str):
        """Initialize extractor for an Excel file."""
        self.filepath = filepath
        self._wb: Optional[Workbook] = None

    @property
    def wb(self) -> Workbook:
        """
        Get loaded workbook.

        Returns:
            Workbook object

        Raises:
            ExcelDataExtractorError: If workbook not loaded
        """
        if self._wb is None:
            raise ExcelDataExtractorError(
                "Workbook not loaded. Use ExcelDataExtractor as context manager:\n"
                "    with ExcelDataExtractor('file.xlsx') as extractor:\n"
                "        df = extractor.extract_by_columns([...])"
            )
        return self._wb

    def __enter__(self):
        """Context manager entry - opens workbook."""
        self._wb = load_excel_workbook(self.filepath, data_only=True)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager exit - closes workbook."""
        if self._wb:
            self._wb.close()

    def _get_sheet(self, sheet_name: Optional[str] = None):
        """
        Get sheet, with validation.

        Args:
            sheet_name: Sheet name (if None, uses active sheet)

        Returns:
            Worksheet object

        Raises:
            ExcelDataExtractorError: If sheet not found
        """
        # self.wb property ensures non-None
        if sheet_name is None:
            return self.wb.active  # ✅ Type checker happy

        if sheet_name not in self.wb.sheetnames:  # ✅ Type checker happy
            raise ExcelDataExtractorError(
                f"Sheet '{sheet_name}' not found in {self.filepath}. "
                f"Available sheets: {self.wb.sheetnames}"
            )

        return self.wb[
            sheet_name
        ]  # ==================== Table Extraction (PUBLIC API) ====================

    def extract_table_by_columns(
        self,
        required_columns: List[str],
        sheet: Optional[str] = None,
        search_all_sheets: bool = False,
        unmerge_cells: bool = True,
        exclude_summary: bool = True,
        fill_forward: bool = True,
        max_header_search_rows: int = 100,
    ) -> pl.DataFrame:
        """
        Find and extract table by searching for column headers.

        Args:
            required_columns: Column names that must be present
            sheet: Sheet name to search (if None, searches active sheet or all sheets)
            search_all_sheets: If True, searches all sheets for the table
            unmerge_cells: Unmerge and fill merged cells
            exclude_summary: Exclude summary rows
            fill_forward: Fill forward for hierarchical data
            max_header_search_rows: Max rows to search for headers

        Returns:
            Polars DataFrame

        Raises:
            TableNotFoundError: If table not found

        Examples:
            >>> # Search in active sheet
            >>> df = extractor.extract_by_columns(['Company', 'Amount'])

            >>> # Search in specific sheet
            >>> df = extractor.extract_by_columns(['Company', 'Amount'], sheet='Data')

            >>> # Search all sheets
            >>> df = extractor.extract_by_columns(
            ...     ['Company', 'Amount'],
            ...     search_all_sheets=True
            ... )
        """
        if search_all_sheets:
            # Search all sheets until found
            for sheet_name in self.wb.sheetnames:
                try:
                    return self._extract_by_columns_from_sheet(
                        sheet_name,
                        required_columns,
                        unmerge_cells,
                        exclude_summary,
                        fill_forward,
                        max_header_search_rows,
                    )
                except TableNotFoundError:
                    # Not in this sheet, continue to next
                    continue

            # Not found in any sheet
            raise TableNotFoundError(
                f"Could not find table with columns {required_columns} "
                f"in any sheet of {self.filepath}"
            )
        else:
            # Search specific sheet (or active)
            sheet_obj = self._get_sheet(sheet)
            return self._extract_by_columns_from_sheet(
                sheet_obj.title,
                required_columns,
                unmerge_cells,
                exclude_summary,
                fill_forward,
                max_header_search_rows,
            )

    def _extract_table_by_columns_from_sheet(
        self,
        sheet_name: str,
        required_columns: List[str],
        unmerge_cells: bool,
        exclude_summary: bool,
        fill_forward: bool,
        max_header_search_rows: int,
    ) -> pl.DataFrame:
        """
        Internal method to extract from a specific sheet.

        Args:
            sheet_name: Name of sheet to extract from
            required_columns: Columns to find
            unmerge_cells: Unmerge cells
            exclude_summary: Exclude summary
            fill_forward: Fill forward
            max_header_search_rows: Max rows to search

        Returns:
            Polars DataFrame

        Raises:
            TableNotFoundError: If table not found in this sheet
        """
        sheet = self._get_sheet(sheet_name)

        # Step 1: Unmerge cells if requested
        if unmerge_cells:
            self._unmerge_and_fill_sheet(sheet)

        # Step 2: Find header row
        header_row = self._find_header_row_in_sheet(
            sheet, required_columns, max_header_search_rows
        )

        if header_row is None:
            raise TableNotFoundError(
                f"Could not find table with columns: {required_columns} "
                f"in sheet '{sheet_name}' (searched first {max_header_search_rows} rows)"
            )

        # Step 3: Detect table boundaries from header row
        # Start from column 1 (A) by default
        start_col = 1
        boundaries = self._detect_boundaries_in_sheet(
            sheet, header_row, start_col, has_headers=True
        )

        # Step 4: Extract the range
        df = self._extract_range_from_sheet(
            sheet,
            *boundaries,
            has_headers=True,
            column_names=None,
            exclude_summary=exclude_summary,
        )

        # Step 5: Fill forward if requested
        if fill_forward and len(df) > 0:
            df = df.with_columns([pl.col(col).forward_fill() for col in df.columns])

        return df

    def extract_table_by_range(
        self,
        range_str: str,
        sheet: str,
        has_headers: bool = True,
        column_names: Optional[List[str]] = None,
        unmerge_cells: bool = True,
        exclude_summary: bool = True,
        fill_forward: bool = True,
    ) -> pl.DataFrame:
        """
        Extract table from explicit range.

        Args:
            range_str: Excel range (e.g., "A5:C20")
            sheet: Sheet name (REQUIRED for range extraction)
            has_headers: First row contains headers
            column_names: Manual column names if has_headers=False
            unmerge_cells: Unmerge cells
            exclude_summary: Exclude summary rows
            fill_forward: Fill forward

        Returns:
            Polars DataFrame

        Example:
            >>> df = extractor.extract_by_range('A5:C20', sheet='Data')
            >>> df = extractor.extract_by_range(
            ...     'B10:D30',
            ...     sheet='Sheet1',
            ...     has_headers=False,
            ...     column_names=['ID', 'Name', 'Value']
            ... )
        """
        # Get sheet
        sheet_obj = self._get_sheet(sheet)

        # Step 1: Unmerge cells if requested
        if unmerge_cells:
            self._unmerge_and_fill_sheet(sheet_obj)

        # Step 2: Parse range
        min_col, min_row, max_col, max_row = range_boundaries(range_str)

        # assert values are not None
        assert min_row is not None
        assert min_col is not None
        assert max_row is not None
        assert max_col is not None

        # Step 3: Extract
        df = self._extract_range_from_sheet(
            sheet_obj,
            min_row,
            min_col,
            max_row,
            max_col,
            has_headers=has_headers,
            column_names=column_names,
            exclude_summary=exclude_summary,
        )

        # Step 4: Fill forward
        if fill_forward and len(df) > 0:
            df = df.with_columns([pl.col(col).forward_fill() for col in df.columns])

        return df

    def extract_from_cell(
        self,
        start_cell: str,
        sheet: str,
        has_headers: bool = True,
        column_names: Optional[List[str]] = None,
        unmerge_cells: bool = True,
        exclude_summary: bool = True,
        fill_forward: bool = True,
        max_empty_rows: int = 2,
    ) -> pl.DataFrame:
        """
        Extract table starting from cell, auto-detecting boundaries.

        Args:
            start_cell: Starting cell (e.g., "A5")
            sheet: Sheet name (REQUIRED)
            has_headers: First row is headers
            column_names: Manual column names
            unmerge_cells: Unmerge cells
            exclude_summary: Exclude summary
            fill_forward: Fill forward
            max_empty_rows: Stop threshold

        Returns:
            Polars DataFrame

        Example:
            >>> df = extractor.extract_from_cell('A5', sheet='Data')
        """
        # Get sheet
        sheet_obj = self._get_sheet(sheet)

        # Step 1: Unmerge cells if requested
        if unmerge_cells:
            self._unmerge_and_fill_sheet(sheet_obj)

        # Step 2: Parse start cell
        start_row, start_col = coordinate_to_tuple(start_cell)

        # Step 3: Detect boundaries
        boundaries = self._detect_boundaries_in_sheet(
            sheet_obj, start_row, start_col, has_headers, max_empty_rows
        )

        # Step 4: Extract
        df = self._extract_range_from_sheet(
            sheet_obj,
            *boundaries,
            has_headers=has_headers,
            column_names=column_names,
            exclude_summary=exclude_summary,
        )

        # Step 5: Fill forward
        if fill_forward and len(df) > 0:
            df = df.with_columns([pl.col(col).forward_fill() for col in df.columns])

        return df

    # ==================== Utility Methods (INTERNAL) ====================

    def _unmerge_and_fill_sheet(self, sheet):
        """
        Unmerge all merged cells in the sheet and fill values forward.

        Args:
            sheet: Worksheet object
        """
        merged_ranges = list(sheet.merged_cells.ranges)

        for merged_range in merged_ranges:
            # Get bounds of merged range
            min_col, min_row, max_col, max_row = merged_range.bounds

            # Get value from top-left cell (where value is stored)
            value = sheet.cell(min_row, min_col).value

            # Unmerge
            sheet.unmerge_cells(str(merged_range))

            # Fill value to all cells in the range
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    sheet.cell(row, col).value = value

    def _detect_boundaries_in_sheet(
        self,
        sheet,
        start_row: int,
        start_col: int,
        has_headers: bool = True,
        max_empty_rows: int = 2,
    ) -> Tuple[int, int, int, int]:
        """
        Auto-detect table boundaries from a starting point.

        Args:
            sheet: Worksheet object
            start_row: Starting row number (1-indexed)
            start_col: Starting column number (1-indexed)
            has_headers: Whether first row is headers
            max_empty_rows: Stop after this many empty rows

        Returns:
            Tuple of (min_row, min_col, max_row, max_col)
        """
        # Detect right boundary (columns)
        max_col = start_col
        for col in range(start_col, min(start_col + 100, sheet.max_column + 1)):
            # Check if this column has data in header row
            cell_value = sheet.cell(start_row, col).value

            if cell_value is None or cell_value == "":
                # Empty column - stop here
                break

            max_col = col

        # Detect bottom boundary (rows)
        data_start_row = start_row + 1 if has_headers else start_row
        max_row = start_row
        consecutive_empty = 0

        for row in range(data_start_row, sheet.max_row + 1):
            # Check if row has any data in our column range
            has_data = any(
                sheet.cell(row, col).value is not None
                and sheet.cell(row, col).value != ""
                for col in range(start_col, max_col + 1)
            )

            if has_data:
                max_row = row
                consecutive_empty = 0
            else:
                consecutive_empty += 1
                if consecutive_empty >= max_empty_rows:
                    # Too many empty rows, stop
                    break

        return (start_row, start_col, max_row, max_col)

    @staticmethod
    def _is_summary_row(row_values: List[Any]) -> bool:
        """
        Check if row contains summary keywords.

        Args:
            row_values: List of cell values in the row

        Returns:
            True if row appears to be a summary/total row
        """
        # Convert all values to lowercase string
        text = " ".join(str(v).lower() for v in row_values if v is not None)

        # Check for summary keywords
        keywords = ["total", "sum", "subtotal", "grand total", "summary"]

        return any(keyword in text for keyword in keywords)

    def _extract_range_from_sheet(
        self,
        sheet,
        min_row: int,
        min_col: int,
        max_row: int,
        max_col: int,
        has_headers: bool = True,
        column_names: Optional[List[str]] = None,
        exclude_summary: bool = False,
    ) -> pl.DataFrame:
        """
        Extract data from specific cell range.

        Args:
            sheet: Worksheet object
            min_row, min_col: Top-left corner (1-indexed)
            max_row, max_col: Bottom-right corner (1-indexed)
            has_headers: Whether first row contains headers
            column_names: Manual column names if has_headers=False
            exclude_summary: Skip summary rows

        Returns:
            Polars DataFrame
        """
        # Step 1: Get headers
        if has_headers:
            headers = [
                sheet.cell(min_row, col).value for col in range(min_col, max_col + 1)
            ]
            data_start = min_row + 1
        else:
            # Use provided column names or generate default
            if column_names:
                headers = column_names
            else:
                headers = [f"col_{i}" for i in range(max_col - min_col + 1)]
            data_start = min_row

        # Step 2: Get data rows
        data = []
        for row in range(data_start, max_row + 1):
            row_data = [
                sheet.cell(row, col).value for col in range(min_col, max_col + 1)
            ]

            # Skip summary rows if requested
            if exclude_summary and self._is_summary_row(row_data):
                continue

            data.append(row_data)

        # Step 3: Create DataFrame
        if not data:
            # Empty table - return empty DataFrame with schema
            return pl.DataFrame(schema={h: pl.Utf8 for h in headers})

        return pl.DataFrame(data, schema=headers, orient="row")

    def _find_header_row_in_sheet(
        self, sheet, required_columns: List[str], max_search_rows: int = 100
    ) -> Optional[int]:
        """
        Find row containing all required column headers.

        Args:
            sheet: Worksheet object
            required_columns: List of column names to find
            max_search_rows: Maximum rows to search

        Returns:
            Row number (1-indexed) or None if not found
        """
        search_limit = min(max_search_rows, sheet.max_row)

        for row_idx in range(1, search_limit + 1):
            # Get all non-empty values in this row
            row_values = [
                cell.value for cell in sheet[row_idx] if cell.value is not None
            ]

            # Check if all required columns are present
            if all(col in row_values for col in required_columns):
                return row_idx

        return None
