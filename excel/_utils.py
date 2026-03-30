from io import BytesIO
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from pathlib import Path
from typing import Union
import os

from excel.exceptions import (
    ExcelCorruptedError,
    ExcelError,
    ExcelFileNotFoundError,
    ExcelPermissionError,
)


def load_excel_workbook(
    filepath: Union[str, os.PathLike, bytes], read_only: bool = False, data_only: bool = False
) -> Workbook:
    """
    Load Excel workbook with error handling.

    Args:
        filepath: Path to Excel file or raw bytes (e.g., from an email attachment)
        read_only: Open in read-only mode (faster, less memory)
        data_only: Read cell values only, not formulas

    Returns:
        Openpyxl Workbook object

    Raises:
        ExcelFileNotFoundError: If file doesn't exist
        ExcelPermissionError: If no permission to read
        ExcelCorruptedError: If file is invalid/corrupted

    Example:
        >>> wb = load_excel_workbook('data.xlsx', data_only=True)
        >>> sheet = wb['Sheet1']
        >>> wb.close()
    """
    if isinstance(filepath, bytes):
        try:
            return load_workbook(BytesIO(filepath), read_only=read_only, data_only=data_only)
        except Exception as e:
            raise ExcelCorruptedError("Invalid or corrupted Excel bytes") from e

    filepath = Path(filepath)

    try:
        wb = load_workbook(filepath, read_only=read_only, data_only=data_only)
        return wb

    except FileNotFoundError as e:
        raise ExcelFileNotFoundError(f"File not found: {filepath}") from e

    except PermissionError as e:
        raise ExcelPermissionError(f"No permission to read: {filepath}") from e

    except Exception as e:
        raise ExcelCorruptedError(f"Invalid or corrupted Excel file: {filepath}") from e


def get_sheet_names(filepath: Union[str, os.PathLike]) -> list[str]:
    """
    Get list of sheet names without fully loading workbook.

    Args:
        filepath: Path to Excel file

    Returns:
        List of sheet names

    Example:
        >>> sheets = get_sheet_names('report.xlsx')
        >>> print(sheets)  # ['Sheet1', 'Data', 'Summary']
    """
    wb = load_excel_workbook(filepath, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()
    return sheet_names


def sheet_exists(filepath: Union[str, os.PathLike], sheet_name: str) -> bool:
    """
    Check if sheet exists in workbook.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of sheet to check

    Returns:
        True if sheet exists, False otherwise
    """
    return sheet_name in get_sheet_names(filepath)


def validate_excel_file(filepath: Union[str, os.PathLike]) -> bool:
    """
    Validate that file is a readable Excel file.

    Args:
        filepath: Path to Excel file

    Returns:
        True if valid, False otherwise
    """
    try:
        wb = load_excel_workbook(filepath, read_only=True)
        wb.close()
        return True
    except ExcelError:
        return False
